import pandas as pd
from datetime import datetime
from .database import upsert_urun, upsert_firma_stok, get_client, upsert_yoldaki_urun, upsert_g5f_stok, depo_kanonik
from shared.utils import normalize_tr
import math

def safe_float(val, default=0.0):
    """NaN ve None değerlerini güvenli şekilde float'a çevirir."""
    try:
        v = float(val or default)
        return default if (math.isnan(v) or math.isinf(v)) else v
    except:
        return default

def safe_int(val, default=0):
    """NaN ve None değerlerini güvenli şekilde int'e çevirir."""
    try:
        v = float(val or default)
        if math.isnan(v) or math.isinf(v):
            return default
        return int(v)
    except:
        return default

def safe_str(val, default=""):
    """NaN ve None değerlerini güvenli şekilde str'ye çevirir."""
    if val is None:
        return default
    s = str(val).strip()
    return default if s.lower() in ("nan", "none", "nat") else s


def tr_upper(s):
    """Türkçe karakterleri de doğru büyüten upper fonksiyonu"""
    return str(s).strip().upper().replace("İ","I").replace("Ğ","G").replace("Ü","U").replace("Ş","S").replace("Ç","C").replace("Ö","O")

def normalize_sku(sku):
    """Fazeon/FAZEON gibi marka prefix'lerini SKU'dan temizler ve büyük harfe çevirir."""
    sku = str(sku).strip()
    for prefix in ["FAZEON ", "Fazeon ", "fazeon "]:
        if sku.startswith(prefix):
            sku = sku[len(prefix):]
            break
    return sku.strip().upper()


FIRMA_LISTESI = ["ITOPYA", "HB", "VATAN", "MONDAY", "KANAL", "DIGER"]

def excel_yukle_ana_stok(dosya_yolu):
    """
    Ana stok sekmesini yükler.
    Beklenen kolonlar: SKU, Ürün Adı, Kategori, Marka, Fiyat, Özellikler, Bizim Stok, Trendyol Stok
    """
    try:
        df = pd.read_excel(dosya_yolu, sheet_name=0)  # İlk sekmeyi oku (G5F STOK)
        df.columns = [tr_upper(c) for c in df.columns]
        
        kolon_esleme = {
            "SKU": ["SKU", "KOD", "URUN KODU", "BARKOD"],
            "URUN_ADI": ["URUN ADI", "AD", "URUN", "PRODUCT"],
            "KATEGORI": ["KATEGORI", "CATEGORY"],
            "MARKA": ["MARKA", "BRAND"],
            "SATIS_FIYATI": ["SATIS FIYATI ($)", "SATIS FIYATI", "FIYAT", "PRICE"],
            "HEDEF_KAR": ["HEDEF KAR MARJI (%)", "HEDEF KAR MARJI", "HEDEF KAR", "KAR MARJI", "MARGIN"],
            "BIZIM_STOK": ["BIZIM STOK", "DEPO STOK", "STOK", "G5F STOK"],
            "YOLDAKI_MIKTAR": ["YOLDAKI MIKTAR", "YOL MIKTAR", "YOLDA"],
            "VARIS_TARIHI": ["TAHMINI VARIS TARIHI", "VARIS TARIHI", "TAHMINI VARIS", "ETA"],
            "YOLDAKI_TEDARIKCI": ["YOLDAKI TEDARIKCI", "TEDARIKCI"],
        }
        
        kolon_map = {}
        for hedef, alternatifler in kolon_esleme.items():
            for alt in alternatifler:
                if tr_upper(alt) in df.columns:
                    kolon_map[hedef] = tr_upper(alt)
                    break
        
        if "SKU" not in kolon_map:
            return False, "SKU/Ürün Kodu kolonu bulunamadı."
        if "URUN_ADI" not in kolon_map:
            return False, "Ürün Adı kolonu bulunamadı."
        
        basarili = 0
        hatali = 0
        hata_mesajlari = []
        for _, row in df.iterrows():
            try:
                sku = normalize_sku(row[kolon_map["SKU"]])
                if not sku or sku == "nan":
                    continue
                urun_adi = str(row.get(kolon_map.get("URUN_ADI", ""), "")).strip()
                kategori = str(row.get(kolon_map.get("KATEGORI", ""), "")).strip() if "KATEGORI" in kolon_map else ""
                marka = str(row.get(kolon_map.get("MARKA", ""), "")).strip() if "MARKA" in kolon_map else ""
                ozellikler = str(row.get(kolon_map.get("OZELLIKLER", ""), "")).strip() if "OZELLIKLER" in kolon_map else ""
                satis_fiyati = safe_float(row.get(kolon_map.get("SATIS_FIYATI", ""), 0))
                hedef_kar = safe_float(row.get(kolon_map.get("HEDEF_KAR", ""), 0))
                bizim_stok = safe_int(row.get(kolon_map.get("BIZIM_STOK", ""), 0))
                yoldaki_miktar = safe_int(row.get(kolon_map.get("YOLDAKI_MIKTAR", ""), 0))
                varis_tarihi = safe_str(row.get(kolon_map.get("VARIS_TARIHI", ""), ""))
                yoldaki_tedarikci = safe_str(row.get(kolon_map.get("YOLDAKI_TEDARIKCI", ""), ""))
                kategori = safe_str(row.get(kolon_map.get("KATEGORI", ""), ""))
                marka = safe_str(row.get(kolon_map.get("MARKA", ""), ""))

                upsert_urun(sku, urun_adi, kategori, marka, satis_fiyati, 0, hedef_kar, "", bizim_stok, 0)

                # Yoldaki veriyi de kaydet
                if yoldaki_miktar > 0 or varis_tarihi:
                    upsert_yoldaki_urun(sku, urun_adi, yoldaki_miktar, varis_tarihi)
                basarili += 1
            except Exception as e:
                hatali += 1
                hata_mesajlari.append(f"{sku}: {str(e)}")
        
        hata_detay = f" | Hatalar: {'; '.join(hata_mesajlari[:3])}" if hata_mesajlari else ""
        return True, f"{basarili} ürün başarıyla yüklendi. {hatali} satır atlandı.{hata_detay}"
    except Exception as e:
        return False, f"Dosya okunamadı: {str(e)}"


def excel_yukle_firma_stoklari(dosya_yolu):
    """
    Firma stok sekmelerini yükler.
    Her firma için ayrı sekme: ITOPYA, HB, VATAN, MONDAY, KANAL, DIGER
    Beklenen kolonlar: SKU, Ürün Adı, Stok Miktarı, Haftalık Satış
    """
    try:
        xl = pd.ExcelFile(dosya_yolu)
        mevcut_sekmeler = [s.strip().upper() for s in xl.sheet_names]
        
        sonuclar = []
        for firma in FIRMA_LISTESI:
            # Sekme adı eşleştirme (DİĞER -> DIGER vb.)
            eslesen_sekme = None
            for sekme in xl.sheet_names:
                if normalize_tr(sekme) == normalize_tr(firma):
                    eslesen_sekme = sekme
                    break
                if normalize_tr(firma) in normalize_tr(sekme):
                    eslesen_sekme = sekme
                    break
            
            if not eslesen_sekme:
                sonuclar.append(f"⚠️ {firma}: Sekme bulunamadı, atlandı.")
                continue
            
            df = pd.read_excel(dosya_yolu, sheet_name=eslesen_sekme)
            df.columns = [tr_upper(c) for c in df.columns]
            
            kolon_esleme = {
                "SKU": ["SKU", "KOD", "URUN KODU", "BARKOD"],
                "URUN_ADI": ["URUN ADI", "AD", "URUN"],
                "STOK": ["STOK MIKTARI", "STOK", "MEVCUT STOK", "ADET"],
                "SATIS": ["HAFTALIK SATIS", "SATIS", "SATIS ADEDI"],
            }
            
            kolon_map = {}
            for hedef, alternatifler in kolon_esleme.items():
                for alt in alternatifler:
                    if tr_upper(alt) in df.columns:
                        kolon_map[hedef] = tr_upper(alt)
                        break
            
            if "SKU" not in kolon_map:
                sonuclar.append(f"❌ {firma}: SKU kolonu bulunamadı.")
                continue
            
            basarili = 0
            for _, row in df.iterrows():
                try:
                    sku = normalize_sku(row[kolon_map["SKU"]])
                    if not sku or sku == "NAN":
                        continue
                    urun_adi = safe_str(row.get(kolon_map.get("URUN_ADI", ""), "")) if "URUN_ADI" in kolon_map else ""
                    stok = safe_int(row.get(kolon_map.get("STOK", ""), 0)) if "STOK" in kolon_map else 0
                    satis = safe_int(row.get(kolon_map.get("SATIS", ""), 0)) if "SATIS" in kolon_map else 0

                    # Eğer SKU urunler tablosunda yoksa otomatik ekle
                    _sb = get_client()
                    _mev = _sb.table("urunler").select("sku").eq("sku", sku).execute().data
                    if not _mev:
                        from datetime import date as _date
                        _sb.table("urunler").insert({
                            "sku": sku, "urun_adi": urun_adi or sku,
                            "kategori": "", "marka": "", "satis_fiyati": 0,
                            "alis_fiyati": 0, "hedef_kar_marji": 0, "ozellikler": "",
                            "bizim_stok": 0, "trendyol_stok": 0,
                            "guncelleme_tarihi": _date.today().isoformat(),
                        }).execute()

                    upsert_firma_stok(firma, sku, urun_adi, stok, satis)
                    basarili += 1
                except Exception as ex:
                    pass
            
            sonuclar.append(f"✅ {firma}: {basarili} ürün yüklendi.")
        
        return True, "\n".join(sonuclar)
    except Exception as e:
        return False, f"Dosya okunamadı: {str(e)}"


def excel_yukle_yoldaki_urunler(dosya_yolu):
    """
    Yoldaki ürünler sekmesini yükler.
    Beklenen kolonlar: SKU, Ürün Adı, Yoldaki Miktar, Tahmini Varış Tarihi
    """
    try:
        xl = pd.ExcelFile(dosya_yolu)
        eslesen_sekme = None
        for sekme in xl.sheet_names:
            s = sekme.strip().upper()
            if "YOLDAK" in s or "YOL" in s or "TRANSIT" in s or "SIPARIS" in s:
                eslesen_sekme = sekme
                break

        if not eslesen_sekme:
            return False, "❌ 'YOLDAKI' adında sekme bulunamadı."

        df = pd.read_excel(dosya_yolu, sheet_name=eslesen_sekme)
        df.columns = [str(c).strip().upper() for c in df.columns]

        kolon_esleme = {
            "SKU": ["SKU", "KOD", "ÜRÜN KODU", "URUN KODU"],
            "URUN_ADI": ["ÜRÜN ADI", "URUN ADI", "AD", "ÜRÜN"],
            "MIKTAR": ["YOLDAKI MIKTAR", "YOLDAKİ MİKTAR", "MİKTAR", "MIKTAR", "ADET", "SİPARİŞ MİKTARI"],
            "VARIS": ["TAHMİNİ VARIŞ", "TAHMINI VARIS", "VARIŞ TARİHİ", "VARIS TARIHI", "ETD", "ETA"],
        }

        kolon_map = {}
        for hedef, alternatifler in kolon_esleme.items():
            for alt in alternatifler:
                if tr_upper(alt) in df.columns:
                    kolon_map[hedef] = tr_upper(alt)
                    break

        if "SKU" not in kolon_map:
            return False, "❌ SKU kolonu bulunamadı."

        basarili = 0
        for _, row in df.iterrows():
            try:
                sku = normalize_sku(row[kolon_map["SKU"]])
                if not sku or sku == "nan":
                    continue
                urun_adi = str(row.get(kolon_map.get("URUN_ADI", ""), "")).strip() if "URUN_ADI" in kolon_map else ""
                miktar = int(row.get(kolon_map.get("MIKTAR", ""), 0) or 0) if "MIKTAR" in kolon_map else 0
                varis = str(row.get(kolon_map.get("VARIS", ""), "")).strip() if "VARIS" in kolon_map else ""
                if varis == "nan":
                    varis = ""
                upsert_yoldaki_urun(sku, urun_adi, miktar, varis)
                basarili += 1
            except:
                pass

        return True, f"✅ Yoldaki ürünler: {basarili} satır yüklendi."
    except Exception as e:
        return False, f"Dosya okunamadı: {str(e)}"



def create_sample_excel_bytes():
    """Örnek Excel şablonunu bellekte oluşturur ve bytes döndürür (Streamlit Cloud için)"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "G5F STOK"
    basliklar = ["SKU", "Ürün Adı", "Kategori", "Marka", "Satış Fiyatı ($)", "Hedef Kar Marjı (%)", "Bizim Stok", "Yoldaki Miktar", "Tahmini Varış Tarihi", "Yoldaki Tedarikçi"]
    for i, b in enumerate(basliklar, 1):
        cell = ws1.cell(row=1, column=i, value=b)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="1F4E79")
        cell.alignment = Alignment(horizontal="center")

    for row in [
        ["SKU001", "Samsung Galaxy S24", "Telefon", "Samsung", 980, 30, 50, 100, "2026-05-15", "ABC Elektronik"],
        ["SKU002", "iPhone 15", "Telefon", "Apple", 1500, 25, 20, 0, "", ""],
        ["SKU003", "Xiaomi Redmi Note 13", "Telefon", "Xiaomi", 340, 35, 80, 50, "2026-06-01", "XYZ İthalat"],
    ]:
        ws1.append(row)
    for col in ws1.columns:
        ws1.column_dimensions[col[0].column_letter].width = 18

    for firma in ["ITOPYA", "HB", "VATAN", "MONDAY", "KANAL", "DIGER"]:
        ws = wb.create_sheet(firma)
        for i, b in enumerate(["SKU", "Ürün Adı", "Stok Miktarı", "Haftalık Satış"], 1):
            cell = ws.cell(row=1, column=i, value=b)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", start_color="2E7D32")
            cell.alignment = Alignment(horizontal="center")
        for row in [["SKU001", "Samsung Galaxy S24", 10, 5], ["SKU002", "iPhone 15", 3, 2], ["SKU003", "Xiaomi Redmi Note 13", 25, 12]]:
            ws.append(row)
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18

    ws_yol = wb.create_sheet("YOLDAKI")
    for i, b in enumerate(["SKU", "Ürün Adı", "Yoldaki Miktar", "Tahmini Varış Tarihi"], 1):
        cell = ws_yol.cell(row=1, column=i, value=b)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="6A1B9A")
        cell.alignment = Alignment(horizontal="center")
    for row in [["SKU001", "Samsung Galaxy S24", 30, "2026-05-01"], ["SKU002", "iPhone 15", 10, "2026-04-25"]]:
        ws_yol.append(row)
    for col in ws_yol.columns:
        ws_yol.column_dimensions[col[0].column_letter].width = 22

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def create_sample_excel():
    """Örnek Excel şablonu oluşturur"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    
    # Ana Stok sekmesi
    ws1 = wb.active
    ws1.title = "G5F STOK"
    basliklar = ["SKU", "Ürün Adı", "Kategori", "Marka", "Satış Fiyatı ($)", "Hedef Kar Marjı (%)", "Bizim Stok", "Yoldaki Miktar", "Tahmini Varış Tarihi", "Yoldaki Tedarikçi"]
    for i, b in enumerate(basliklar, 1):
        cell = ws1.cell(row=1, column=i, value=b)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="1F4E79")
        cell.alignment = Alignment(horizontal="center")

    ornek_veri = [
        ["SKU001", "Samsung Galaxy S24", "Telefon", "Samsung", 980, 30, 50, 100, "2026-05-15", "ABC Elektronik"],
        ["SKU002", "iPhone 15", "Telefon", "Apple", 1500, 25, 20, 0, "", ""],
        ["SKU003", "Xiaomi Redmi Note 13", "Telefon", "Xiaomi", 340, 35, 80, 50, "2026-06-01", "XYZ İthalat"],
    ]
    for row in ornek_veri:
        ws1.append(row)
    for col in ws1.columns:
        ws1.column_dimensions[col[0].column_letter].width = 18
    
    # Firma sekmeleri
    firma_listesi_tr = ["ITOPYA", "HB", "VATAN", "MONDAY", "KANAL", "DIGER"]
    for firma in firma_listesi_tr:
        ws = wb.create_sheet(firma)
        firma_basliklar = ["SKU", "Ürün Adı", "Stok Miktarı", "Haftalık Satış"]
        for i, b in enumerate(firma_basliklar, 1):
            cell = ws.cell(row=1, column=i, value=b)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", start_color="2E7D32")
            cell.alignment = Alignment(horizontal="center")
        ornek = [
            ["SKU001", "Samsung Galaxy S24", 10, 5],
            ["SKU002", "iPhone 15", 3, 2],
            ["SKU003", "Xiaomi Redmi Note 13", 25, 12],
        ]
        for row in ornek:
            ws.append(row)
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18

    # Yoldaki ürünler sekmesi
    ws_yol = wb.create_sheet("YOLDAKI")
    yol_basliklar = ["SKU", "Ürün Adı", "Yoldaki Miktar", "Tahmini Varış Tarihi"]
    for i, b in enumerate(yol_basliklar, 1):
        cell = ws_yol.cell(row=1, column=i, value=b)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="6A1B9A")
        cell.alignment = Alignment(horizontal="center")
    ornek_yol = [
        ["SKU001", "Samsung Galaxy S24", 30, "2026-05-01"],
        ["SKU002", "iPhone 15", 10, "2026-04-25"],
    ]
    for row in ornek_yol:
        ws_yol.append(row)
    for col in ws_yol.columns:
        ws_yol.column_dimensions[col[0].column_letter].width = 22

    path = "/home/claude/stok_app/SABLON_STOK_TAKIP.xlsx"
    wb.save(path)
    return path


def _firma_normalize(s):
    """Türkçe karakterleri sadeleştirip büyük harfe çevirir (İTOPYA → ITOPYA)."""
    return (str(s or "").strip().upper()
            .replace("İ", "I").replace("Ğ", "G").replace("Ü", "U")
            .replace("Ş", "S").replace("Ç", "C").replace("Ö", "O"))


def _firma_coz(firma_ham):
    """Excel'deki firma adını/kodunu standart koda çözer (ITOPYA/HB/VATAN/MONDAY/KANAL/DIGER).
    Doğrudan kod, ref_no tespit anahtarları (HEPSİBURADA→HB, EERA→ITOPYA, D-MARKET→HB...)
    ve önek üzerinden eşleştirir. Çözülemezse None döner."""
    fn = _firma_normalize(firma_ham)
    if not fn:
        return None
    if fn in FIRMA_LISTESI:
        return fn
    try:
        from .ref_no import FIRMA_ESLESME
        for rol, cfg in FIRMA_ESLESME.items():
            for t in cfg.get("tespit", ()):
                if _firma_normalize(t) == fn:
                    return rol
            if _firma_normalize(cfg.get("onek", "")) == fn:
                return rol
    except Exception:
        pass
    if fn in ("DIGER", "DIĞER"):
        return "DIGER"
    return None


def excel_yukle_firma_birlesik(dosya_yolu):
    """YENİ birleşik tek-sayfa firma stok+satış şablonunu yükler.
    Beklenen sütunlar: FİRMA ADI · KATEGORİ · MARKA · STOK KODU · STOK ADI ·
                       STOK · STOK-MAĞAZA · SATIŞ · SATIŞ-MAĞAZA
    Her satır bir firma-SKU. STOK ve STOK-MAĞAZA / SATIŞ ve SATIŞ-MAĞAZA AYRI saklanır.
    Sadece firma satırları işlenir; GSF STOK / YOLDAKI bu sayfada beklenmez (atlanır)."""
    try:
        df = pd.read_excel(dosya_yolu, sheet_name=0)
        df.columns = [tr_upper(c) for c in df.columns]

        kolon_esleme = {
            "FIRMA":        ["FIRMA ADI", "FIRMA", "MAGAZA", "FIRMA ADI "],
            "SKU":          ["STOK KODU", "SKU", "KOD", "URUN KODU", "BARKOD"],
            "URUN_ADI":     ["STOK ADI", "URUN ADI", "AD", "URUN"],
            "KATEGORI":     ["KATEGORI", "CATEGORY"],
            "MARKA":        ["MARKA", "BRAND"],
            "STOK":         ["STOK", "STOK MIKTARI", "ONLINE STOK"],
            "STOK_MAGAZA":  ["STOK-MAGAZA", "STOK MAGAZA", "MAGAZA STOK", "STOK_MAGAZA"],
            "SATIS":        ["SATIS", "HAFTALIK SATIS", "SATIS ADEDI", "ONLINE SATIS"],
            "SATIS_MAGAZA": ["SATIS-MAGAZA", "SATIS MAGAZA", "MAGAZA SATIS", "SATIS_MAGAZA"],
        }
        kolon_map = {}
        for hedef, alternatifler in kolon_esleme.items():
            for alt in alternatifler:
                if tr_upper(alt) in df.columns:
                    kolon_map[hedef] = tr_upper(alt)
                    break

        if "FIRMA" not in kolon_map:
            return False, "❌ 'FİRMA ADI' kolonu bulunamadı."
        if "SKU" not in kolon_map:
            return False, "❌ 'STOK KODU' kolonu bulunamadı."

        # MAĞAZA kolonları SAYI mı METİN mi? (yeni raporlarda mağaza ADI geliyor:
        # 'ACB Depo', 'AirPort Magaza'... → satırlar SKU bazında TOPLANMALI, ezilmemeli)
        def _metin_mi(kol_adi):
            if kol_adi not in kolon_map:
                return False
            _v = df[kolon_map[kol_adi]].dropna()
            if _v.empty:
                return False
            _say = 0
            for x in _v.head(30):
                try:
                    float(str(x).replace(",", "."))
                except Exception:
                    _say += 1
            return _say >= max(1, int(len(_v.head(30)) * 0.5))
        _stok_mgz_metin = _metin_mi("STOK_MAGAZA")
        _satis_mgz_metin = _metin_mi("SATIS_MAGAZA")
        _kirilim_modu = _stok_mgz_metin or _satis_mgz_metin

        gecerli_firmalar = {f: f for f in FIRMA_LISTESI}  # normalize edilmiş hâlleri
        basarili, atlanan = 0, 0
        firma_sayac = {}
        _firma_satir = {}
        atlanan_firma = set()

        _agg = {}  # (firma, sku) → {urun_adi, stok, stok_magaza, satis, satis_magaza}
        for _, row in df.iterrows():
            try:
                firma_ham = safe_str(row.get(kolon_map["FIRMA"], ""))
                sku = normalize_sku(row.get(kolon_map["SKU"], ""))
                if not sku or sku.lower() == "nan" or not firma_ham:
                    atlanan += 1
                    continue
                firma_n = _firma_normalize(firma_ham)
                if firma_n in ("GSF STOK", "G5F STOK", "YOLDAKI", "YOLDAKİ", "BIZIM STOK", "DEPO"):
                    atlanan += 1
                    continue
                firma = _firma_coz(firma_ham)
                if not firma:
                    atlanan += 1
                    atlanan_firma.add(firma_ham)
                    continue

                urun_adi = safe_str(row.get(kolon_map.get("URUN_ADI", ""), ""))
                stok = safe_int(row.get(kolon_map.get("STOK", ""), 0))
                satis = safe_int(row.get(kolon_map.get("SATIS", ""), 0))

                if _kirilim_modu:
                    # Satır = tek mağazanın kırılımı → TOPLA. Mağaza ADI'na göre ayır:
                    #   stok: adı MAGAZA/TESHIR/PAZARLAMA içeren → mağaza stoku; diğerleri (Depo/Internet/Servis) → merkez stok
                    #   satış: kanal İNTERNET/ONLINE → online satış; diğer (mağaza adı) → mağaza satışı
                    _smgz_ad = tr_upper(safe_str(row.get(kolon_map.get("STOK_MAGAZA", ""), "")))
                    _vmgz_ad = tr_upper(safe_str(row.get(kolon_map.get("SATIS_MAGAZA", ""), "")))
                    _stok_mgz_mi = any(k in _smgz_ad for k in ("MAGAZA", "TESHIR", "PAZARLAMA", "SHOWROOM"))
                    _satis_online_mi = (not _vmgz_ad or _vmgz_ad in ("0", "NAN", "GENEL")
                                        or "INTERNET" in _vmgz_ad or "ONLINE" in _vmgz_ad
                                        or "E-TICARET" in _vmgz_ad or "ETICARET" in _vmgz_ad)
                    stok_ana = 0 if _stok_mgz_mi else stok
                    stok_magaza = stok if _stok_mgz_mi else 0
                    satis_ana = satis if _satis_online_mi else 0
                    satis_magaza = 0 if _satis_online_mi else satis
                else:
                    stok_ana = stok
                    stok_magaza = safe_int(row.get(kolon_map.get("STOK_MAGAZA", ""), 0))
                    satis_ana = satis
                    satis_magaza = safe_int(row.get(kolon_map.get("SATIS_MAGAZA", ""), 0))

                _k = (firma, sku)
                _firma_satir[firma] = _firma_satir.get(firma, 0) + 1
                _o = _agg.get(_k)
                if _o is None:
                    _agg[_k] = {"urun_adi": urun_adi, "stok": stok_ana, "stok_magaza": stok_magaza,
                                "satis": satis_ana, "satis_magaza": satis_magaza}
                else:
                    _o["stok"] += stok_ana
                    _o["stok_magaza"] += stok_magaza
                    _o["satis"] += satis_ana
                    _o["satis_magaza"] += satis_magaza
                    if urun_adi and not _o["urun_adi"]:
                        _o["urun_adi"] = urun_adi
            except Exception:
                atlanan += 1

        for (firma, sku), _o in _agg.items():
            try:
                upsert_firma_stok(firma, sku, _o["urun_adi"], _o["stok"], _o["satis"],
                                  stok_magaza=_o["stok_magaza"], satis_magaza=_o["satis_magaza"])
                firma_sayac[firma] = firma_sayac.get(firma, 0) + 1
                basarili += 1
            except Exception:
                atlanan += 1

        ozet = " · ".join(
            (f"{f}: {n} SKU ({_firma_satir.get(f, n)} satırdan)"
             if _firma_satir.get(f, n) != n else f"{f}: {n} SKU")
            for f, n in firma_sayac.items()) or "kayıt yok"
        uyari = ""
        if atlanan_firma:
            uyari = f" | ⚠️ Tanınmayan firma(lar) atlandı: {', '.join(sorted(atlanan_firma)[:5])}"
        _mod_not = (" | 📊 Mağaza kırılımlı format algılandı: aynı SKU'nun satırları TOPLANDI "
                    "(stok: mağaza/merkez ayrımı mağaza adından; satış: İnternet=online, diğerleri mağaza)."
                    if _kirilim_modu else "")
        return True, f"✅ {basarili} firma-SKU yüklendi ({ozet}). {atlanan} satır atlandı.{uyari}{_mod_not}"
    except Exception as e:
        return False, f"❌ Dosya okunamadı: {type(e).__name__}: {str(e)[:160]}"


# G5F depo kırılımı — "satılabilir ana stok" sayılan depolar (sipariş önerisi/analitik için)
G5F_SATILABILIR_DEPOLAR = {"MERKEZ DEPO", "HAPPY LIFE"}


def excel_yukle_g5f_depolar(dosya_yolu):
    """G5F (bizim depo) çok-depolu stok şablonu.
    Beklenen sütunlar: DEPO ADI · STOK KODU · STOK İSMİ · MİKTAR
    Her SKU için depo kırılımı saklanır (TÜM depolar). bizim_stok (analitik) =
    satılabilir depolar (Merkez + Happy Life) toplamı. Fiyat/kategori/marka KORUNUR."""
    try:
        from collections import defaultdict
        df = pd.read_excel(dosya_yolu, sheet_name=0)
        df.columns = [tr_upper(c) for c in df.columns]
        kolon_esleme = {
            "DEPO":     ["DEPO ADI", "DEPO", "DEPO ISMI", "AMBAR", "DEPO ISMI "],
            "SKU":      ["STOK KODU", "SKU", "KOD", "URUN KODU", "BARKOD"],
            "URUN_ADI": ["STOK ISMI", "STOK ADI", "URUN ADI", "AD", "URUN"],
            "MIKTAR":   ["MIKTAR", "ADET", "STOK", "STOK MIKTARI"],
        }
        kolon_map = {}
        for hedef, alts in kolon_esleme.items():
            for alt in alts:
                if tr_upper(alt) in df.columns:
                    kolon_map[hedef] = tr_upper(alt)
                    break
        if "SKU" not in kolon_map:
            return False, "❌ 'STOK KODU' kolonu bulunamadı."
        if "DEPO" not in kolon_map:
            return False, "❌ 'DEPO ADI' kolonu bulunamadı."
        if "MIKTAR" not in kolon_map:
            return False, "❌ 'MİKTAR' kolonu bulunamadı."

        kirilim = defaultdict(dict)   # sku -> {depo: miktar}
        adlar = {}
        depolar_set = set()
        atlanan = 0
        for _, row in df.iterrows():
            sku = normalize_sku(row.get(kolon_map["SKU"], ""))
            if not sku or sku.lower() == "nan":
                atlanan += 1
                continue
            depo = depo_kanonik(safe_str(row.get(kolon_map["DEPO"], "")) or "Bilinmeyen")
            mik = safe_int(row.get(kolon_map["MIKTAR"], 0))
            kirilim[sku][depo] = kirilim[sku].get(depo, 0) + mik
            depolar_set.add(depo)
            if "URUN_ADI" in kolon_map and not adlar.get(sku):
                adlar[sku] = safe_str(row.get(kolon_map["URUN_ADI"], ""))

        # Mevcut ürünleri çek → SKU'yu büyük/küçük harf farkı GÖZETMEDEN eşleştir.
        # (Excel 'Faze1' → normalize 'FAZE1'; sistemde 'Faze1' kayıtlıysa yine eşleşsin,
        #  yeni mükerrer kayıt oluşmasın ve mevcut ürünün stoğu güncellensin.)
        mevcut_sku_map = {}   # {NORMALIZE: gercek_sku}
        try:
            _mr = get_client().table("urunler").select("sku").execute().data or []
            for _r in _mr:
                _gs = str(_r.get("sku") or "").strip()
                if _gs:
                    mevcut_sku_map[normalize_sku(_gs)] = _gs
        except Exception:
            pass

        basarili, toplam_adet, eslesen, yeni = 0, 0, 0, 0
        for sku, dd in kirilim.items():
            gercek_sku = mevcut_sku_map.get(sku, sku)   # mevcut varsa onun yazımıyla güncelle
            if sku in mevcut_sku_map:
                eslesen += 1
            else:
                yeni += 1
            satilabilir = sum(m for d, m in dd.items()
                              if _firma_normalize(d) in G5F_SATILABILIR_DEPOLAR)
            upsert_g5f_stok(gercek_sku, adlar.get(sku, ""), satilabilir, dd)
            basarili += 1
            toplam_adet += sum(dd.values())

        # ── BİREBİR SENKRON: Excel'de OLMAYAN ürünlerin eski kırılımını sıfırla ──
        # Böylece sistemdeki depo stoğu, yüklenen dosyanın birebir aynısı olur;
        # sonraki ithalat teslimleri (Model B) bu temiz tabanın ÜZERİNE işlemeye devam eder.
        sifirlanan = 0
        try:
            _tum = get_client().table("urunler").select("sku, depo_kirilim").execute().data or []
            for _r in _tum:
                _gs = str(_r.get("sku") or "").strip()
                _dk = _r.get("depo_kirilim") or {}
                if not _gs or not isinstance(_dk, dict) or not any(safe_int(v) for v in _dk.values()):
                    continue
                if normalize_sku(_gs) not in kirilim:
                    get_client().table("urunler").update(
                        {"depo_kirilim": {}, "bizim_stok": 0}).eq("sku", _gs).execute()
                    sifirlanan += 1
        except Exception:
            pass

        depo_liste = ", ".join(sorted(depolar_set))
        _sfr = (f" · 🧹 Excel'de olmayan {sifirlanan} ürünün eski kırılımı sıfırlandı (birebir senkron)"
                if sifirlanan else "")
        return True, (f"✅ {basarili} ürün yüklendi · {eslesen} mevcut güncellendi, {yeni} yeni eklendi · "
                      f"{len(depolar_set)} depo ({depo_liste}) · toplam {toplam_adet:,} adet{_sfr}. "
                      f"'bizim stok' = Merkez + Happy Life.")
    except Exception as e:
        return False, f"❌ Dosya okunamadı: {type(e).__name__}: {str(e)[:160]}"


# ═══════════ HAFTALIK STOK+SATIŞ — firma başına 2 sekme, PORTAL formatları ═══════════
_HSS_FIRMA_TOKEN = [("ITOPYA", "ITOPYA"), ("EERA", "ITOPYA"),
                    ("VATAN", "VATAN"),
                    ("HEPSIBURADA", "HB"), ("HEPSİBURADA", "HB"), ("HB", "HB"),
                    ("MONDAY", "MONDAY"), ("KANAL", "KANAL"), ("DIGER", "DIGER"), ("DİĞER", "DIGER")]


def _hss_kolon(df, *adaylar):
    """Başlık adaylarından ilk eşleşen kolonu döndürür (Türkçe-İ uyumlu, boşluk esnek)."""
    def _n(s):
        s = tr_upper(str(s)).replace(" ", "").replace("_", "").replace("-", "")
        return s
    kolmap = {_n(c): c for c in df.columns}
    for a in adaylar:
        _a = _n(a)
        if _a in kolmap:
            return kolmap[_a]
    for a in adaylar:
        _a = _n(a)
        for k, v in kolmap.items():
            if _a in k:
                return v
    return None


def excel_yukle_haftalik_stok_satis(dosya_yolu):
    """Firma başına AYRI 'X STOK' + 'X SATIŞ' sekmeleri olan haftalık dosyayı yükler.
    Her firmanın PORTAL formatı desteklenir (STOKKODU/Kod/Sku/Malzeme/Ürün Kodu...).
    Satışlar SKU ile stok satırlarının yanına bağlanır → firma_stok'a tek özet yazılır.
    KATEGORİ dosyadan BEKLENMEZ (boş/yok → hata vermez) — kategori bizim ürün kartından gelir."""
    try:
        sheets = pd.read_excel(dosya_yolu, sheet_name=None)
    except Exception as e:
        return False, f"❌ Dosya okunamadı: {type(e).__name__}: {str(e)[:140]}"

    def _sayfa_firma(ad):
        _u = tr_upper(str(ad))
        for tok, kod in _HSS_FIRMA_TOKEN:
            if tr_upper(tok) in _u:
                return kod
        return None

    # firma → {"stok": df, "satis": df}
    gruplar = {}
    for ad, df in (sheets or {}).items():
        kod = _sayfa_firma(ad)
        if not kod:
            continue
        _u = tr_upper(str(ad))
        tur = "satis" if ("SATIS" in _u.replace("Ş", "S") or "SATIŞ" in _u) else \
              ("stok" if "STOK" in _u else None)
        if tur:
            gruplar.setdefault(kod, {})[tur] = df

    if not gruplar:
        return False, ("❌ Firma sekmesi bulunamadı. Sekme adları 'VATAN STOK', 'VATAN SATIŞ' "
                       "gibi FİRMA + STOK/SATIŞ içermeli.")

    _SKU_ADAY = ("STOKKODU", "STOK KODU", "SKU", "MALZEME", "ÜRÜN KODU", "URUN KODU", "KOD")
    _AD_ADAY = ("STOKADI", "STOK ADI", "SKU ADI", "ÜRÜN ADI", "URUN ADI", "TANIM", "AD")
    _STOK_ADAY = ("ADET", "STOK", "DEPO MIKTAR", "MİKTAR", "MIKTAR")
    _SATIS_ADAY = ("MIKTAR", "MİKTAR", "ADET", "SIPARIŞ MIKTARI", "SİPARİŞ MİKTARI", "SATIŞ MIKTAR", "SATIS MIKTAR")
    _KANAL_ADAY = ("MAĞAZA", "MAGAZA", "DEPO")

    def _mgz_mi(ad):
        _u = tr_upper(str(ad))
        return any(k in _u for k in ("MAGAZA", "MAĞAZA", "TESHIR", "TEŞHIR", "PAZARLAMA", "SHOWROOM"))

    def _online_mi(ad):
        _u = tr_upper(str(ad)).strip()
        return (not _u or _u in ("0", "NAN", "GENEL")
                or "INTERNET" in _u or "İNTERNET" in _u or "ONLINE" in _u
                or "E-TICARET" in _u or "ETICARET" in _u)

    firma_ozet, atlanan_sayfa = {}, []
    basarili = 0
    for kod, g in gruplar.items():
        agg = {}  # sku → {ad, stok, stok_magaza, satis, satis_magaza}

        # ── STOK sekmesi ──
        sdf = g.get("stok")
        if sdf is not None and len(sdf):
            c_sku = _hss_kolon(sdf, *_SKU_ADAY)
            c_ad = _hss_kolon(sdf, *_AD_ADAY)
            c_adet = _hss_kolon(sdf, *_STOK_ADAY)
            c_depo = _hss_kolon(sdf, "DEPO")
            if c_sku is None or c_adet is None:
                atlanan_sayfa.append(f"{kod} STOK (SKU/adet kolonu yok)")
            else:
                for _, r in sdf.iterrows():
                    sku = normalize_sku(r.get(c_sku, ""))
                    if not sku or sku.lower() == "nan":
                        continue
                    adet = safe_int(r.get(c_adet, 0))
                    o = agg.setdefault(sku, {"ad": "", "stok": 0, "stok_magaza": 0,
                                             "satis": 0, "satis_magaza": 0})
                    if c_ad is not None and not o["ad"]:
                        o["ad"] = safe_str(r.get(c_ad, ""))
                    if c_depo is not None and _mgz_mi(r.get(c_depo, "")):
                        o["stok_magaza"] += adet
                    else:
                        o["stok"] += adet

        # ── SATIŞ sekmesi → SKU ile stokun yanına bağlanır ──
        vdf = g.get("satis")
        if vdf is not None and len(vdf):
            c_sku = _hss_kolon(vdf, *_SKU_ADAY)
            c_ad = _hss_kolon(vdf, *_AD_ADAY)
            c_adet = _hss_kolon(vdf, *_SATIS_ADAY)
            c_kanal = _hss_kolon(vdf, *_KANAL_ADAY)
            if c_sku is None or c_adet is None:
                atlanan_sayfa.append(f"{kod} SATIŞ (SKU/adet kolonu yok)")
            else:
                for _, r in vdf.iterrows():
                    sku = normalize_sku(r.get(c_sku, ""))
                    if not sku or sku.lower() == "nan":
                        continue
                    adet = safe_int(r.get(c_adet, 0))
                    o = agg.setdefault(sku, {"ad": "", "stok": 0, "stok_magaza": 0,
                                             "satis": 0, "satis_magaza": 0})
                    if c_ad is not None and not o["ad"]:
                        o["ad"] = safe_str(r.get(c_ad, ""))
                    if c_kanal is not None and not _online_mi(r.get(c_kanal, "")):
                        o["satis_magaza"] += adet
                    else:
                        o["satis"] += adet

        # ── Özet yaz (kategori bizim ürün kartından gelir; dosyadan beklenmez) ──
        _n_sku, _t_stok, _t_satis = 0, 0, 0
        for sku, o in agg.items():
            try:
                upsert_firma_stok(kod, sku, o["ad"], o["stok"], o["satis"],
                                  stok_magaza=o["stok_magaza"], satis_magaza=o["satis_magaza"])
                _n_sku += 1
                _t_stok += o["stok"] + o["stok_magaza"]
                _t_satis += o["satis"] + o["satis_magaza"]
                basarili += 1
            except Exception:
                pass
        if _n_sku:
            firma_ozet[kod] = (_n_sku, _t_stok, _t_satis)

    if not basarili:
        return False, ("❌ Hiç kayıt yazılamadı. Sekmelerde veri var mı ve SKU/adet kolonları dolu mu kontrol et."
                       + (f" Atlanan: {', '.join(atlanan_sayfa)}" if atlanan_sayfa else ""))
    ozet = " · ".join(f"{k}: {n} SKU (stok {s:,} / satış {v:,})"
                      for k, (n, s, v) in firma_ozet.items())
    msg = f"✅ Haftalık stok+satış yüklendi → {ozet}."
    if atlanan_sayfa:
        msg += f" ⚠️ Atlanan sekme: {', '.join(atlanan_sayfa)}."
    msg += " Kategoriler ürün kartlarından eşlenir (dosyada kategori gerekmez)."
    return True, msg


# ═══════════ STOK KARTLARI (TAM LİSTE) — kart aç · barkod tamamla · kategori eşitle ═══════════
def excel_yukle_stok_kartlari(dosya_yolu):
    """MARKA · STOK KODU · STOK ADI · BARKOD · KATEGORİ kolonlu tam kart listesi.
    Kurallar (bütünlük bozulmaz, mükerrer oluşmaz):
      • Yeni SKU → kart açılır (ad+kategori+marka+barkod).
      • Mevcut SKU → yalnız HEDEFLİ alanlar güncellenir (fiyat/stok/paçala DOKUNULMAZ):
          - barkod: karttaki boşsa dosyadakiyle doldurulur
          - kategori: dosyadakinden farklıysa eşitlenir (SSD&RAM → ssd / ram ayrımı böyle yapılır)
          - marka / ürün adı: karttaki boşsa doldurulur
    Kategoriler programın standardında (Türkçe küçük harf) yazılır; dosyadaki yeni
    kategoriler (örn. mouse pad) otomatik oluşur."""
    from .database import tr_kucuk
    try:
        df = pd.read_excel(dosya_yolu)
    except Exception as e:
        return False, f"❌ Dosya okunamadı: {type(e).__name__}: {str(e)[:140]}"

    def _n(s):
        return tr_upper(str(s)).replace(" ", "").replace("_", "")
    kol = {_n(c): c for c in df.columns}

    def _bul(*adlar):
        for a in adlar:
            if _n(a) in kol:
                return kol[_n(a)]
        for a in adlar:
            for k, v in kol.items():
                if _n(a) in k:
                    return v
        return None
    c_sku = _bul("STOK KODU", "SKU", "KOD")
    c_ad = _bul("STOK ADI", "ÜRÜN ADI", "AD")
    c_bar = _bul("BARKOD")
    c_kat = _bul("KATEGORİ", "KATEGORI")
    c_mar = _bul("MARKA")
    if c_sku is None:
        return False, "❌ 'STOK KODU' kolonu bulunamadı."

    def _bar_str(v):
        s = safe_str(v)
        if not s:
            return ""
        try:  # 4718009611191.0 → 4718009611191
            f = float(s.replace(",", "."))
            if f == int(f):
                return str(int(f))
        except Exception:
            pass
        return s

    def _kn(s):
        s = str(s or "").strip().lower()
        for a, b in (("i̇", "i"), ("ı", "i"), ("ş", "s"), ("ğ", "g"),
                     ("ü", "u"), ("ö", "o"), ("ç", "c")):
            s = s.replace(a, b)
        return s

    try:
        sb = get_client()
        mevcutlar = {}
        _bas = 0
        while True:
            r = sb.table("urunler").select("sku, urun_adi, kategori, marka, barkod") \
                  .range(_bas, _bas + 999).execute()
            _chunk = r.data or []
            for u in _chunk:
                mevcutlar[str(u.get("sku") or "").strip()] = u
            if len(_chunk) < 1000:
                break
            _bas += 1000
    except Exception as e:
        return False, f"❌ Ürünler okunamadı: {type(e).__name__}: {str(e)[:120]}"

    yeni = barkod_eklendi = kategori_esitlendi = alan_dolduruldu = ayni = hatali = 0
    kat_degisim = {}
    for _, r in df.iterrows():
        sku = normalize_sku(r.get(c_sku, ""))
        if not sku or sku.lower() == "nan":
            continue
        ad = safe_str(r.get(c_ad, "")) if c_ad is not None else ""
        bar = _bar_str(r.get(c_bar)) if c_bar is not None else ""
        kat = tr_kucuk(safe_str(r.get(c_kat, ""))) if c_kat is not None else ""
        mar = safe_str(r.get(c_mar, "")) if c_mar is not None else ""

        u = mevcutlar.get(sku)
        try:
            if u is None:
                upsert_urun(sku, ad, kategori=kat, marka=mar)
                if bar:
                    sb.table("urunler").update({"barkod": bar}).eq("sku", sku).execute()
                yeni += 1
                continue
            _upd = {}
            if bar and not safe_str(u.get("barkod")):
                _upd["barkod"] = bar
                barkod_eklendi += 1
            _eski_kat = safe_str(u.get("kategori"))
            if kat and _kn(kat) != _kn(_eski_kat):
                _upd["kategori"] = kat
                kategori_esitlendi += 1
                kat_degisim[f"{_eski_kat or '(boş)'} → {kat}"] = \
                    kat_degisim.get(f"{_eski_kat or '(boş)'} → {kat}", 0) + 1
            if mar and not safe_str(u.get("marka")):
                _upd["marka"] = mar
                alan_dolduruldu += 1
            if ad and not safe_str(u.get("urun_adi")):
                _upd["urun_adi"] = ad
                alan_dolduruldu += 1
            if _upd:
                sb.table("urunler").update(_upd).eq("sku", sku).execute()
            else:
                ayni += 1
        except Exception:
            hatali += 1

    try:
        from .database import _cache_temizle
        _cache_temizle()
    except Exception:
        pass

    _dosya_skular = {normalize_sku(r.get(c_sku, "")) for _, r in df.iterrows()}
    _dosyada_olmayan = [s for s in mevcutlar if s and s not in _dosya_skular]
    msg = (f"✅ Stok kartları işlendi: 🆕 {yeni} yeni kart · 🏷️ {barkod_eklendi} karta barkod eklendi · "
           f"📂 {kategori_esitlendi} kartın kategorisi eşitlendi · ✏️ {alan_dolduruldu} boş alan dolduruldu · "
           f"{ayni} kart zaten günceldi.")
    if kat_degisim:
        _ozet = " | ".join(f"{k}: {v}" for k, v in sorted(kat_degisim.items())[:8])
        msg += f" Kategori değişimleri → {_ozet}."
    if hatali:
        msg += f" ⚠️ {hatali} satır yazılamadı."
    if _dosyada_olmayan:
        msg += (f" ℹ️ Programda olup dosyada OLMAYAN {len(_dosyada_olmayan)} kart var "
                f"(dokunulmadı): {', '.join(sorted(_dosyada_olmayan)[:6])}"
                + (" …" if len(_dosyada_olmayan) > 6 else ""))
    return True, msg
