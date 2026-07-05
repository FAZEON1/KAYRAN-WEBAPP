"""
KAYRAN — Ürün & Stok Yönetim Sistemi
Modüler olarak KAYRAN portal içinden çağrılır.

Kullanım:
    from kayranpm.main import run
    run()
"""
import streamlit as st
import logging
_log = logging.getLogger(__name__)
# Türkiye saat dilimi için ortak yardımcılar
from shared.utils import tr_today, tr_now, tr_now_str, tr_tomorrow, tr_yesterday as _tr_today_iso_dummy
from shared.utils import tr_kucuk
from shared.utils import firma_gorunen_ad
from shared.utils import sidebar_stil, sidebar_baslik, sidebar_kullanici
from shared.utils import metrik_satiri, metric_css
from shared.ui import tablo_h
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
import os, sys
from datetime import datetime, date
from io import BytesIO

# Modül bazlı importlar (relative)
from .database import (initialize_db, onayla_siparis, reddet_siparis,
                      get_siparis_onerileri, ekle_siparis_onerisi,
                      get_gecmis_satis_firma_bazli, get_urun_detay,
                      ekle_kampanya, get_kampanyalar, get_kampanya,
                      guncelle_kampanya, kapat_kampanya, sil_kampanya,
                      ekle_kampanya_urun, get_kampanya_urunler, get_tum_kampanya_urunler,
                      guncelle_kampanya_urun, sil_kampanya_urun,
                      get_tum_sku_listesi, get_client,
                      get_gecmis_satis_tum_firmalar,
                      get_kampanya_destek_ortalamalari,
                      get_firma_listesi, get_musteri_haftalik_satis,
                      sku_fazeon_temizle_onizle, sku_fazeon_temizle_uygula)
from .analitik import dashboard_hesapla, tum_urunler_listesi, siparis_onerisi_listesi
from .excel_islemler import (excel_yukle_ana_stok, excel_yukle_firma_stoklari,
                            excel_yukle_yoldaki_urunler, create_sample_excel_bytes,
                            excel_yukle_firma_birlesik, excel_yukle_g5f_depolar,
                            excel_yukle_haftalik_stok_satis, excel_yukle_stok_kartlari)


def render_renkli_tablo(df, para=None, yuzde=None, kar=None, sol=None,
                        kisalt=None, gizle=None, satir_durum=None):
    """Bir DataFrame'i KAYRAN renkli HTML tablo temasinda (salt-okunur) cizer."""
    para = set(para or []); yuzde = set(yuzde or []); kar = set(kar or [])
    sol = set(sol or []); kisalt = kisalt or {}; gizle = set(gizle or [])
    if df is None or len(df) == 0:
        st.info("Gosterilecek veri yok.")
        return
    durum_kolon, durum_harita = (satir_durum or (None, {}))
    kolonlar = [c for c in df.columns if c not in gizle]

    def _isnum(c):
        try:
            return pd.api.types.is_numeric_dtype(df[c])
        except Exception:
            return False

    def _sag(c):
        return (c in para or c in yuzde or _isnum(c)) and c not in sol

    def _fmt(c, v):
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return "\u2014"
        try:
            if c in para:
                return f"${float(v):,.2f}"
            if c in yuzde:
                return f"%{float(v):.1f}"
            if _isnum(c):
                fv = float(v)
                return f"{int(fv):,}" if fv == int(fv) else f"{fv:,.2f}"
        except Exception:
            pass
        s = str(v)
        mx = kisalt.get(c)
        return (s[:mx-1] + "\u2026") if (mx and len(s) > mx) else s

    rows_html = ""
    for _, row in df.iterrows():
        drenk = durum_harita.get(str(row.get(durum_kolon, "")), "") if durum_kolon else ""
        tds = ""
        for c in kolonlar:
            v = row[c]
            base = "rk-num" if _sag(c) else "rk-txt"
            cls = base
            if c in kar:
                try:
                    fv = float(v)
                    cls = "rk-pos" if fv > 0 else ("rk-neg" if fv < 0 else "rk-num")
                except Exception:
                    cls = "rk-num"
            elif drenk:
                cls = base + " " + drenk
            full = str(v)
            mx = kisalt.get(c)
            ttl = f' title="{full.replace(chr(34), "&quot;")}"' if (mx and len(full) > mx) else ""
            tds += f'<td class="{cls}"{ttl}>{_fmt(c, v)}</td>'
        rows_html += f"<tr>{tds}</tr>"

    ths = "".join(f'<th class="{"" if _sag(c) else "l"}">{c}</th>' for c in kolonlar)
    css = (
        "<style>"
        ".rkw{overflow-x:auto;border-radius:12px;box-shadow:0 2px 14px rgba(0,0,0,0.25);margin:4px 0}"
        ".rkt{width:100%;border-collapse:collapse;font-family:Inter,sans-serif}"
        ".rkt thead tr{background:linear-gradient(135deg,#1E293B,#0F172A)}"
        ".rkt thead th{padding:8px 12px;color:#CBD5E1;font-size:11px;font-weight:700;letter-spacing:.4px;text-transform:uppercase;white-space:nowrap;text-align:right}"
        ".rkt thead th.l{text-align:left}"
        ".rkt tbody{background:#131C35}"
        ".rkt td{padding:8px 12px;font-size:11px;max-width:300px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}"
        ".rkt tbody tr{border-bottom:1px solid rgba(255,255,255,0.05)}"
        ".rkt tbody tr:hover{background:rgba(99,102,241,0.06)}"
        ".rk-txt{text-align:left;color:#CBD5E1}"
        ".rk-num{text-align:right;color:#CBD5E1;font-family:'JetBrains Mono',monospace}"
        ".rk-pos{text-align:right;color:#4ADE80 !important;font-weight:700;font-family:'JetBrains Mono',monospace}"
        ".rk-neg{text-align:right;color:#F87171 !important;font-weight:700;font-family:'JetBrains Mono',monospace}"
        ".rk-red{color:#F87171 !important;font-weight:600}.rk-org{color:#FB923C !important;font-weight:600}"
        ".rk-yel{color:#FCD34D !important;font-weight:600}.rk-grn{color:#4ADE80 !important;font-weight:600}.rk-dim{color:#64748B !important}"
        "</style>"
    )
    st.html(css + f'<div class="rkw"><table class="rkt"><thead><tr>{ths}</tr></thead><tbody>' + rows_html + "</tbody></table></div>")


def run():
    """KAYRAN ana çalıştırıcı. Portal tarafından çağrılır."""
    initialize_db()

    # İthalat'taki tüm modeller Ürün Yönetimi'ne otomatik yansısın (oturumda bir kez · ekleme-only, silme yok)
    if not st.session_state.get("_ith_autosync"):
        try:
            from .database import ithalat_eksikleri_ekle
            _yeni = ithalat_eksikleri_ekle()
            if _yeni:
                st.toast(f"🚢 İthalat'tan {_yeni} model ürünlere eklendi")
        except Exception:
            pass
        st.session_state["_ith_autosync"] = True

    st.markdown("""
    <style>
    /* ── GLOBAL ─────────────────────────────────────────────────────────── */
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');
    
    html, body, .stApp {
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif !important;
        background: #0A0F1E !important;
    }

    /* ── METRIC KARTLARI (modern · sade · profesyonel) ───────────────── */
    div[data-testid="stMetric"] {
        background: linear-gradient(180deg, rgba(255,255,255,0.045), rgba(255,255,255,0.015));
        border: 1px solid rgba(255,255,255,0.07);
        border-radius: 16px;
        padding: 16px 16px;
        box-shadow: 0 1px 3px rgba(0,0,0,0.22);
        transition: border-color .15s ease;
    }
    div[data-testid="stMetric"]:hover {
        border-color: rgba(99,102,241,0.32);
    }
    div[data-testid="stMetricLabel"],
    div[data-testid="stMetricLabel"] p,
    div[data-testid="stMetricLabel"] div {
        color: #94A3B8 !important;
        font-size: 11px !important;
        font-weight: 600 !important;
        letter-spacing: .5px !important;
        text-transform: uppercase !important;
    }
    div[data-testid="stMetricValue"] {
        color: #F1F5F9 !important;
        font-size: 25px !important;
        font-weight: 700 !important;
        font-variant-numeric: tabular-nums;
        line-height: 1.22 !important;
        margin-top: 3px;
        white-space: normal !important;
        word-break: break-word;
    }
    div[data-testid="stMetricDelta"] { font-size: 12px !important; }

    /* ── ALERT BOX'LARI (warning/error/info/success) ─────────────────── */
    /* Streamlit'in default renkleri dark tema'da okunmuyor. Manuel override. */
    div[data-testid="stAlert"] {
        border-radius: 10px !important;
        font-family: 'Inter', sans-serif !important;
        font-size: 13px !important;
        font-weight: 500 !important;
        padding: 0 !important;
        border: 1px solid transparent !important;
        box-shadow: none !important;
    }
    div[data-testid^="stAlertContent"] {
        background: transparent !important;
        border: none !important;
        padding: 11px 15px !important;
        color: #E2E8F0 !important;
    }
    div.stAlert:has([data-testid="stAlertContentWarning"]) {
        background: rgba(245,158,11,0.07) !important;
        border: 1px solid rgba(245,158,11,0.20) !important;
        border-left: 3px solid rgba(245,158,11,0.60) !important;
    }
    div.stAlert:has([data-testid="stAlertContentError"]) {
        background: rgba(239,68,68,0.07) !important;
        border: 1px solid rgba(239,68,68,0.20) !important;
        border-left: 3px solid rgba(239,68,68,0.60) !important;
    }
    div.stAlert:has([data-testid="stAlertContentInfo"]) {
        background: rgba(59,130,246,0.06) !important;
        border: 1px solid rgba(59,130,246,0.20) !important;
        border-left: 3px solid rgba(59,130,246,0.55) !important;
    }
    div.stAlert:has([data-testid="stAlertContentSuccess"]) {
        background: rgba(34,197,94,0.06) !important;
        border: 1px solid rgba(34,197,94,0.20) !important;
        border-left: 3px solid rgba(34,197,94,0.50) !important;
    }
    div[data-testid="stAlert"] p,
    div[data-testid="stAlert"] span,
    div[data-testid="stAlert"] strong,
    div[data-testid="stAlert"] div {
        color: #E2E8F0 !important;
    }
    div[data-testid="stAlert"] strong,
    div[data-testid="stAlert"] b {
        font-weight: 700 !important;
    }
    
    /* ── METRIC KARTLARI (özet barlar) ── eski 'metric-container' + yeni 'stMetric' */
    [data-testid="metric-container"],
    [data-testid="stMetric"] {
        background: linear-gradient(135deg, rgba(255,255,255,0.07) 0%, rgba(255,255,255,0.03) 100%);
        border-radius: 14px;
        padding: 16px 16px;
        border: 1px solid rgba(255,255,255,0.10);
        transition: border-color 0.2s;
        min-width: 0;
        overflow: hidden;
    }
    [data-testid="metric-container"]:hover,
    [data-testid="stMetric"]:hover { border-color: rgba(66,165,245,0.4); }
    [data-testid="stMetricLabel"],
    [data-testid="stMetricLabel"] * {
        color: #78909C !important; font-size:11px !important; font-weight:700 !important;
        letter-spacing:0.5px; text-transform:uppercase;
        white-space:nowrap !important; overflow:hidden !important; text-overflow:ellipsis !important;
    }
    [data-testid="stMetricValue"],
    [data-testid="stMetricValue"] * {
        color: #ECEFF1 !important; font-weight: 800 !important; font-size:23px !important;
        white-space:nowrap !important; overflow:hidden !important; text-overflow:ellipsis !important;
        line-height:1.15 !important;
    }
    [data-testid="stMetricDelta"] { font-size:13px !important; }
    
    /* ── BAŞLIK STİLLERİ ─────────────────────────────────────────────────── */
    .baslik {
        font-size: 22px;
        font-weight: 800;
        color: #ECEFF1;
        margin-bottom: 2px;
        letter-spacing: -0.3px;
    }
    .alt-baslik {
        font-size: 12px;
        color: #546E7A;
        margin-bottom: 24px;
        font-weight: 500;
        letter-spacing: 0.3px;
    }
    .sayfa-baslik-cizgi {
        height: 3px;
        background: linear-gradient(90deg, #1565C0, #42A5F5, transparent);
        border-radius: 2px;
        margin-bottom: 24px;
    }
    
    /* ── ETİKET KUTUCUKLARI ──────────────────────────────────────────────── */
    .tag-kirmizi { background:#7F0000; color:#FFCDD2; padding:4px 8px; border-radius:20px; font-size:11px; font-weight:700; }
    .tag-turuncu { background:#BF360C; color:#FFE0B2; padding:4px 8px; border-radius:20px; font-size:11px; font-weight:700; }
    .tag-sari    { background:#F57F17; color:#FFF9C4; padding:4px 8px; border-radius:20px; font-size:11px; font-weight:700; }
    .tag-yesil   { background:#1B5E20; color:#C8E6C9; padding:4px 8px; border-radius:20px; font-size:11px; font-weight:700; }
    .tag-mavi    { background:#0D47A1; color:#BBDEFB; padding:4px 8px; border-radius:20px; font-size:11px; font-weight:700; }
    .tag-gri     { background:#263238; color:#90A4AE; padding:4px 8px; border-radius:20px; font-size:11px; }
    
    /* ── BİLGİ KUTULARI ─────────────────────────────────────────────────── */
    .uyari-box {
        background: linear-gradient(135deg,#3E1800,#2E1200);
        border-left: 3px solid #FF6F00;
        color: #FFE0B2;
        padding: 12px 16px;
        border-radius: 0 8px 8px 0;
        margin: 8px 0;
        font-size: 13px;
        font-weight: 500;
    }
    .info-box {
        background: linear-gradient(135deg,#0A1929,#071526);
        border-left: 3px solid #42A5F5;
        color: #BBDEFB;
        padding: 12px 16px;
        border-radius: 0 8px 8px 0;
        margin: 8px 0;
        font-size: 13px;
        font-weight: 500;
    }
    .basari-box {
        background: linear-gradient(135deg,#0F2910,#091E0A);
        border-left: 3px solid #66BB6A;
        color: #C8E6C9;
        padding: 12px 16px;
        border-radius: 0 8px 8px 0;
        margin: 8px 0;
        font-size: 13px;
        font-weight: 500;
    }
    
    /* ── SIDEBAR ─────────────────────────────────────────────────────────── */
    section[data-testid="stSidebar"] {
        background: linear-gradient(180deg, #050B16 0%, #0A1628 100%) !important;
        border-right: 1px solid rgba(255,255,255,0.06) !important;
    }
    section[data-testid="stSidebar"] * { color: #CFD8DC !important; }
    /* Sidebar nav stili shared/utils.py → sidebar_stil() tarafından yönetilir */
    section[data-testid="stSidebar"] .stButton button {
        background: linear-gradient(135deg,#1A3A5C,#1565C0) !important;
        color: #ECEFF1 !important;
        border: none !important;
        border-radius: 8px !important;
        font-weight: 600 !important;
        font-size: 13px !important;
    }
    section[data-testid="stSidebar"] hr {
        border-color: rgba(255,255,255,0.08) !important;
        margin: 12px 0 !important;
    }
    
    /* ── BUTONLAR ────────────────────────────────────────────────────────── */
    .stButton > button {
        border-radius: 8px !important;
        font-weight: 600 !important;
        font-size: 13px !important;
        transition: all 0.2s !important;
    }
    .stButton > button[kind="primary"] {
        background: linear-gradient(135deg,#1565C0,#1E88E5) !important;
        border: none !important;
        color: white !important;
        box-shadow: 0 4px 12px rgba(21,101,192,0.3) !important;
    }
    .stButton > button[kind="primary"]:hover {
        box-shadow: 0 6px 20px rgba(21,101,192,0.5) !important;
        transform: translateY(-1px) !important;
    }
    
    /* ── FORM ALANLARI ───────────────────────────────────────────────────── */
    .stTextInput > div > div > input,
    .stNumberInput > div > div > input,
    .stTextArea > div > div > textarea,
    .stSelectbox > div > div {
        background: rgba(255,255,255,0.05) !important;
        border: 1px solid rgba(255,255,255,0.12) !important;
        border-radius: 8px !important;
        color: #ECEFF1 !important;
        font-size: 13px !important;
    }
    .stTextInput > div > div > input:focus,
    .stNumberInput > div > div > input:focus {
        border-color: #42A5F5 !important;
        box-shadow: 0 0 0 2px rgba(66,165,245,0.15) !important;
    }
    label[data-testid="stWidgetLabel"] p {
        color: #78909C !important;
        font-size: 12px !important;
        font-weight: 600 !important;
        letter-spacing: 0.3px !important;
    }
    
    /* ── EXPANDER ────────────────────────────────────────────────────────── */
    .streamlit-expanderHeader {
        background: rgba(255,255,255,0.04) !important;
        border-radius: 10px !important;
        border: 1px solid rgba(255,255,255,0.08) !important;
        color: #90CAF9 !important;
        font-weight: 600 !important;
        font-size: 13px !important;
    }
    .streamlit-expanderContent {
        background: rgba(255,255,255,0.02) !important;
        border: 1px solid rgba(255,255,255,0.06) !important;
        border-top: none !important;
        border-radius: 0 0 10px 10px !important;
    }
    
    /* ── TABS ────────────────────────────────────────────────────────────── */
    .stTabs [data-baseweb="tab-list"] {
        background: rgba(255,255,255,0.03) !important;
        border-radius: 10px !important;
        padding: 4px !important;
        gap: 2px !important;
    }
    .stTabs [data-baseweb="tab"] {
        border-radius: 8px !important;
        color: #78909C !important;
        font-weight: 600 !important;
        font-size: 13px !important;
        padding: 6px 16px !important;
    }
    .stTabs [aria-selected="true"] {
        background: rgba(21,101,192,0.4) !important;
        color: #90CAF9 !important;
    }
    
    /* ── DATAFRAME ───────────────────────────────────────────────────────── */
    .stDataFrame { border-radius: 10px !important; overflow: hidden !important; }
    .stDataFrame [data-testid="stDataFrameResizable"] {
        border: 1px solid rgba(255,255,255,0.08) !important;
        border-radius: 10px !important;
    }
    
    /* ── DIVIDER ─────────────────────────────────────────────────────────── */
    hr { border-color: rgba(255,255,255,0.06) !important; margin: 20px 0 !important; }
    
    /* ── TABLO HÜCRE RENKLERİ ────────────────────────────────────────────── */
    .hucre-acil    { background:#7F0000 !important; color:#FFCDD2 !important; font-weight:700; }
    .hucre-turuncu { background:#BF360C !important; color:#FFE0B2 !important; font-weight:600; }
    .hucre-sari    { background:#827717 !important; color:#F9A825 !important; font-weight:600; }
    .hucre-yesil   { background:#1B5E20 !important; color:#A5D6A7 !important; font-weight:600; }
    .hucre-gri     { background:#263238 !important; color:#CFD8DC !important; }
    .fcp-vurgu { color:#FFD54F; font-weight:800; font-size:15px; }
    
    /* ── SCROLLBAR ───────────────────────────────────────────────────────── */
    ::-webkit-scrollbar { width: 6px; height: 6px; }
    ::-webkit-scrollbar-track { background: rgba(255,255,255,0.02); }
    ::-webkit-scrollbar-thumb { background: rgba(255,255,255,0.15); border-radius: 3px; }
    ::-webkit-scrollbar-thumb:hover { background: rgba(255,255,255,0.25); }
    </style>
    """, unsafe_allow_html=True)
    
    # ── Yardımcı fonksiyonlar ────────────────────────────────────────────
    STOK_YAS_ETIKET = {
        "kirmizi": '<span class="tag-kirmizi">🔴 {g} gün</span>',
        "turuncu": '<span class="tag-turuncu">🟠 {g} gün</span>',
        "sari":    '<span class="tag-sari">🟡 {g} gün</span>',
        "yesil":   '<span class="tag-yesil">🟢 {g} gün</span>',
        "yok":     '<span class="tag-gri">—</span>',
    }
    GUN_ETIKET = {
        "kirmizi": '<span class="tag-kirmizi">🔴 {g} gün</span>',
        "turuncu": '<span class="tag-turuncu">🟠 {g} gün</span>',
        "yesil":   '<span class="tag-yesil">🟢 {g} gün</span>',
        "yok":     '<span class="tag-gri">—</span>',
    }
    PERF_ETIKET = {
        "Çok İyi": '<span class="tag-yesil">⭐ Çok İyi</span>',
        "İyi":     '<span class="tag-sari">👍 İyi</span>',
        "Düşük":   '<span class="tag-kirmizi">📉 Düşük</span>',
        "veri yok":'<span class="tag-gri">—</span>',
    }
    YOL_ETIKET = {
        "yesil":   "🟢",
        "sari":    "🟡",
        "kirmizi": "🔴",
        "yok":     "—",
    }
    
    def stok_yas_html(renk, gun):
        return STOK_YAS_ETIKET.get(renk, STOK_YAS_ETIKET["yok"]).format(g=gun)
    
    def gun_html(renk, gun):
        if gun is None: return '<span class="tag-gri">—</span>'
        return GUN_ETIKET.get(renk, GUN_ETIKET["yok"]).format(g=gun)
    
    def perf_html(perf):
        return PERF_ETIKET.get(perf, PERF_ETIKET["veri yok"])
    
    # ── Sidebar navigasyon ───────────────────────────────────────────────
    with st.sidebar:
        st.markdown('<script>var sidebarEl=window.parent.document.querySelector("[data-testid=stSidebar] > div");if(sidebarEl)sidebarEl.scrollTop=0;</script>', unsafe_allow_html=True)
        aktif_kullanici = st.session_state.get("aktif_kullanici", "")
        st.markdown(sidebar_stil(), unsafe_allow_html=True)
        st.markdown(sidebar_baslik("📦", "Ürün Yönetimi", "Stok Sistemi"), unsafe_allow_html=True)

        if aktif_kullanici:
            st.markdown(sidebar_kullanici(aktif_kullanici), unsafe_allow_html=True)
            if st.button("Çıkış Yap", use_container_width=True):
                st.session_state.giris_yapildi = False
                st.session_state.aktif_kullanici = ""
                st.rerun()
    
        st.markdown('<div style="height:6px;"></div>', unsafe_allow_html=True)
        sayfa = st.radio("Sayfa", [
            "📊  Dashboard",
            "📋  Tüm Ürünler",
            "📈  Müşteri Satışları",
            "🎯  Kampanya Takip",
            "📦  Sipariş Önerisi",
            "🔖  Ref No Takibi",
            "📂  Veri Yükleme",
        ], label_visibility="collapsed")

        st.markdown('<div style="height:12px;"></div>', unsafe_allow_html=True)
        # ── STOK KARTI — hızlı erişim (sık kullanılan eylem, öne çıkarılmış) ──
        st.markdown(
            '<div style="background:linear-gradient(135deg,rgba(99,102,241,0.14),rgba(34,211,238,0.05) 70%,transparent);'
            'border:1px solid rgba(129,140,248,0.28);border-radius:12px;padding:10px 12px 12px;margin-bottom:8px">'
            '<div style="display:flex;align-items:center;gap:8px;margin-bottom:8px">'
            '<span style="font-size:15px">🗂️</span>'
            '<span style="color:#C7D2FE;font-size:12px;font-weight:800;'
            'text-transform:uppercase;letter-spacing:.6px">Stok Kartı Görüntüle</span></div>',
            unsafe_allow_html=True)
        _skl = get_tum_sku_listesi() or []
        _opts = [r["sku"] for r in _skl]
        _ssec = st.selectbox(
            "SKU ara/seç", ["—"] + _opts, key="stok_karti_sec",
            label_visibility="collapsed",
            help="SKU veya model kodu yaz → seç → kartı aç")
        st.markdown('</div>', unsafe_allow_html=True)
        # Modal içinden "başka SKU'ya geç" isteği (selectbox'tan bağımsız çalışır).
        _gec = st.session_state.pop("_stok_gec_sku", None)
        if _gec:
            st.session_state["_son_acilan_sku"] = _ssec
            from kayranpm.stok_karti import goster as _stok_goster
            _stok_goster(_gec)
        elif _ssec and _ssec != "—" and _ssec != st.session_state.get("_son_acilan_sku"):
            st.session_state["_son_acilan_sku"] = _ssec
            from kayranpm.stok_karti import goster as _stok_goster
            _stok_goster(_ssec)
        # Kart kapatıldıktan sonra aynı kodu tekrar açabilmek için: kapanınca
        # son-açılan takibini sıfırla (böylece aynı SKU'ya tekrar tıklama açar).
        if _ssec == "—":
            st.session_state.pop("_son_acilan_sku", None)

        st.markdown(f"""
        <div style="text-align:center; margin-top:20px; padding-bottom:8px;">
            <div style="color:#475569; font-size:11px;">🕐 {tr_now().strftime('%d.%m.%Y  %H:%M')}</div>
        </div>
        """, unsafe_allow_html=True)
    
    # ════════════════════════════════════════════════════════════════════
    # 1) DASHBOARD
    # ════════════════════════════════════════════════════════════════════
    if sayfa == "📊  Dashboard":
        st.markdown('<div class="baslik">📊 Dashboard</div>', unsafe_allow_html=True)
        st.markdown('<div class="alt-baslik">Stok durumu · Satış performansı · Uyarılar</div>', unsafe_allow_html=True)
        st.markdown('<div class="sayfa-baslik-cizgi"></div>', unsafe_allow_html=True)
    
        # Veri yükle (seçici için SKU listesi gerekli)
        try:
            veri = dashboard_hesapla()
        except Exception as e:
            _log.error("Dashboard veri hatası: %s", e)
            st.error(f"Veri yüklenemedi: {e}")
            st.stop()

        # Filtreler
        _kat_list_d = sorted({tr_kucuk(u.get("kategori")) for u in veri if tr_kucuk(u.get("kategori"))})
        col_f1, col_f2, col_f3 = st.columns([1.6, 1.6, 0.9])
        with col_f1:
            filtre_firma = st.selectbox("Firma Filtresi", ["Tüm Firmalar", "ITOPYA", "HB", "VATAN", "MONDAY", "KANAL", "DİĞER"], format_func=firma_gorunen_ad)
        with col_f2:
            filtre_kat = st.selectbox("Kategori", ["Tüm Kategoriler"] + _kat_list_d, key="dash_kat")
        with col_f3:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("🔄 Yenile", use_container_width=True):
                st.cache_data.clear()
                st.rerun()
    
        # Filtrele
        gosterilecek = []
        for urun in veri:
            # Stoku olan firmalar
            firmali_satirlar = [fd for fd in urun["firma_detay"] if fd.get("stok", 0) > 0]
    
            # Firma filtresi varsa sadece o firmayı göster
            if filtre_firma != "Tüm Firmalar":
                hedef = filtre_firma.replace("İ","I").replace("Ğ","G").replace("Ü","U").replace("Ş","S").replace("Ç","C").replace("Ö","O")
                firmali_satirlar = [fd for fd in firmali_satirlar if hedef in fd["firma"].replace("İ","I").replace("Ğ","G").replace("Ü","U").replace("Ş","S").replace("Ç","C").replace("Ö","O")]
    
            # Kategori filtresi
            if filtre_kat != "Tüm Kategoriler" and tr_kucuk(urun.get("kategori")) != filtre_kat:
                continue
    
            if firmali_satirlar:
                # Firmada stok var — her firma için ayrı satır
                for fd in firmali_satirlar:
                    gosterilecek.append((urun, fd))
            else:
                # Hiçbir firmada stok yok — ürünü yine de göster (sadece G5F depo bilgisiyle)
                if filtre_firma == "Tüm Firmalar":
                    # Boş bir firma satırı oluştur
                    bos_fd = {"firma": "—", "stok": 0, "haftalik_satis": 0,
                              "siparis_uyarisi": False, "muadil_gerekli": False}
                    gosterilecek.append((urun, bos_fd))
    
        # İstatistik kartları
        toplam_sku = len(set(u["sku"] for u in veri))
        acil_urunler = [u for u in veri if u.get("siparis_durum") == "acil"]
        yaklasan_urunler = [u for u in veri if u.get("siparis_durum") == "yaklasıyor"]
        planlama_urunler = [u for u in veri if u.get("siparis_durum") == "planlama"]
        uyari_sayisi = sum(1 for u, fd in gosterilecek if fd.get("siparis_uyarisi"))
        kritik_sayisi = sum(1 for u in veri if u.get("stok_renk") == "kirmizi")
    
        metrik_satiri([
            {"label": "📦 Toplam Ürün", "value": f"{toplam_sku:,}", "renk": "#818CF8"},
            {"label": "🔴 Acil Sipariş", "value": f"{len(acil_urunler):,}", "renk": "#F87171"},
            {"label": "🟠 Yaklaşıyor", "value": f"{len(yaklasan_urunler):,}", "renk": "#FB923C"},
            {"label": "🟡 Planlama", "value": f"{len(planlama_urunler):,}", "renk": "#FBBF24"},
        ])
    
        # ── UYARI PENCERELERİ — shared/ui.py standardı: yan yana kart + iç scroll ──
        from shared.ui import RENK, pencere_css, pencere, pencere_grid, bos_durum
        st.markdown(pencere_css(), unsafe_allow_html=True)

        acil_items_list = []
        for u in acil_urunler:
            gun = u.get('stok_bitis_gun', '?')
            toplam = u.get('toplam_stok', u.get('bizim_stok', 0))
            ad = u['urun_adi']
            ad_kisalt = (ad[:46] + '…') if len(ad) > 46 else ad
            acil_items_list.append(
                f'<div style="display:flex;justify-content:space-between;align-items:center;'
                f'padding:4px 8px;margin:4px 0;border-radius:6px;background:linear-gradient(180deg,#152036,#0F172A);">'
                f'<span style="color:#E2E8F0;font-weight:600;font-size:13px;">⚡ {ad_kisalt}</span>'
                f'<div style="display:flex;gap:12px;flex-shrink:0;margin-left:8px;">'
                f'<span style="color:#94A3B8;font-size:11px;">📦 {toplam:,}</span>'
                f'<span style="color:#F87171;font-size:11px;font-weight:700;">{gun}g</span>'
                f'</div></div>'
            )
        acil_html = "".join(acil_items_list) or bos_durum("Acil sipariş gerektiren ürün yok")

        yak_items_list = []
        for u in yaklasan_urunler:
            gun = u.get('siparis_son_gun', u.get('stok_bitis_gun', '?'))
            ad = u['urun_adi']
            ad_kisalt = (ad[:46] + '…') if len(ad) > 46 else ad
            yak_items_list.append(
                f'<div style="display:flex;justify-content:space-between;align-items:center;'
                f'padding:4px 8px;margin:4px 0;border-radius:6px;background:linear-gradient(180deg,#152036,#0F172A);">'
                f'<span style="color:#E2E8F0;font-size:13px;">📌 {ad_kisalt}</span>'
                f'<span style="color:#FBBF24;font-size:11px;font-weight:600;flex-shrink:0;margin-left:8px;">'
                f'{gun}g içinde</span></div>'
            )
        yak_html = "".join(yak_items_list) or bos_durum("30 gün içinde sipariş gereken ürün yok")

        st.markdown(pencere_grid(
            pencere("🚨 ACİL SİPARİŞ", RENK["kirmizi"], acil_html,
                    rozet=f"{len(acil_urunler)} ürün"),
            pencere("⚠️ 30 GÜN İÇİNDE SİPARİŞ", RENK["amber"], yak_html,
                    rozet=f"{len(yaklasan_urunler)} ürün"),
        ), unsafe_allow_html=True)
    
    
        st.markdown("---")
    
        # ── GÜNCEL KAMPANYA ÖZETİ (kapalı panel + tablo) ──
        try:
            from .database import get_kampanyalar as _get_kmp
            _kmps = _get_kmp(durum="aktif") or []
        except Exception:
            _kmps = []
        import datetime as _kdt2

        def _bts(k):
            try:
                return _kdt2.date.fromisoformat(str(k.get("bitis_tarihi"))[:10])
            except Exception:
                return _kdt2.date.max

        _bg = _kdt2.date.today()
        @st.dialog(f"🎯 Güncel Kampanyalar ({len(_kmps)} aktif)", width="large")
        def _dlg_dash_kampanyalar():
            if not _kmps:
                st.info("Şu an aktif kampanya yok. Kampanya eklemek için **Kampanya Takip** sayfasını kullan.")
            else:
                _tur_say = {}
                for k in _kmps:
                    _t = (k.get("kampanya_turu") or "").strip() or "Belirsiz"
                    _tur_say[_t] = _tur_say.get(_t, 0) + 1
                _dagilim = " · ".join(f"{v} {t}" for t, v in sorted(_tur_say.items(), key=lambda x: -x[1]))
                _yakin = sum(1 for k in _kmps if _bts(k) != _kdt2.date.max and 0 <= (_bts(k) - _bg).days <= 7)
                _ozet = f'<b style="color:#A5B4FC">{len(_kmps)} aktif kampanya</b>'
                if _dagilim:
                    _ozet += f' · <span style="color:#94A3B8">{_dagilim}</span>'
                if _yakin:
                    _ozet += f' · <span style="color:#FB923C">⏳ {_yakin} tanesi 7 gün içinde bitiyor</span>'
                st.markdown(f'<div style="font-size:13px;margin-bottom:8px">{_ozet}</div>', unsafe_allow_html=True)

                _rows_k = []
                for k in sorted(_kmps, key=_bts):
                    _bt = _bts(k)
                    if _bt == _kdt2.date.max:
                        _kalan = ""
                        _durum_k = "—"
                    else:
                        _kg = (_bt - _bg).days
                        _kalan = _kg
                        _durum_k = ("Süresi doldu" if _kg < 0 else
                                    "Bugün bitiyor" if _kg == 0 else
                                    "Yakında bitiyor" if _kg <= 7 else "Aktif")
                    try:
                        _sp = float(k.get("spiff_tl") or 0)
                    except Exception:
                        _sp = 0.0
                    _rows_k.append({
                        "Kampanya": k.get("kampanya_adi", "—"),
                        "Tür": (k.get("kampanya_turu") or "—"),
                        "Firma": k.get("firma", "") or "",
                        "Kategori": k.get("kategori", "") or "",
                        "Başlangıç": str(k.get("baslangic_tarihi", "") or "")[:10],
                        "Bitiş": str(k.get("bitis_tarihi", "") or "")[:10],
                        "Kalan (gün)": _kalan,
                        "Durum": _durum_k,
                        "Spiff ₺": (f"{_sp:,.0f}" if _sp else ""),
                    })
                st.dataframe(pd.DataFrame(_rows_k), hide_index=True, use_container_width=True, height=tablo_h(len(_rows_k)))
        if st.button(f"🎯 Güncel Kampanyalar ({len(_kmps)} aktif)", key="btn_dash_kmp", use_container_width=True):
            _dlg_dash_kampanyalar()

    elif sayfa == "📋  Tüm Ürünler":
        st.markdown('<div class="baslik">📋 Tüm Ürünler</div>', unsafe_allow_html=True)
        st.markdown('<div class="alt-baslik">FOB Price · Cost · Cost Price · Final Cost Price (Paçal) · Stok Dağılımı</div>', unsafe_allow_html=True)
        st.markdown('<div class="sayfa-baslik-cizgi"></div>', unsafe_allow_html=True)
    
        # Ürün verilerini yükle
        try:
            urun_data = tum_urunler_listesi()
        except Exception as e:
            _log.error("Hata: %s", e)
            st.error(f"Veri yüklenemedi: {e}")
            st.stop()
    
        if not urun_data:
            st.info("Henüz ürün yüklenmemiş. 'Veri Yükleme' sekmesinden G5F STOK dosyasını yükleyin.")
            st.stop()
    
        # Özet metrikler
        toplam_stok_degeri = sum(u.get("stok_degeri_fcp", 0) for u in urun_data)
        toplam_satis_degeri = sum(u.get("stok_degeri_satis", 0) for u in urun_data)
        toplam_genel_stok = sum(u.get("toplam_stok", u.get("bizim_stok", 0)) for u in urun_data)
    
        metrik_satiri([
            {"label": "📦 Toplam Ürün", "value": f"{len(urun_data):,}", "renk": "#818CF8"},
            {"label": "🏭 Toplam Stok (Tüm Kanallar)", "value": f"{toplam_genel_stok:,} adet", "renk": "#22D3EE"},
            {"label": "💰 Depo Stok Değeri (Cost)", "value": f"${toplam_stok_degeri:,.0f}", "renk": "#FB923C"},
            {"label": "💵 Depo Stok Değeri (Satış)", "value": f"${toplam_satis_degeri:,.0f}", "renk": "#34D399"},
        ])

        # 🩺 Veri sağlığı (tek satır · eksik alanlar)
        _eksik_kat = sum(1 for u in urun_data if not (u.get("kategori") or "").strip())
        _eksik_mar = sum(1 for u in urun_data if not (u.get("marka") or "").strip())
        _eksik_fiy = sum(1 for u in urun_data if not (u.get("satis_fiyati") or 0))
        _eksik_mal = sum(1 for u in urun_data if not (u.get("final_cost_price") or 0))
        _sg = []
        if _eksik_kat: _sg.append(f'<span style="color:#FBBF24">⚠ {_eksik_kat} kategorisiz</span>')
        if _eksik_mar: _sg.append(f'<span style="color:#FBBF24">⚠ {_eksik_mar} markasız</span>')
        if _eksik_fiy: _sg.append(f'<span style="color:#F87171">⚠ {_eksik_fiy} satış fiyatsız</span>')
        if _eksik_mal: _sg.append(f'<span style="color:#94A3B8">{_eksik_mal} İthalat maliyeti yok</span>')
        if _sg:
            st.markdown('<div style="font-size:13px;color:#94A3B8;margin:8px 0 0px">🩺 <b>Veri sağlığı:</b> '
                        + '  ·  '.join(_sg)
                        + ' <span style="color:#64748B">— Veri Yükleme’deki 🏷️/💲 toplu araçlardan doldurabilirsin</span></div>',
                        unsafe_allow_html=True)
        else:
            st.markdown('<div style="font-size:13px;color:#34D399;margin:8px 0 0px">🩺 <b>Veri sağlığı:</b> ✓ tüm alanlar dolu</div>',
                        unsafe_allow_html=True)

        st.markdown("---")
    
        # SKU arama + ürün seçimi
        st.markdown('<div style="font-size:11px;font-weight:700;color:#94A3B8;letter-spacing:1px;text-transform:uppercase;margin:0px 0 8px;text-align:center;">🔎 Ürün Ara / Seç</div>', unsafe_allow_html=True)
        _sl_tu, col_sec, _sr_tu = st.columns([1.6, 2.0, 1.6])
        with col_sec:
            sku_secenekler = {u['sku']: u['sku'] for u in urun_data}
            secim = st.selectbox("Ürün Seç", list(sku_secenekler.keys()),
                                 label_visibility="collapsed", key="tu_sku")
    
        secilen_sku = sku_secenekler[secim]
        secilen = next(u for u in urun_data if u["sku"] == secilen_sku)
        firma_st = secilen.get("firma_stoklari", {})
    
        # ── Ürün Başlığı (modern kart) ──
        st.markdown(f"""<div style="display:flex;align-items:center;gap:16px;padding:16px 16px;margin:4px 0 16px;background:linear-gradient(135deg,rgba(99,102,241,0.10),rgba(59,130,246,0.04));border:1px solid rgba(99,102,241,0.18);border-radius:16px;">
 <div style="width:48px;height:48px;border-radius:13px;flex-shrink:0;background:linear-gradient(135deg,#6366F1,#7C3AED);display:flex;align-items:center;justify-content:center;font-size:23px;box-shadow:0 6px 18px rgba(99,102,241,0.35);">📦</div>
 <div style="min-width:0;">
 <div style="font-family:'Manrope','Inter',sans-serif;font-size:19px;font-weight:800;color:#F8FAFC;line-height:1.3;letter-spacing:-0.3px;">{secilen["urun_adi"]}</div>
 <div style="margin-top:8px;"><span style="display:inline-block;padding:4px 12px;border-radius:7px;background:rgba(99,102,241,0.15);border:1px solid rgba(99,102,241,0.25);color:#A5B4FC;font-family:'JetBrains Mono',monospace;font-size:11px;font-weight:600;letter-spacing:0.5px;">{secilen["sku"]}</span></div>
 </div></div>""", unsafe_allow_html=True)
    
        bizim_stok = secilen.get("bizim_stok", 0)
        toplam_firma = secilen.get("toplam_firma_stok", 0)
        toplam = secilen.get("toplam_stok", bizim_stok + toplam_firma)
    
        # Stok kartları — ortak tema (renkli sol şeritli kart)
        _stok_cards = [{"label": "G5F DEPO", "value": f"{bizim_stok:,}", "alt": "adet", "renk": "#60A5FA"}]
        for firma, adet in firma_st.items():
            if adet > 0:
                _stok_cards.append({"label": firma, "value": f"{adet:,}", "alt": "adet"})
        st.markdown(
            f'<div style="display:flex; justify-content:space-between; align-items:center; margin:8px 0 8px;">'
            f'<span style="color:#8B97A8; font-size:11px; font-weight:700; text-transform:uppercase; letter-spacing:1.5px;">STOK DAĞILIMI</span>'
            f'<span style="color:#FBBF24; font-size:19px; font-weight:800;">{toplam:,} adet</span>'
            f'</div>',
            unsafe_allow_html=True)
        metrik_satiri(_stok_cards)

        # G5F depo kırılımı (tüm depolar) — bizim deponun depo bazlı dağılımı + genel toplam
        _dk = secilen.get("depo_kirilim") or {}
        if isinstance(_dk, dict) and _dk:
            _dk_toplam = sum(int(v or 0) for v in _dk.values())
            _chips = "".join(
                f'<span style="display:inline-flex;gap:8px;align-items:center;background:rgba(255,255,255,0.04);'
                f'border:1px solid rgba(148,163,184,0.2);border-radius:8px;padding:4px 12px;font-size:13px;color:#CBD5E1">'
                f'{_d} <b style="color:#93C5FD;font-family:monospace">{int(_v or 0):,}</b></span>'
                for _d, _v in sorted(_dk.items(), key=lambda x: -int(x[1] or 0)))
            st.markdown(
                f'<div style="margin:0px 0 12px">'
                f'<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px">'
                f'<span style="color:#8B97A8;font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:1.5px">🏬 G5F DEPO KIRILIMI</span>'
                f'<span style="color:#34D399;font-size:15px;font-weight:800;font-family:monospace">Tüm depolar: {_dk_toplam:,} adet</span></div>'
                f'<div style="display:flex;flex-wrap:wrap;gap:8px">{_chips}</div>'
                f'<div style="color:#64748B;font-size:11px;margin-top:8px">Sipariş önerisinde kullanılan stok '
                f'(Merkez + Happy Life): <b style="color:#93C5FD">{bizim_stok:,}</b></div>'
                f'</div>', unsafe_allow_html=True)
    
        # Fiyat ve karlılık kartı
        fob = secilen.get("fob_price") or 0
        cost = secilen.get("cost") or 0
        cost_price = secilen.get("cost_price") or 0
        fcp = secilen.get("final_cost_price") or 0
        son_fob = secilen.get("son_fob") or 0
        son_fcp = secilen.get("son_final") or 0
        son_tarih = secilen.get("son_tarih") or ""
        ithalat_dosya = secilen.get("ithalat_dosya_sayisi", 0) or 0
        satis = secilen.get("satis_fiyati") or 0
        mal_y = secilen.get("mal_yuzde") or secilen.get("son_mal_yuzde") or 0
    
        if fob > 0 or fcp > 0:
            _son_alt = f"En yeni dosya · {son_tarih}" if son_tarih else "En yeni ithalat dosyası"
            _fiyat_cards = [
                {"label": "PAÇAL FOB", "value": f"${fob:,.2f}", "renk": "#60A5FA",
                 "alt": "Adet-ağırlıklı ortalama"},
                {"label": "SON FOB", "value": f"${son_fob:,.2f}" if son_fob else "—", "renk": "#38BDF8",
                 "alt": _son_alt},
                {"label": f"COST (%{mal_y:.1f})", "value": f"${cost:,.2f}", "renk": "#FB923C"},
                {"label": "⭐ PAÇAL MALİYET", "value": f"${fcp:,.2f}", "renk": "#FBBF24",
                 "alt": "Landed · İthalat"},
                {"label": "SON MALİYET", "value": f"${son_fcp:,.2f}" if son_fcp else "—", "renk": "#FCD34D",
                 "alt": _son_alt},
            ]
            if satis > 0:
                _fiyat_cards.append({"label": "SATIŞ FİYATI", "value": f"${satis:,.2f}", "renk": "#22D3EE"})
                if fcp > 0:
                    _kar = satis - fcp
                    _marj = (_kar / satis * 100) if satis else 0
                    if _kar >= 0:
                        _fiyat_cards.append({"label": "KÂR", "value": f"${_kar:,.2f}",
                                             "renk": "#34D399", "alt": f"Marj %{_marj:.1f} · paçala göre"})
                    else:
                        _fiyat_cards.append({"label": "⚠️ ZARAR", "value": f"${_kar:,.2f}",
                                             "renk": "#F87171", "alt": "Satış, paçal maliyetin altında"})
            st.markdown(
                f'<div style="display:flex;align-items:center;justify-content:space-between;margin:8px 0 8px">'
                f'<div style="color:#90CAF9;font-size:13px;font-weight:700;text-transform:uppercase;letter-spacing:1px">FİYAT ANALİZİ</div>'
                f'<div style="color:#A5D6A7;font-size:11px;font-weight:600;background:rgba(34,197,94,0.12);border:1px solid rgba(34,197,94,0.25);border-radius:6px;padding:4px 8px">🚢 İthalat · {ithalat_dosya} parti</div></div>',
                unsafe_allow_html=True)
            metrik_satiri(_fiyat_cards)
        else:
            st.markdown('<div style="background:rgba(148,163,184,0.06);border:1px dashed rgba(148,163,184,0.25);border-radius:12px;padding:16px;text-align:center;color:#94A3B8;font-size:13px;margin-bottom:16px">🚢 Bu ürün için İthalat maliyet verisi yok — İthalat modülünden bu SKU ile dosya girilince maliyet/paçal otomatik gelecek.</div>', unsafe_allow_html=True)

        # EOL rozeti
        if secilen.get("eol"):
            st.markdown(
                '<div style="background:rgba(248,113,113,0.10);border:1px solid rgba(248,113,113,0.35);'
                'border-radius:10px;padding:8px 16px;margin:4px 0 8px;color:#F87171;font-size:13px;font-weight:700">'
                '⛔ EOL — Bu ürün üretimi/satışı sonlandı olarak işaretli; sipariş önerisine girmez.</div>',
                unsafe_allow_html=True)

        # Müşteri bazlı satış fiyat listesi
        _fl = secilen.get("satis_fiyat_listesi") or {}
        if _fl:
            st.markdown('<div style="color:#90CAF9;font-size:13px;font-weight:700;text-transform:uppercase;letter-spacing:1px;margin:8px 0 8px">🏷️ Müşteri Bazlı Satış Fiyatları</div>', unsafe_allow_html=True)
            metrik_satiri([{"label": _m, "value": f"${_v:,.2f}", "renk": "#22D3EE"}
                           for _m, _v in _fl.items()])

        st.markdown("---")

        # Detayli Gorunum (satis trendi - siparis - yoldaki)
        try:
            _veri_detay = dashboard_hesapla()
            urun = next((u for u in _veri_detay if u["sku"] == secilen_sku), secilen)
        except Exception:
            urun = secilen
        # Üst bilgi kartları
        c1, c2, c3, c4, c5 = st.columns(5)
        toplam_stok_ud = urun.get("toplam_stok", urun.get("bizim_stok", 0))
        stok_bitis = urun.get('stok_bitis_gun')
        stok_bitis_str = f"{stok_bitis} gün" if stok_bitis is not None and stok_bitis != 0 else "Veri yok"
        _risk = urun.get('risk_skor', 0) or 0
        _risk_renk = "#F87171" if _risk >= 70 else ("#FBBF24" if _risk >= 40 else "#34D399")
        metrik_satiri([
            {"label": "📦 Toplam Stok", "value": f"{toplam_stok_ud:,}", "renk": "#818CF8"},
            {"label": "📊 Ort. Hft. Satış", "value": f"{round(urun.get('ortalama_haftalik_satis', 0)):,}", "renk": "#22D3EE"},
            {"label": "⚡ Risk Skoru", "value": f"{_risk}/100", "renk": _risk_renk},
            {"label": "📅 Stok Biter", "value": stok_bitis_str, "renk": "#A78BFA"},
            {"label": "📦 Sipariş Önerisi", "value": f"{urun.get('oneri_miktar',0)} adet", "renk": "#FB923C"},
        ])
    
        # Sipariş durumu banner (yumuşak, tek katman)
        siparis_durum = urun.get("siparis_durum", "veri_yok")
        siparis_mesaj = urun.get("siparis_mesaj", "")
        _oneri_mesaj = urun.get("oneri_mesaj", "")
        _durum_stil = {
            "acil":       ("rgba(239,68,68,0.07)", "rgba(239,68,68,0.55)", "🚨", "#F87171"),
            "yaklasıyor": ("rgba(245,158,11,0.07)", "rgba(245,158,11,0.5)", "⚠️", "#FBBF24"),
            "planlama":   ("rgba(59,130,246,0.07)", "rgba(59,130,246,0.5)", "📋", "#93C5FD"),
            "veri_yok":   ("rgba(34,197,94,0.06)", "rgba(34,197,94,0.45)", "✅", "#4ADE80"),
        }
        _bg, _brd, _ik, _tc = _durum_stil.get(siparis_durum, _durum_stil["veri_yok"])
        _detay = f' <span style="color:#94A3B8;font-weight:400;">· {_oneri_mesaj}</span>' if _oneri_mesaj else ""
        st.markdown(
            f'<div style="background:{_bg};border-left:3px solid {_brd};border-radius:8px;padding:12px 16px;margin:8px 0;font-size:13px;">'
            f'<span style="color:{_tc};font-weight:700;">{_ik} {siparis_mesaj}</span>{_detay}</div>',
            unsafe_allow_html=True
        )
    
        st.markdown("---")
    
        # Tüm ürünler özet tablosu
        @st.dialog("📊 Tüm Ürünler Özet — filtrele, sırala, incele", width="large")
        def _dlg_urunler_ozet():
            _kat_oz = sorted({tr_kucuk(u.get("kategori")) for u in urun_data if tr_kucuk(u.get("kategori"))})
            _mar_oz = sorted({(u.get("marka") or "").strip() for u in urun_data if (u.get("marka") or "").strip()})
            _ozf0, _ozf1, _ozf2, _ozf3, _ozf4 = st.columns([1.6, 1.2, 1.2, 1.5, 1.0])
            _sku_ad_map_oz = {}
            for _uu in urun_data:
                _ss = str(_uu.get("sku", "") or "").strip()
                if _ss and _ss not in _sku_ad_map_oz:
                    _sku_ad_map_oz[_ss] = _uu.get("urun_adi", "") or ""
            _sku_secenek_oz = ["Tümü"] + [f"{s} — {a}" if a else s for s, a in sorted(_sku_ad_map_oz.items())]
            with _ozf0:
                f_ara_oz = st.selectbox(f"🔍 SKU / Ürün ({len(_sku_ad_map_oz)} model · yazarak ara)",
                                        _sku_secenek_oz, key="oz_ara")
            with _ozf1:
                f_kat_oz = st.selectbox("Kategori", ["Tümü"] + _kat_oz, key="oz_kat")
            with _ozf2:
                f_mar_oz = st.selectbox("Marka", ["Tümü"] + _mar_oz, key="oz_mar")
            with _ozf3:
                f_sira_oz = st.selectbox("Sırala", ["Stok Yaşı (gün)", "Net Kâr ($)", "Net Marj (%)", "Maliyet %", "Toplam Stok", "Satış ($)", "FOB ($)", "Risk Skoru", "SKU (A-Z)"], key="oz_sira")
            with _ozf4:
                f_yon_oz = st.selectbox("Yön", ["Azalan", "Artan"], key="oz_yon")
            _sadece_zarar = st.checkbox("⚠️ Sadece zararına satılanlar (satış < paçal maliyet)", value=False, key="oz_zarar")
            _sira_map_oz = {"Stok Yaşı (gün)": "_stok_yas", "Net Kâr ($)": "Net Kar ($)", "Net Marj (%)": "Net Marj (%)", "Maliyet %": "Maliyet %", "Toplam Stok": "Toplam", "Satış ($)": "Satış ($)", "FOB ($)": "FOB ($)", "Risk Skoru": "_risk", "SKU (A-Z)": "SKU"}
            rows_oz = []
            for u in urun_data:
                fs = u.get("firma_stoklari", {})
                satis = u.get('satis_fiyati') or 0
                ith_var = (u.get("ithalat_dosya_sayisi", 0) or 0) > 0
                fcp = (u.get('final_cost_price') or 0) if ith_var else 0
                fob = (u.get('fob_price') or 0) if ith_var else 0
                son_fob = (u.get('son_fob') or 0) if ith_var else 0
                son_fcp = (u.get('son_final') or 0) if ith_var else 0
                maliyet_yuzde = ((fcp / fob - 1) * 100) if (fob > 0 and fcp > 0) else None
                net_kar = (satis - fcp) if (satis > 0 and fcp > 0) else None
                net_marj = ((net_kar / satis) * 100) if (net_kar is not None and satis > 0) else None
                rows_oz.append({
                    "SKU": u["sku"],
                    "Ürün Adı": u.get("urun_adi", ""),
                    "Kategori": u.get("kategori", ""),
                    "Marka": u.get("marka", ""),
                    "_stok_yas": int(u.get("stok_gun", 0) or 0),
                    "_stok_renk": u.get("stok_renk", "yok"),
                    "_risk": float(u.get("risk_skor", 0) or 0),
                    "G5F Depo": int(u.get("bizim_stok", 0) or 0),
                    "ITOPYA": int(fs.get("ITOPYA", 0) or 0),
                    "HB": int(fs.get("HB", 0) or 0),
                    "VATAN": int(fs.get("VATAN", 0) or 0),
                    "MONDAY": int(fs.get("MONDAY", 0) or 0),
                    "KANAL": int(fs.get("KANAL", 0) or 0),
                    "Toplam": int(u.get("toplam_stok", u.get("bizim_stok", 0)) or 0),
                    "FOB ($)": (float(fob) if ith_var else None),
                    "Son FOB ($)": (float(son_fob) if (ith_var and son_fob) else None),
                    "Maliyet %": float(maliyet_yuzde) if maliyet_yuzde is not None else None,
                    "Final Cost ($)": (float(fcp) if ith_var else None),
                    "Son Maliyet ($)": (float(son_fcp) if (ith_var and son_fcp) else None),
                    "Satış ($)": float(satis or 0),
                    "Net Marj (%)": float(net_marj) if net_marj is not None else None,
                    "Net Kar ($)": float(net_kar) if net_kar is not None else None,
                })
            _toplam_oz = len(rows_oz)
            _zarar_say = sum(1 for r in rows_oz
                             if r.get("Net Kar ($)") is not None and r.get("Net Kar ($)") < 0)
            # Filtre + sıralama uygula
            if f_ara_oz and f_ara_oz != "Tümü":
                _sel_sku = f_ara_oz.split(" — ")[0].strip()
                rows_oz = [r for r in rows_oz if str(r.get("SKU") or "").strip() == _sel_sku]
            if f_kat_oz != "Tümü":
                rows_oz = [r for r in rows_oz if tr_kucuk(r.get("Kategori")) == f_kat_oz]
            if f_mar_oz != "Tümü":
                rows_oz = [r for r in rows_oz if (r.get("Marka") or "").strip() == f_mar_oz]
            if _sadece_zarar:
                rows_oz = [r for r in rows_oz
                           if r.get("Net Kar ($)") is not None and r.get("Net Kar ($)") < 0]
            _sk_oz = _sira_map_oz.get(f_sira_oz, "_stok_yas")
            _rev_oz = (f_yon_oz == "Azalan")
            if _sk_oz == "SKU":
                rows_oz.sort(key=lambda r: str(r.get("SKU") or "").lower(), reverse=_rev_oz)
            else:
                _dolu = [r for r in rows_oz if r.get(_sk_oz) not in (None, "")]
                _bos = [r for r in rows_oz if r.get(_sk_oz) in (None, "")]
                _dolu.sort(key=lambda r: float(r.get(_sk_oz) or 0), reverse=_rev_oz)
                rows_oz = _dolu + _bos

            st.caption(f"📦 {len(rows_oz)} / {_toplam_oz} ürün gösteriliyor"
                       + (f"  ·  ⚠️ {_zarar_say} ürün zararına satılıyor (satış < paçal)" if _zarar_say else ""))
            df_oz = pd.DataFrame(rows_oz)
            if df_oz.empty:
                st.info("Henüz ürün yok.")
            else:
                def _fmt_para(v):
                    try:
                        return f"${float(v):,.2f}" if v not in (None, "") and float(v) != 0 else "—"
                    except Exception:
                        return "—"
                def _fmt_int(v):
                    try:
                        return f"{int(v):,}" if v not in (None, "") else "0"
                    except Exception:
                        return "0"
                def _fmt_pct(v):
                    try:
                        return f"%{float(v):.1f}" if v not in (None, "") else "—"
                    except Exception:
                        return "—"
                def _stok_cls(v):
                    try:
                        return "c-num" if (v and int(v) > 0) else "c-muted"
                    except Exception:
                        return "c-muted"

                satir_html = ""
                for r in rows_oz:
                    ad = str(r.get("Ürün Adı", "") or "")
                    ad_kisa = ad if len(ad) <= 46 else ad[:45] + "…"
                    ad_title = ad.replace(chr(34), "&quot;")
                    nk = r.get("Net Kar ($)")
                    if nk is None:
                        nk_cls, nk_txt = "c-muted", "—"
                    elif nk > 0:
                        nk_cls, nk_txt = "c-pos", f"${nk:,.2f}"
                    elif nk < 0:
                        nk_cls, nk_txt = "c-neg", f"${nk:,.2f}"
                    else:
                        nk_cls, nk_txt = "c-num", f"${nk:,.2f}"
                    nm = r.get("Net Marj (%)")
                    if nm is None:
                        nm_cls, nm_txt = "c-muted", "—"
                    elif nm > 0:
                        nm_cls, nm_txt = "c-pos", f"%{nm:.1f}"
                    elif nm < 0:
                        nm_cls, nm_txt = "c-neg", f"%{nm:.1f}"
                    else:
                        nm_cls, nm_txt = "c-num", f"%{nm:.1f}"
                    tot = r.get("Toplam") or 0
                    tot_cls = "c-tot" if tot else "c-muted"
                    fcp_raw = r.get("Final Cost ($)")
                    fcp = fcp_raw or 0
                    fcp_cls = "c-fcp" if fcp else "c-muted"
                    fob_raw = r.get("FOB ($)")
                    fob_v = fob_raw or 0
                    fob_cls = "c-num" if fob_v else "c-muted"
                    son_fob_raw = r.get("Son FOB ($)")
                    son_fob_v = son_fob_raw or 0
                    son_fob_cls = "c-num" if son_fob_v else "c-muted"
                    son_fcp_raw = r.get("Son Maliyet ($)")
                    son_fcp_v = son_fcp_raw or 0
                    son_fcp_cls = "c-fcp" if son_fcp_v else "c-muted"
                    satis = r.get("Satış ($)") or 0
                    satis_cls = "c-money" if satis else "c-muted"
                    mal = r.get("Maliyet %")
                    mal_cls = "c-mal" if (mal is not None) else "c-muted"
                    _kanallar = []
                    for _kn, _kl in (("ITOPYA", "IT"), ("HB", "HB"), ("VATAN", "VT"), ("MONDAY", "MN"), ("KANAL", "KN")):
                        _kv = r.get(_kn) or 0
                        if _kv:
                            _kanallar.append(f"{_kl}:{int(_kv)}")
                    kanal_str = " · ".join(_kanallar) if _kanallar else "—"
                    kanal_title = kanal_str.replace(chr(34), "&quot;")
                    _loss = (nk is not None and nk < 0)
                    _tr_attr = ' style="background:rgba(239,68,68,0.07);box-shadow:inset 3px 0 0 #EF4444"' if _loss else ""
                    _yas_v = r.get("_stok_yas")
                    _yas_renk = r.get("_stok_renk", "yok")
                    _yas_renk_map = {"yesil": "#4ADE80", "sari": "#FBBF24", "turuncu": "#FB923C", "kirmizi": "#F87171"}
                    if _yas_renk in _yas_renk_map:
                        _yas_col = _yas_renk_map[_yas_renk]
                        _yas_txt = f"{int(_yas_v or 0)}g"
                    else:
                        _yas_col = "#475569"
                        _yas_txt = "—"
                    satir_html += (
                        f"<tr{_tr_attr}>"
                        f'<td class="c-sku">{r.get("SKU","")}</td>'
                        f'<td class="c-name" title="{ad_title}">{ad_kisa}</td>'
                        f'<td class="c-kat">{r.get("Kategori","")}</td>'
                        f'<td style="text-align:right;font-family:\'JetBrains Mono\',monospace;color:{_yas_col};font-weight:600">{_yas_txt}</td>'
                        f'<td class="{_stok_cls(r.get("G5F Depo"))}">{_fmt_int(r.get("G5F Depo"))}</td>'
                        f'<td class="c-kanal" title="{kanal_title}">{kanal_str}</td>'
                        f'<td class="{tot_cls}">{_fmt_int(tot)}</td>'
                        f'<td class="{fob_cls}">{_fmt_para(fob_v) if fob_raw is not None else "—"}</td>'
                        f'<td class="{son_fob_cls}">{_fmt_para(son_fob_v) if son_fob_raw is not None else "—"}</td>'
                        f'<td class="{mal_cls}">{_fmt_pct(mal)}</td>'
                        f'<td class="{fcp_cls}">{_fmt_para(fcp) if fcp_raw is not None else "—"}</td>'
                        f'<td class="{son_fcp_cls}">{_fmt_para(son_fcp_v) if son_fcp_raw is not None else "—"}</td>'
                        f'<td class="{satis_cls}">{_fmt_para(satis)}</td>'
                        f'<td class="{nm_cls}">{nm_txt}</td>'
                        f'<td class="{nk_cls}">{nk_txt}</td>'
                        "</tr>"
                    )
                css = (
                    "<style>"
                    ".urun-wrap{overflow-x:auto;border-radius:14px;box-shadow:0 2px 16px rgba(0,0,0,0.25);margin-top:8px}"
                    ".urun-tbl{width:100%;table-layout:fixed;border-collapse:collapse;font-family:Inter,sans-serif}"
                    ".urun-tbl thead tr{background:linear-gradient(135deg,#1E293B,#0F172A)}"
                    ".urun-tbl thead th{padding:8px 8px;color:#CBD5E1;font-size:11px;font-weight:700;letter-spacing:.2px;text-transform:uppercase;white-space:normal;line-height:1.2;text-align:center;vertical-align:middle}"
                    ".urun-tbl thead th.l{text-align:left}"
                    ".urun-tbl tbody{background:#131C35}"
                    ".urun-tbl td{padding:8px 8px;font-size:11px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}"
                    ".urun-tbl tbody tr{border-bottom:1px solid rgba(255,255,255,0.05)}"
                    ".urun-tbl tbody tr:hover{background:rgba(99,102,241,0.06)}"
                    ".c-sku{color:#E2E8F0;font-family:'JetBrains Mono',monospace;font-weight:600;white-space:nowrap}"
                    ".c-name{color:#CBD5E1;max-width:300px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}"
                    ".c-kat{color:#94A3B8;font-size:11px}"
                    ".c-kanal{text-align:left;color:#94A3B8;font-family:'JetBrains Mono',monospace;font-size:11px;max-width:175px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}"
                    ".c-num{text-align:right;color:#CBD5E1;font-family:'JetBrains Mono',monospace}"
                    ".c-dim{text-align:right;color:#94A3B8;font-family:'JetBrains Mono',monospace}"
                    ".c-money{text-align:right;color:#E2E8F0;font-family:'JetBrains Mono',monospace;font-weight:600}"
                    ".c-mal{text-align:right;color:#C4B5FD;font-family:'JetBrains Mono',monospace}"
                    ".c-pos{text-align:right;color:#4ADE80;font-weight:700;font-family:'JetBrains Mono',monospace}"
                    ".c-neg{text-align:right;color:#F87171;font-weight:700;font-family:'JetBrains Mono',monospace}"
                    ".c-fcp{text-align:right;color:#FBBF24;font-weight:600;font-family:'JetBrains Mono',monospace}"
                    ".c-tot{text-align:right;color:#93C5FD;font-weight:700;font-family:'JetBrains Mono',monospace}"
                    ".c-muted{text-align:right;color:#475569;font-family:'JetBrains Mono',monospace}"
                    "</style>"
                )
                thead = (
                    '<div class="urun-wrap"><table class="urun-tbl">'
                    '<colgroup>'
                    '<col style="width:6%"><col style="width:12%"><col style="width:6%"><col style="width:5%">'
                    '<col style="width:5%"><col style="width:9%"><col style="width:5%">'
                    '<col style="width:6%"><col style="width:6%"><col style="width:6%">'
                    '<col style="width:7%"><col style="width:7%"><col style="width:6%">'
                    '<col style="width:6%"><col style="width:6%">'
                    '</colgroup>'
                    '<thead><tr>'
                    '<th class="l">SKU</th><th class="l">Ürün Adı</th><th class="l">Kategori</th>'
                    '<th>📅 Stok Yaşı</th>'
                    '<th>G5F</th><th class="l">Kanal Stok</th><th>Toplam</th>'
                    "<th>Paçal FOB</th><th>Son FOB</th><th>Maliyet %</th>"
                    "<th>⭐ Paçal Maliyet</th><th>Son Maliyet</th><th>Satış</th>"
                    "<th>📊 Net Marj %</th><th>💰 Net Kâr $</th>"
                    "</tr></thead><tbody>"
                )
                st.html(css + thead + satir_html + "</tbody></table></div>")
            st.caption("💡 FOB/Maliyet iki türlü: Paçal = adet-ağırlıklı ortalama (kâr/marj buna göre) · Son = en yeni ithalat dosyası · Net Kâr $ = Satış − Paçal Maliyet")

            # ── 📤 Rapor Al — Excel / PDF (ekrandaki filtreye göre) ──
            if rows_oz:
                _rapor_meta = (f"Kategori: {f_kat_oz} · Marka: {f_mar_oz} · "
                               f"Sıra: {f_sira_oz} ({f_yon_oz})"
                               + (" · Sadece zararına" if _sadece_zarar else ""))
                with st.expander(f"📤 Rapor Al — Excel / PDF  ·  {len(rows_oz)} ürün (filtreye göre)", expanded=False):
                    st.caption(f"Aktif filtre → {_rapor_meta}")
                    _rr1, _rr2 = st.columns(2)
                    with _rr1:
                        if st.button("📊 Excel Oluştur", use_container_width=True, type="primary", key="oz_rapor_excel"):
                            from .rapor import tum_urunler_excel
                            import tempfile
                            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as _tmp:
                                _tp = _tmp.name
                            _ok, _msg = tum_urunler_excel(rows_oz, _tp, _rapor_meta)
                            if _ok:
                                with open(_tp, "rb") as _f:
                                    st.download_button("⬇️ Excel İndir", _f.read(),
                                        f"Tum_Urunler_{tr_now().strftime('%Y%m%d_%H%M')}.xlsx",
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                        use_container_width=True, key="oz_rapor_excel_dl")
                                os.unlink(_tp)
                            else:
                                st.error(_msg)
                    with _rr2:
                        if st.button("📑 PDF Oluştur", use_container_width=True, key="oz_rapor_pdf"):
                            from .rapor import tum_urunler_pdf
                            import tempfile
                            with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as _tmp:
                                _tp = _tmp.name
                            _ok, _msg = tum_urunler_pdf(rows_oz, _tp, _rapor_meta)
                            if _ok:
                                with open(_tp, "rb") as _f:
                                    st.download_button("⬇️ PDF İndir", _f.read(),
                                        f"Tum_Urunler_{tr_now().strftime('%Y%m%d_%H%M')}.pdf",
                                        mime="application/pdf",
                                        use_container_width=True, key="oz_rapor_pdf_dl")
                                os.unlink(_tp)
                            else:
                                st.error(_msg)
        if st.button("📊 Tüm Ürünler Özet — filtrele, sırala, incele", key="btn_urun_ozet", use_container_width=True):
            _dlg_urunler_ozet()

        @st.dialog("✏️ Ürün Düzenle", width="large")
        def _dlg_urun_duzenle():
            st.caption("Açılır listeden ürünü seç, alanları düzenle, **Kaydet**'e bas.")
            _sec_list = {f'{u["sku"]} — {(u.get("urun_adi") or "")[:50]}': u["sku"] for u in urun_data}
            if _sec_list:
                _sec_label = st.selectbox("Düzenlenecek ürün", list(_sec_list.keys()), key="urun_duzen_sec")
                _sec_sku = _sec_list[_sec_label]
                _u = next((x for x in urun_data if x["sku"] == _sec_sku), {})
                with st.form("urun_duzen_form"):
                    fc1, fc2, fc3, fc4 = st.columns([3, 1.5, 1.2, 1])
                    with fc1:
                        d_ad = st.text_input("Ürün Adı", value=_u.get("urun_adi", "") or "")
                    with fc2:
                        d_kat = st.text_input("Kategori", value=_u.get("kategori", "") or "")
                    with fc3:
                        d_satis = st.number_input("Genel Satış ($)", value=float(_u.get("satis_fiyati", 0) or 0), min_value=0.0, step=1.0, format="%.2f",
                                                  help="Genel/liste fiyatı — kâr marjı ve dashboard bundan hesaplanır.")
                    with fc4:
                        d_stok = st.number_input("G5F Depo", value=int(_u.get("bizim_stok", 0) or 0), min_value=0, step=1)

                    # Müşteri bazlı satış fiyat listesi (ana müşteriler hazır + satır ekleyerek yeni müşteri)
                    from .database import ANA_MUSTERILER as _ANA_MUST
                    _mevcut_liste = _u.get("satis_fiyat_listesi") or {}
                    _musteriler = list(_ANA_MUST) + [m for m in _mevcut_liste if m not in _ANA_MUST]
                    _liste_df = pd.DataFrame(
                        [{"Müşteri": m, "Fiyat ($)": float(_mevcut_liste.get(m, 0) or 0)} for m in _musteriler]
                    )
                    st.markdown("**🏷️ Satış Fiyat Listesi (müşteri bazlı)** — ana müşteriler hazır gelir; "
                                "müşteriye özel fiyat gir. **Satır ekleyip** yeni müşteri de yazabilirsin (0 bıraktığın satır kaydedilmez).")
                    _liste_edit = st.data_editor(
                        _liste_df, num_rows="dynamic", use_container_width=True, key="urun_fiyat_liste",
                        column_config={
                            "Müşteri": st.column_config.TextColumn("Müşteri", required=False),
                            "Fiyat ($)": st.column_config.NumberColumn("Fiyat ($)", min_value=0.0, step=1.0, format="%.2f"),
                        },
                    )
                    d_eol = st.checkbox(
                        "⛔ EOL — üretimi/satışı sonlandı (bu ürüne sipariş ÖNERİLMESİN)",
                        value=bool(_u.get("eol")))

                    if st.form_submit_button("💾 Kaydet", type="primary", use_container_width=True):
                        from .database import upsert_urun as _upsert_urun
                        # Fiyat listesini editörden topla
                        _yeni_liste = {}
                        try:
                            for _, _r in _liste_edit.iterrows():
                                _m = str(_r.get("Müşteri", "") or "").strip()
                                _fy = float(_r.get("Fiyat ($)", 0) or 0)
                                if _m and _fy != 0:
                                    _yeni_liste[_m] = _fy
                        except Exception:
                            _yeni_liste = {}
                        try:
                            _upsert_urun(
                                _sec_sku, d_ad.strip(), d_kat.strip(),
                                _u.get("marka", "") or "", float(d_satis or 0),
                                float(_u.get("alis_fiyati", 0) or 0), float(_u.get("hedef_kar_marji", 0) or 0),
                                _u.get("ozellikler", "") or "", int(d_stok or 0),
                                int(_u.get("trendyol_stok", 0) or 0),
                                satis_fiyat_listesi=_yeni_liste, eol=d_eol,
                            )
                            st.cache_data.clear()
                            st.success(f"✅ {_sec_sku} güncellendi.")
                            st.rerun()
                        except Exception as _e:
                            st.error(f"Kaydedilemedi: {_e}")

                # ── Yanlış girilen stok kartını sil ──
                st.markdown("<div style='height:1px;background:rgba(255,255,255,0.06);margin:12px 0 8px'></div>",
                            unsafe_allow_html=True)
                _sil_onay = st.checkbox(f"⚠️ '{_sec_sku}' stok kartını kalıcı olarak sil", key="urun_sil_onay")
                if st.button("🗑️ Stok Kartını Sil", key="urun_sil_btn",
                             disabled=not _sil_onay, use_container_width=True):
                    from .database import sil_urun as _sil_urun
                    try:
                        _sil_urun(_sec_sku)
                        st.cache_data.clear()
                        st.success(f"🗑️ {_sec_sku} stok kartı silindi.")
                        st.rerun()
                    except Exception as _e:
                        st.error(f"Silinemedi: {_e}")
        if st.button("✏️ Ürün Düzenle", key="btn_urun_duz", use_container_width=True):
            _dlg_urun_duzenle()
    
    
    
        st.markdown("---")
        st.markdown('<div style="font-size:13px;font-weight:700;color:#94A3B8;letter-spacing:1px;text-transform:uppercase;margin:8px 0 8px;display:flex;align-items:center;gap:8px"><span style="width:4px;height:14px;border-radius:3px;background:linear-gradient(180deg,#38BDF8,#0EA5E9);display:inline-block"></span>🚢 Yoldaki Ürün Durumu</div>', unsafe_allow_html=True)
        _yr = urun.get("yol_renk", "yok")
        _ymik = urun.get("yol_miktar", 0)
        _ymsg = urun.get("yol_mesaj", "")
        _yol_map = {"yesil": ("rgba(34,197,94,0.10)", "#4ADE80", "🟢", f"{_ymik} adet yolda · {_ymsg}"), "sari": ("rgba(245,158,11,0.10)", "#FBBF24", "🟡", f"{_ymik} adet yolda · {_ymsg}"), "kirmizi": ("rgba(239,68,68,0.10)", "#F87171", "🔴", _ymsg)}
        _yb, _yc, _yi, _yt = _yol_map.get(_yr, ("rgba(148,163,184,0.08)", "#94A3B8", "⚪", "Yolda ürün kaydı bulunmuyor."))
        st.markdown(f'<div style="background:{_yb};border-left:3px solid {_yc};border-radius:7px;padding:8px 12px;font-size:13px;color:{_yc};font-weight:600;display:inline-block">{_yi} {_yt}</div>', unsafe_allow_html=True)


    elif sayfa == "📈  Müşteri Satışları":
        from shared.tarih import hizli_tarih_araligi
        st.markdown('<div class="baslik">📈 Müşteri Haftalık Satışları</div>', unsafe_allow_html=True)
        st.markdown('<div class="alt-baslik">Müşteri (firma) bazında haftalık satış geçmişi · müşteri · tarih aralığı · ürün/SKU filtresi</div>', unsafe_allow_html=True)
        st.markdown('<div class="sayfa-baslik-cizgi"></div>', unsafe_allow_html=True)
        _bas, _bit = hizli_tarih_araligi("mhs", varsayilan="Son 90 gün")
        _mc1, _mc2 = st.columns([1, 1.4])
        _firmalar = ["Tümü"] + get_firma_listesi()
        _f = _mc1.selectbox("Müşteri", _firmalar, key="mhs_firma", format_func=firma_gorunen_ad)
        _sku_ara = _mc2.text_input("Ürün / SKU ara", key="mhs_sku",
                                   placeholder="SKU veya ürün adı ile filtrele…")
        _rows = get_musteri_haftalik_satis(_bas, _bit, _f, _sku_ara)
        if not _rows:
            st.info("Bu filtrelerle haftalık satış kaydı bulunamadı.")
        else:
            _df = pd.DataFrame([{
                "Tarih": str(r.get("yukleme_tarihi", ""))[:10],
                "Müşteri": firma_gorunen_ad(r.get("firma", "")),
                "SKU": r.get("sku", ""),
                "Ürün": (r.get("urun_adi", "") or "")[:45],
                "Haftalık Satış": int(r.get("haftalik_satis", 0) or 0),
                "Stok": int(r.get("stok_miktari", 0) or 0),
            } for r in _rows])
            metrik_satiri([
                {"label": "📈 Toplam Haftalık Satış", "value": f"{int(_df['Haftalık Satış'].sum()):,}", "renk": "#818CF8"},
                {"label": "🧾 Kayıt", "value": f"{len(_df):,}", "renk": "#22D3EE"},
                {"label": "👥 Müşteri Sayısı", "value": f"{int(_df['Müşteri'].nunique()):,}", "renk": "#34D399"},
            ])
            @st.dialog("📊 Müşteri bazında toplam satış", width="large")
            def _dlg_musteri_toplam():
                _grp = (_df.groupby("Müşteri")["Haftalık Satış"].sum()
                        .sort_values(ascending=False).reset_index())
                st.dataframe(_grp, hide_index=True, use_container_width=True, height=tablo_h(len(_grp)))
            if st.button("📊 Müşteri bazında toplam satış", key="btn_mus_top", use_container_width=True):
                _dlg_musteri_toplam()
            st.dataframe(_df, hide_index=True, use_container_width=True, height=460)
            st.download_button("⬇️ CSV indir",
                               _df.to_csv(index=False).encode("utf-8-sig"),
                               "musteri_haftalik_satis.csv", "text/csv", key="mhs_csv")

        st.markdown("---")
        @st.dialog("📤 Müşteri Satış / Stok Verisi Yükle", width="large")
        def _dlg_musteri_yukle():
            st.caption("Firma adını tam yazabilirsin (HEPSİBURADA, EERA, D-MARKET…) — sistem otomatik tanır. "
                       "Önerilen: birleşik tek sayfa. Eski çok-sekmeli dosya da desteklenir.")
            st.markdown("**📊 Firma Stok + Satış · Birleşik Tek Sayfa**")
            st.caption("Sütunlar: FİRMA ADI · KATEGORİ · MARKA · STOK KODU · STOK ADI · STOK · STOK-MAĞAZA · SATIŞ · SATIŞ-MAĞAZA")
            _dosya_b = st.file_uploader("Birleşik Excel'i Seç", type=["xlsx", "xls"], key="mhs_birlesik_dosya")
            if _dosya_b and st.button("⬆️ Firma Stok + Satış Yükle", type="primary", use_container_width=True, key="mhs_birlesik_btn"):
                import tempfile
                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as _tb:
                    _tb.write(_dosya_b.read())
                    _tbp = _tb.name
                _ok, _msg = excel_yukle_firma_birlesik(_tbp)
                os.unlink(_tbp)
                st.cache_data.clear()
                (st.success if _ok else st.error)(_msg)
            st.markdown("---")
            st.markdown("**📅 Haftalık STOK + SATIŞ · Firma Başına 2 Sekme (portal formatları)**")
            st.caption("Sekmeler: `ITOPYA STOK` · `ITOPYA SATIŞ` · `VATAN STOK` · `VATAN SATIŞ` · "
                       "`HEPSİBURADA STOK/SATIŞ` · `MONDAY STOK/SATIŞ`. Her firmanın **kendi portal başlıkları** "
                       "olduğu gibi kalır (STOKKODU/Kod/Sku/Malzeme/Ürün Kodu…). Satışlar SKU ile stokun yanına "
                       "bağlanıp tek özet oluşturulur. **Kategori dosyada gerekmez** — bizim ürün kartından eşlenir; "
                       "boş bırakılan kategori kolonları hata vermez.")
            _dosya_hss = st.file_uploader("Haftalık STOK+SATIŞ Excel'i Seç", type=["xlsx", "xls"], key="mhs_hss_dosya")
            if _dosya_hss and st.button("⬆️ Haftalık STOK+SATIŞ Yükle", type="primary",
                                        use_container_width=True, key="mhs_hss_btn"):
                import tempfile
                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as _tb2:
                    _tb2.write(_dosya_hss.read())
                    _tbp2 = _tb2.name
                _ok2, _msg2 = excel_yukle_haftalik_stok_satis(_tbp2)
                os.unlink(_tbp2)
                st.cache_data.clear()
                (st.success if _ok2 else st.error)(_msg2)
            st.markdown("---")
            st.markdown("**📥 Çok-Sekmeli Haftalık Dosya**")
            st.caption("Her firma ayrı sekmede (ITOPYA, HB, VATAN, MONDAY, KANAL, DIGER) · SKU · Ürün Adı · Stok · Haftalık Satış")
            _dosya_h = st.file_uploader("Excel Dosyasını Seç", type=["xlsx", "xls"], key="mhs_haftalik_dosya")
            if _dosya_h and st.button("⬆️ Tüm Veriyi Yükle", type="primary", use_container_width=True, key="mhs_haftalik_btn"):
                import tempfile
                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as _th:
                    _th.write(_dosya_h.read())
                    _thp = _th.name
                _ok2, _msg2 = excel_yukle_firma_stoklari(_thp)
                os.unlink(_thp)
                st.cache_data.clear()
                (st.success if _ok2 else st.error)(_msg2)
        if st.button("📤 Müşteri Satış / Stok Verisi Yükle", key="btn_mus_yuk", use_container_width=True):
            _dlg_musteri_yukle()

    elif sayfa == "🎯  Kampanya Takip":
        st.markdown('<div class="baslik">🎯 Kampanya Takip</div>', unsafe_allow_html=True)
        st.markdown('<div class="alt-baslik">Firma desteği · Net kar · Kampanya performansı</div>', unsafe_allow_html=True)
        st.markdown('<div class="sayfa-baslik-cizgi"></div>', unsafe_allow_html=True)
    
        FIRMA_LISTESI_K = ["ITOPYA", "HB", "VATAN", "MONDAY", "KANAL", "DİĞER"]
        KAMPANYA_TURLERI = ["(Seçilmedi)", "Sellout", "Rebate", "Marketing", "Spiff"]
    
        # Ürün verisi (paçal maliyet için)
        try:
            urun_data_k = tum_urunler_listesi()
            urun_dict_k = {u["sku"]: u for u in urun_data_k}
            sku_listesi_k = {f"{u['sku']} — {u['urun_adi']}": u['sku'] for u in urun_data_k}
        except Exception:
            urun_data_k = []
            urun_dict_k = {}
            sku_listesi_k = {}
    
        # ── Durum sekmesi (üstte, sekme gibi) ──────────────────────────
        _kt_durum = st.radio("Durum", ["📢 Aktif Kampanyalar", "📁 Geçmiş Kampanyalar"],
                             horizontal=True, key="kt_durum", label_visibility="collapsed")

        # ── Filtreler tek kompakt panelde (kapalı başlar → özet + giriş üste çıkar) ──
        _kt_kat_list = sorted({tr_kucuk(u.get("kategori")) for u in urun_data_k if tr_kucuk(u.get("kategori"))})
        _yil_set = set()
        for _kk in get_kampanyalar():
            _b = str(_kk.get("baslangic_tarihi") or "")[:4]
            if _b.isdigit():
                _yil_set.add(_b)
        _yil_opts = ["Tümü"] + sorted(_yil_set, reverse=True)
        _KT_AYLAR = ["Tümü", "Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
                     "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"]
        with st.expander("🔍 Filtrele — Müşteri · Kategori · Dönem", expanded=False):
            _kt_firma = st.radio("Müşteri", ["Tümü", "HB", "VATAN", "ITOPYA", "DİĞER"],
                                 horizontal=True, key="kt_firma",
                                 format_func=firma_gorunen_ad)
            _fk1, _fk2, _fk3, _fk4, _fk5 = st.columns(5)
            _kt_kat = _fk1.selectbox("Kategori", ["Tümü"] + _kt_kat_list, key="kt_kat")
            _kt_yil = _fk2.selectbox("Yıl", _yil_opts, key="kt_yil")
            _kt_ceyrek = _fk3.selectbox("Çeyrek", ["Tümü", "Q1", "Q2", "Q3", "Q4"], key="kt_ceyrek")
            _kt_ay = _fk4.selectbox("Ay", _KT_AYLAR, key="kt_ay")
            _kt_sira = _fk5.selectbox("Sırala", ["Yeniden eskiye", "Eskiden yeniye"], key="kt_sira")

        def _kt_firma_filtre(liste):
            if _kt_firma == "Tümü":
                return liste
            _ana = {"HB", "VATAN", "ITOPYA"}
            if _kt_firma == "DİĞER":
                return [k for k in liste if str(k.get("firma", "")).strip().upper() not in _ana]
            return [k for k in liste if str(k.get("firma", "")).strip().upper() == _kt_firma]

        def _kt_ceyrek_of(tarih_str):
            try:
                _m = int(str(tarih_str)[5:7])
                return f"Q{(_m - 1)//3 + 1}"
            except Exception:
                return ""

        def _kt_ay_of(tarih_str):
            try:
                return int(str(tarih_str)[5:7])
            except Exception:
                return 0

        def _kt_uygula(liste):
            """Firma + kategori + yıl + çeyrek + ay filtresi, ardından tarihe göre sıralama."""
            liste = _kt_firma_filtre(liste)
            if _kt_kat != "Tümü":
                liste = [k for k in liste if tr_kucuk(k.get("kategori")) == _kt_kat]
            if _kt_yil != "Tümü":
                liste = [k for k in liste if str(k.get("baslangic_tarihi") or "")[:4] == _kt_yil]
            if _kt_ceyrek != "Tümü":
                liste = [k for k in liste if _kt_ceyrek_of(k.get("baslangic_tarihi")) == _kt_ceyrek]
            if _kt_ay != "Tümü":
                _ay_no = _KT_AYLAR.index(_kt_ay)
                liste = [k for k in liste if _kt_ay_of(k.get("baslangic_tarihi")) == _ay_no]
            return sorted(liste, key=lambda k: str(k.get("baslangic_tarihi") or ""),
                          reverse=(_kt_sira == "Yeniden eskiye"))


        st.markdown('<div style="height:8px"></div>', unsafe_allow_html=True)

        # Tüm kampanya ürünleri TEK sorguda (N+1 önleme)
        _ku_map = {}
        for _r in get_tum_kampanya_urunler():
            _ku_map.setdefault(_r["kampanya_id"], []).append(_r)

        # ─────────────────────────────────────────────────────────────────
        # TAB 1: AKTİF KAMPANYALAR — üstte özet + yatay kompakt giriş + tam genişlik liste
        # ─────────────────────────────────────────────────────────────────
        if _kt_durum == "📢 Aktif Kampanyalar":
            # 📊 Özet kartlar (tam genişlik, üstte)
            _tum_kmp = _kt_uygula(get_kampanyalar())
            _s_toplam = len(_tum_kmp)
            _bugun_iso = str(tr_today())
            _s_aktif = sum(1 for k in _tum_kmp
                           if str(k.get("baslangic_tarihi") or "") <= _bugun_iso <= str(k.get("bitis_tarihi") or "9999"))
            _s_destek = 0.0
            _s_net = 0.0
            for _k in _tum_kmp:
                for _x in _ku_map.get(_k.get("id"), []):
                    _s = float(_x.get("satis_fiyati") or 0)
                    _fd = float(_x.get("birim_firma_destek") or 0)
                    _ed = float(_x.get("birim_ek_destek") or 0)
                    _ad = int(_x.get("satilan_adet") or 0)
                    _p = float((urun_dict_k.get(_x.get("sku"), {}) or {}).get("pacal_maliyet") or 0)
                    _s_destek += (_fd + _ed) * _ad
                    if _s > 0 and _p > 0:
                        _s_net += ((_s - _p) - (_fd + _ed)) * _ad
            metrik_satiri([
                {"label": "🎯 Toplam", "value": f"{_s_toplam:,}", "renk": "#818CF8"},
                {"label": "📢 Aktif", "value": f"{_s_aktif:,}", "renk": "#34D399"},
                {"label": "💰 Destek", "value": f"${_s_destek:,.0f}", "renk": "#FBBF24"},
                {"label": "📈 Net Kâr", "value": f"${_s_net:,.0f}",
                 "renk": "#34D399" if _s_net >= 0 else "#F87171"},
            ])

            # ➕ Yeni Kampanya — AÇILIR PENCERE (dialog). Ana sayfa uzamaz.
            @st.dialog("➕ Yeni Kampanya Oluştur", width="large")
            def _yeni_kampanya_dialog():
                st.caption("Temel bilgileri gir ve oluştur. Tür · Spiff · Notlar · ürünler, "
                           "kampanya oluştuktan sonra detay penceresinden eklenir.")
                with st.form("yeni_kampanya_form", clear_on_submit=False):
                    k_adi = st.text_input("Kampanya Adı *", placeholder="örn: Hepsiburada Mart Kampanyası")
                    _dc1, _dc2 = st.columns(2)
                    k_firma = _dc1.selectbox("Firma *", ["(Firma seç)"] + FIRMA_LISTESI_K,
                                             format_func=lambda f: f if str(f).startswith("(") else firma_gorunen_ad(f))
                    k_kat = _dc2.selectbox("Kategori", ["(Genel / Karışık)"] + _kt_kat_list)
                    _dc3, _dc4 = st.columns(2)
                    k_bas = _dc3.date_input("Başlangıç Tarihi *", value=tr_today())
                    k_bit = _dc4.date_input("Bitiş Tarihi *", value=tr_today())
                    if st.form_submit_button("🚀 Kampanya Oluştur", type="primary", use_container_width=True):
                        if not k_adi.strip():
                            st.error("Kampanya adı zorunludur.")
                        elif str(k_firma).startswith("("):
                            st.error("Firma seçimi zorunludur.")
                        elif str(k_bit) < str(k_bas):
                            st.error("Bitiş tarihi başlangıç tarihinden önce olamaz.")
                        else:
                            _k_kat_val = "" if str(k_kat).startswith("(") else k_kat
                            _hata = None
                            yeni_id = None
                            try:
                                yeni_id = ekle_kampanya(k_adi.strip(), k_firma, str(k_bas), str(k_bit), "", _k_kat_val,
                                                        kampanya_turu="", spiff_tl=0, spiff_kur=0, spiff_fatura=False)
                            except Exception as _e:
                                _hata = str(_e)
                            if yeni_id:
                                st.cache_data.clear()
                                st.session_state["_kamp_detay_sec"] = yeni_id
                                st.success(f"✅ '{k_adi.strip()}' oluşturuldu.")
                                st.rerun()
                            elif _hata:
                                st.error(f"Kampanya oluşturulamadı — kayıt hatası: {_hata}")
                            else:
                                st.error("Kampanya kaydedilemedi — veritabanı kayıt döndürmedi "
                                         "(muhtemelen 'kampanyalar' tablosunda izin/kolon sorunu).")

            _yk1, _yk2 = st.columns([1, 4])
            if _yk1.button("➕ Yeni Kampanya", type="primary", use_container_width=True, key="yeni_kmp_ac"):
                _yeni_kampanya_dialog()
            _yk2.caption("Yeni kampanya için butona bas — açılır pencerede oluştur. "
                         "Listeden bir kampanyaya tıklayınca detayı açılır pencerede görünür.")
            st.markdown('<div style="height:6px"></div>', unsafe_allow_html=True)
    

            # ── Excel şablonundan YENİ kampanya oluştur + ürünleri ekle (tek dosya) ──
            @st.dialog("📥 Excel Şablonundan Kampanya Oluştur (kampanya + ürünler tek dosyada)", width="large")
            def _dlg_kmp_excel():
                _KMP_TAM_KOL = ["FİRMA ADI", "MARKA", "KATEGORİ", "STOK KODU", "STOK ADI", "BARKOD",
                                "FİYAT", "REBATE", "SELLOUT", "EK SELLOUT", "SPIFF", "NET FİYAT",
                                "KAMPANYA ADI", "KAMPANYA TÜRÜ", "BAŞLANGIÇ TARİHİ", "BİTİŞ TARİHİ"]
                st.caption("Tek şablonla YENİ kampanya oluşturur + ürünleri ekler. **Firma · Kampanya Türü · "
                           "Başlangıç/Bitiş Tarihi · Ek Sellout dahil tüm bilgiler dosyadan otomatik gelir.** "
                           "Eşleme: FİYAT→satış · SELLOUT→firma desteği · EK SELLOUT→ek destek (paçal SKU'dan otomatik).")

                def _knrm(_s):
                    _s = str(_s).strip()
                    for _a, _b in (("İ", "i"), ("I", "i"), ("ı", "i"), ("Ş", "s"), ("ş", "s"),
                                   ("Ğ", "g"), ("ğ", "g"), ("Ü", "u"), ("ü", "u"),
                                   ("Ö", "o"), ("ö", "o"), ("Ç", "c"), ("ç", "c")):
                        _s = _s.replace(_a, _b)
                    return _s.lower()

                _tbuf = BytesIO()
                # Açılır liste kaynakları
                try:
                    from satis.database import get_kanallar as _get_cariler
                    _cariler = [c for c in (_get_cariler() or []) if str(c).strip()]
                except Exception:
                    _cariler = []
                _markalar = sorted({(u.get("marka") or "").strip() for u in urun_data_k
                                    if (u.get("marka") or "").strip()})
                _turler = [t for t in KAMPANYA_TURLERI if not str(t).startswith("(")]
                _katlar = list(_kt_kat_list)
                try:
                    import openpyxl
                    from openpyxl.worksheet.datavalidation import DataValidation
                    _wb = openpyxl.Workbook()
                    _ws = _wb.active
                    _ws.title = "Kampanya"
                    _ws.append(_KMP_TAM_KOL)
                    _lst = _wb.create_sheet("LISTELER")

                    def _lst_yaz(_ci, _bas, _veri):
                        _lst.cell(row=1, column=_ci, value=_bas)
                        for _i, _v in enumerate(_veri, start=2):
                            _lst.cell(row=_i, column=_ci, value=str(_v))
                        return len(_veri)
                    _nf = _lst_yaz(1, "FIRMA", _cariler)
                    _nk = _lst_yaz(2, "KATEGORI", _katlar)
                    _nm = _lst_yaz(3, "MARKA", _markalar)
                    _nt = _lst_yaz(4, "TUR", _turler)
                    _lst.sheet_state = "hidden"

                    def _ekle_dv(_kol, _lcol, _n):
                        if _n <= 0:
                            return
                        _dv = DataValidation(type="list",
                                             formula1=f"=LISTELER!${_lcol}$2:${_lcol}${_n + 1}",
                                             allow_blank=True)
                        _dv.add(f"{_kol}2:{_kol}2000")
                        _ws.add_data_validation(_dv)
                    _ekle_dv("A", "A", _nf)   # FİRMA ADI
                    _ekle_dv("B", "C", _nm)   # MARKA (yeni düzende 2. kolon)
                    _ekle_dv("C", "B", _nk)   # KATEGORİ (yeni düzende 3. kolon)
                    _ekle_dv("N", "D", _nt)   # KAMPANYA TÜRÜ (SPIFF eklendi → N kolonu)
                    _wb.save(_tbuf)
                except Exception:
                    # Açılır liste kurulamazsa düz şablona düş
                    _tbuf = BytesIO()
                    with pd.ExcelWriter(_tbuf, engine="openpyxl") as _w:
                        pd.DataFrame(columns=_KMP_TAM_KOL).to_excel(_w, index=False, sheet_name="Kampanya")
                st.caption(f"⬇️ Şablonda Firma Adı ({len(_cariler)}) · Kategori ({len(_katlar)}) · "
                           f"Marka ({len(_markalar)}) · Kampanya Türü açılır listeden seçilir.")
                st.download_button("⬇️ Kampanya şablonu indir", _tbuf.getvalue(),
                                   "KAMPANYA_OLUSTUR_SABLONU.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   key="kmp_olustur_sablon_dl")
                _kfile = st.file_uploader("Doldurulmuş kampanya şablonu", type=["xlsx", "xls"],
                                          key="kmp_olustur_up")
                if _kfile is not None:
                    try:
                        _kdf = pd.read_excel(_kfile)
                        _kolmap = {_knrm(c): c for c in _kdf.columns}

                        def _col(*_names):
                            _ns = [_knrm(n) for n in _names]
                            for _n in _ns:  # önce tam eşleşme (fiyat vs net fiyat, sellout vs ek sellout)
                                if _n in _kolmap:
                                    return _kolmap[_n]
                            for _n in _ns:  # sonra alt-dizge
                                for _k, _actual in _kolmap.items():
                                    if _n in _k:
                                        return _actual
                            return None

                        def _gv(_row, *_names):
                            _c = _col(*_names)
                            if _c is not None and _c in _row and pd.notna(_row[_c]):
                                return _row[_c]
                            return None

                        def _knum(_v):
                            try:
                                if _v is None or (isinstance(_v, float) and pd.isna(_v)):
                                    return 0.0
                                return float(str(_v).replace(",", ".").replace(" ", ""))
                            except Exception:
                                return 0.0

                        def _tarih(_v):
                            try:
                                if _v is None or (isinstance(_v, float) and pd.isna(_v)):
                                    return None
                                _t = pd.to_datetime(_v, errors="coerce", dayfirst=True)
                                return _t.date() if pd.notna(_t) else None
                            except Exception:
                                return None

                        st.dataframe(_kdf.head(15), use_container_width=True, height=200)

                        _xl_ad = _xl_firma = _xl_kat = _xl_tur = ""
                        _xl_bas = _xl_bit = None
                        _xl_spiff = 0.0
                        for _, _r in _kdf.iterrows():
                            _xl_ad = _xl_ad or str(_gv(_r, "kampanya adi") or "").strip()
                            _xl_firma = _xl_firma or str(_gv(_r, "firma adi", "firma") or "").strip()
                            _xl_kat = _xl_kat or str(_gv(_r, "kategori") or "").strip()
                            _xl_tur = _xl_tur or str(_gv(_r, "kampanya turu", "tur") or "").strip()
                            _xl_bas = _xl_bas or _tarih(_gv(_r, "baslangic"))
                            _xl_bit = _xl_bit or _tarih(_gv(_r, "bitis"))
                            if not _xl_spiff:
                                _xl_spiff = _knum(_gv(_r, "spiff"))
                            if _xl_ad and _xl_firma and _xl_tur and _xl_bas and _xl_bit and _xl_spiff:
                                break

                        # Kategori: mevcut bir kategoriyle yalnızca harf farkı varsa onu kullan (KASA/kasa mükerrerini önle)
                        _kat_final = _xl_kat
                        for _ek in _kt_kat_list:
                            if _ek.strip().lower() == _xl_kat.lower():
                                _kat_final = _ek
                                break

                        _tur_idx = 0
                        for _ti, _t in enumerate(KAMPANYA_TURLERI):
                            if _t.lower() == _xl_tur.lower():
                                _tur_idx = _ti
                                break

                        # Firma: şablondaki TAM cari adını FIRMA_LISTESI_K koduna (HB/VATAN/ITOPYA...) eşle.
                        # Kampanya 'firma' alanı KOD saklar; ekranda firma_gorunen_ad ile tam ad gösterilir.
                        def _firma_koda(_ad):
                            _adn = _knrm(_ad)
                            if not _adn:
                                return ""
                            for _c in FIRMA_LISTESI_K:
                                try:
                                    _tam = _knrm(firma_gorunen_ad(_c))
                                except Exception:
                                    _tam = ""
                                if _tam and (_tam == _adn or _tam in _adn or _adn in _tam):
                                    return _c
                            for _anahtar, _c in (("d-market", "HB"), ("dmarket", "HB"), ("hepsiburada", "HB"),
                                                 ("eera", "ITOPYA"), ("itopya", "ITOPYA"),
                                                 ("vatan", "VATAN"), ("monday", "MONDAY")):
                                if _anahtar in _adn:
                                    return _c
                            return ""
                        _fk = _firma_koda(_xl_firma)
                        _firma_idx = FIRMA_LISTESI_K.index(_fk) if _fk in FIRMA_LISTESI_K else 0
                        _of1, _of2 = st.columns(2)
                        _o_ad = _of1.text_input("Kampanya Adı *", value=_xl_ad, key="kmp_o_ad")
                        _o_firma = _of2.selectbox("Firma *", FIRMA_LISTESI_K, index=_firma_idx,
                                                  format_func=firma_gorunen_ad, key="kmp_o_firma")
                        if _xl_firma:
                            st.caption(f"🏢 Dosyadaki firma: **{_xl_firma[:44]}** → "
                                       + (f"**{_o_firma}** koduyla eşlendi." if _fk
                                          else "otomatik eşlenemedi — yukarıdan doğru firmayı seç."))
                        _of3, _of4, _of5 = st.columns(3)
                        _o_turu = _of3.selectbox("Kampanya Türü", KAMPANYA_TURLERI, index=_tur_idx, key="kmp_o_turu")
                        _o_spiff_tl, _o_spiff_kur = 0.0, 0.0
                        if str(_o_turu).lower() == "spiff" or _xl_spiff > 0:
                            _sp1, _sp2 = st.columns(2)
                            _o_spiff_tl = _sp1.number_input("Spiff Net (₺TL)", min_value=0.0, step=100.0,
                                                            value=float(_xl_spiff or 0), format="%.2f",
                                                            key="kmp_o_spiff_tl",
                                                            help="Dosyadaki SPIFF kolonundan geldi; düzenleyebilirsin.")
                            _o_spiff_kur = _sp2.number_input("Kur (₺/$ tahmini)", min_value=0.0, step=0.1,
                                                             value=None, placeholder="örn. 40",
                                                             format="%.4f", key="kmp_o_spiff_kur") or 0.0
                            if _o_spiff_tl and _o_spiff_kur:
                                st.caption(f"≈ ${(_o_spiff_tl / _o_spiff_kur):,.2f} USD spiff maliyeti (tahmini)")
                        _o_bas = _of4.date_input("Başlangıç Tarihi *", value=(_xl_bas or tr_today()), key="kmp_o_bas")
                        _o_bit = _of5.date_input("Bitiş Tarihi *", value=(_xl_bit or tr_today()), key="kmp_o_bit")
                        st.caption(f"🏷️ Kategori (dosyadan): **{_kat_final or '—'}**"
                                   + ("" if (not _kat_final or _kat_final == _xl_kat)
                                      else f"  · mevcut '{_kat_final}' ile eşleştirildi (yeni mükerrer açılmadı)"))

                        _urun_satir = []
                        for _, _r in _kdf.iterrows():
                            _sku = str(_gv(_r, "stok kodu", "sku") or "").strip()
                            if not _sku or _sku.lower() == "nan":
                                continue
                            _bilgi = urun_dict_k.get(_sku, {})
                            _uad = (str(_gv(_r, "stok adi", "urun adi", "urun") or "").strip()
                                    or _bilgi.get("urun_adi", _sku))
                            _urun_satir.append({
                                "sku": _sku, "urun_adi": _uad,
                                "pacal": _bilgi.get("final_cost_price", 0) or 0,
                                "satis": _knum(_gv(_r, "fiyat", "satis")),
                                "fd": _knum(_gv(_r, "sellout")),
                                "ed": _knum(_gv(_r, "ek sellout")),
                            })
                        st.caption(f"Şablonda **{len(_urun_satir)}** geçerli ürün satırı bulundu.")

                        if st.button("🚀 Şablondan Kampanya Oluştur ve Ürünleri Ekle", type="primary",
                                     use_container_width=True, key="kmp_o_olustur",
                                     disabled=(not _o_ad.strip() or not _o_firma.strip() or not _urun_satir)):
                            _ohata, _oyid = None, None
                            try:
                                _oyid = ekle_kampanya(
                                    _o_ad.strip(), _o_firma.strip(), str(_o_bas), str(_o_bit), "",
                                    _kat_final,
                                    kampanya_turu=("" if str(_o_turu).startswith("(") else _o_turu),
                                    spiff_tl=(_o_spiff_tl or 0), spiff_kur=(_o_spiff_kur or 0))
                            except Exception as _e:
                                _ohata = str(_e)
                            if _oyid:
                                _on = 0
                                for _u in _urun_satir:
                                    try:
                                        ekle_kampanya_urun(_oyid, _u["sku"], _u["urun_adi"], _u["pacal"],
                                                           _u["satis"], _u["fd"], _u["ed"], "")
                                        _on += 1
                                    except Exception:
                                        pass
                                st.cache_data.clear()
                                st.success(f"✅ '{_o_ad.strip()}' kampanyası oluşturuldu ve {_on} ürün eklendi "
                                           f"(Firma: {_o_firma.strip()} · Tür: {_o_turu} · {_o_bas}→{_o_bit}).")
                                st.rerun()
                            elif _ohata:
                                st.error(f"Kampanya oluşturulamadı — {_ohata}")
                            else:
                                st.error("Kampanya oluşturulamadı (tablo izni/kolon olabilir).")
                    except Exception as _e:
                        st.error(f"Excel okunamadı: {_e}")
            if st.button("📥 Excel Şablonundan Kampanya Oluştur (kampanya + ürünler tek dosyada)", key="btn_kmp_exc", use_container_width=True):
                _dlg_kmp_excel()

            # Aktif kampanyaları listele
            aktif_kampanyalar = _kt_uygula(get_kampanyalar(durum="aktif"))
            if not aktif_kampanyalar:
                st.info("Aktif kampanya yok. Yukarıdan yeni kampanya oluşturabilirsiniz.")
            else:
                # ── KAMPANYA PANOSU (üstte hızlı özet) ──
                _bugun = tr_today()
                def _dmy(x):
                    try:
                        return date.fromisoformat(str(x)[:10]).strftime("%d-%m-%Y")
                    except Exception:
                        return str(x or "")[:10]
                _pano = []
                for _k in aktif_kampanyalar:
                    _ku = _ku_map.get(_k["id"], [])
                    _sat = 0
                    _net = 0.0
                    for _x in _ku:
                        _p = _x.get("pacal_maliyet") or 0
                        _s = _x.get("satis_fiyati") or 0
                        _fd = _x.get("birim_firma_destek") or 0
                        _ed = _x.get("birim_ek_destek") or 0
                        _ad = _x.get("satilan_adet") or 0
                        if _s > 0 and _p > 0:
                            _net += ((_s - _p) - (_fd + _ed)) * _ad
                        _sat += _ad
                    try:
                        _bit = date.fromisoformat(_k["bitis_tarihi"]) if _k.get("bitis_tarihi") else None
                        _kalan = (_bit - _bugun).days if _bit else None
                    except Exception:
                        _kalan = None
                    if _kalan is None:
                        _kalan_txt = "—"
                    elif _kalan < 0:
                        _kalan_txt = f"🔴 {abs(_kalan)}g geçti"
                    elif _kalan < 7:
                        _kalan_txt = f"🔴 {_kalan}g kaldı"
                    elif _kalan < 30:
                        _kalan_txt = f"🟠 {_kalan}g kaldı"
                    else:
                        _kalan_txt = f"🟢 {_kalan}g kaldı"
                    _pano.append({
                        "Kampanya": _k["kampanya_adi"],
                        "Firma": _k["firma"],
                        "Tarih": f'{_dmy(_k.get("baslangic_tarihi"))} → {_dmy(_k.get("bitis_tarihi"))}',
                        "Kalan": _kalan_txt,
                        "Ürün": len(_ku),
                        "Satılan": _sat,
                        "Net Kâr ($)": round(_net, 2),
                    })
                st.markdown('<div style="font-size:13px;font-weight:700;color:#E2E8F0;margin:4px 0 8px;">📊 Kampanya Panosu — bir kampanyaya tıkla, detayı açılır pencerede görünsün</div>', unsafe_allow_html=True)
                _pano_evt = st.dataframe(
                    pd.DataFrame(_pano),
                    hide_index=True,
                    use_container_width=True,
                    on_select="rerun",
                    selection_mode="single-row",
                    key="kamp_pano_df",
                    column_config={"Net Kâr ($)": st.column_config.NumberColumn(format="$%.2f")},
                )
                _pano_sel = list(_pano_evt.selection.rows)
                if _pano_sel:
                    _yeni_sec = aktif_kampanyalar[_pano_sel[0]]["id"]
                    # Yalnızca farklı bir kampanya seçilince dialog'u yeniden aç
                    if st.session_state.get("_kamp_detay_sec") != _yeni_sec:
                        st.session_state["_kamp_detay_sec"] = _yeni_sec
                        st.session_state["_kamp_detay_ac"] = True
                _kamp_secili_id = st.session_state.get("_kamp_detay_sec")
                if not (_kamp_secili_id and any(_kk["id"] == _kamp_secili_id for _kk in aktif_kampanyalar)):
                    st.caption("👆 Panodan bir kampanyaya tıklayınca detayı açılır pencerede görünür.")
                _kamp_list = [_kk for _kk in aktif_kampanyalar if _kk['id'] == _kamp_secili_id]
                if _kamp_list:
                    @st.dialog("📢 Kampanya Detayı", width="large")
                    def _kampanya_detay_dialog():
                        kamp = _kamp_list[0]
                        kid = kamp['id']
                        k_urunler = _ku_map.get(kid, [])
                        # Kampanya başlık kartı
                        st.markdown(f"""
                        <div style="background:rgba(255,255,255,0.05); border-radius:12px; padding:16px 20px;
                                    margin:12px 0 8px; border-left:5px solid #42A5F5;">
                          <div style="display:flex; justify-content:space-between; align-items:center;">
                            <div>
                              <span style="color:#90CAF9; font-size:19px; font-weight:800;">📢 {kamp["kampanya_adi"]}</span>
                              <span style="background:#1F4E79; color:#90CAF9; padding:0px 8px; border-radius:10px;
                                          font-size:13px; margin-left:8px; font-weight:600;">{kamp["firma"]}</span>
                              <span style="background:#1B5E20; color:#A5D6A7; padding:0px 8px; border-radius:10px;
                                          font-size:13px; margin-left:8px;">● AKTİF</span>
                            </div>
                            <span style="color:#90A4AE; font-size:13px;">
                              {kamp["baslangic_tarihi"]} → {kamp["bitis_tarihi"]}
                            </span>
                          </div>
                        </div>
                        """, unsafe_allow_html=True)
    
                        # Kampanya düzenleme
                        with st.expander(f"✏️ Kampanya Bilgilerini Düzenle — {kamp['kampanya_adi']}"):
                            # ── Kampanya Genel Bilgileri ──
                            st.markdown("**📋 Kampanya Bilgileri**")
                            with st.form(f"duzenle_kamp_{kid}"):
                                dk1, dk2 = st.columns(2)
                                with dk1:
                                    dk_adi = st.text_input("Kampanya Adı", value=kamp["kampanya_adi"], key=f"dk_adi_{kid}")
                                    dk_firma = st.selectbox("Firma", FIRMA_LISTESI_K,
                                        index=FIRMA_LISTESI_K.index(kamp["firma"]) if kamp["firma"] in FIRMA_LISTESI_K else 0,
                                        format_func=firma_gorunen_ad,
                                        key=f"dk_firma_{kid}")
                                    _dk_kat_cur = (kamp.get("kategori") or "").strip()
                                    _dk_opts = ["(Genel / Karışık)"] + _kt_kat_list
                                    if _dk_kat_cur and _dk_kat_cur not in _dk_opts:
                                        _dk_opts = ["(Genel / Karışık)", _dk_kat_cur] + _kt_kat_list
                                    dk_kat = st.selectbox("Kategori", _dk_opts,
                                        index=_dk_opts.index(_dk_kat_cur) if _dk_kat_cur in _dk_opts else 0,
                                        key=f"dk_kat_{kid}")
                                with dk2:
                                    dk_bas = st.date_input("Başlangıç Tarihi", value=date.fromisoformat(kamp["baslangic_tarihi"]) if kamp["baslangic_tarihi"] else tr_today(), key=f"dk_bas_{kid}")
                                    dk_bit = st.date_input("Bitiş Tarihi", value=date.fromisoformat(kamp["bitis_tarihi"]) if kamp["bitis_tarihi"] else tr_today(), key=f"dk_bit_{kid}")
                                dk_not = st.text_area("Notlar", value=kamp.get("notlar","") or "", key=f"dk_not_{kid}")
                                _dk_turu_cur = (kamp.get("kampanya_turu") or "").strip()
                                _dk_turu_opts = KAMPANYA_TURLERI + ([_dk_turu_cur] if _dk_turu_cur and _dk_turu_cur not in KAMPANYA_TURLERI else [])
                                dk_turu = st.selectbox("Kampanya Türü", _dk_turu_opts,
                                    index=_dk_turu_opts.index(_dk_turu_cur) if _dk_turu_cur in _dk_turu_opts else 0,
                                    key=f"dk_turu_{kid}")
                                st.markdown('<div style="color:#94A3B8;font-size:11px;font-weight:700;margin:8px 0 0px">🎟️ Spiff — TL net + kur. Fatura gelince <b>kuru</b> güncelle, USD/kâr otomatik düzelir.</div>', unsafe_allow_html=True)
                                dsp1, dsp2, dsp3 = st.columns([1.2, 1, 1])
                                dk_spiff_tl = dsp1.number_input("Spiff Net (₺TL)", min_value=0.0, step=100.0, format="%.2f",
                                    value=float(kamp.get("spiff_tl") or 0) or None, placeholder="0", key=f"dk_spiff_tl_{kid}")
                                dk_spiff_kur = dsp2.number_input("Kur (₺/$)", min_value=0.0, step=0.1, format="%.4f",
                                    value=float(kamp.get("spiff_kur") or 0) or None, placeholder="örn. 40", key=f"dk_spiff_kur_{kid}")
                                dk_spiff_fatura = dsp3.checkbox("Fatura geldi (kur kesin)",
                                    value=bool(kamp.get("spiff_fatura")), key=f"dk_spiff_fatura_{kid}")
                                if dk_spiff_tl and dk_spiff_kur:
                                    st.caption(f"≈ ${(dk_spiff_tl / dk_spiff_kur):,.2f} USD spiff maliyeti {'(fatura/kesin)' if dk_spiff_fatura else '(tahmini)'}")
                                dc1_k, dc2_k, dc3_k = st.columns(3)
                                with dc1_k:
                                    if st.form_submit_button("💾 Kampanyayı Güncelle", use_container_width=True, type="primary"):
                                        _dk_kat_val = "" if str(dk_kat).startswith("(") else dk_kat
                                        _dk_turu_val = "" if str(dk_turu).startswith("(") else dk_turu
                                        guncelle_kampanya(kid, dk_adi, dk_firma, str(dk_bas), str(dk_bit), dk_not,
                                                          kategori=_dk_kat_val, kampanya_turu=_dk_turu_val,
                                                          spiff_tl=(dk_spiff_tl or 0), spiff_kur=(dk_spiff_kur or 0),
                                                          spiff_fatura=dk_spiff_fatura)
                                        st.cache_data.clear()
                                        st.toast("✅ Kampanya güncellendi!")
                                        st.session_state['_kamp_detay_ac'] = True
                                        st.rerun()
                                with dc2_k:
                                    kapat_flag = st.form_submit_button("🔒 Kampanyayı Kapat", use_container_width=True)
                                with dc3_k:
                                    if st.form_submit_button("🗑️ Kampanyayı Sil", use_container_width=True):
                                        sil_kampanya(kid)
                                        st.cache_data.clear()
                                        st.warning("Kampanya silindi.")
                                        st.session_state.pop("_kamp_detay_sec", None)
                                        st.session_state["_kamp_detay_ac"] = False
                                        st.rerun()
    
                            # Kampanya kapat — adet sor (kapat_flag form içinden geliyor, güvenli)
                            if locals().get('kapat_flag', False):
                                st.session_state[f"kapat_onay_{kid}"] = True
    
                            if st.session_state.get(f"kapat_onay_{kid}"):
                                st.markdown("---")
                                st.markdown("**🔒 Kampanyayı Kapat — Satış Adetlerini Gir**")
                                st.caption("Kampanya kapanmadan önce her ürün için satılan adeti girin.")
                                with st.form(f"kapat_form_{kid}", clear_on_submit=True):
                                    adet_girisleri = {}
                                    for ku in k_urunler:
                                        kf1, kf2, kf3 = st.columns([2, 2, 1])
                                        with kf1:
                                            st.markdown(f'<span style="color:#90CAF9; font-size:13px; font-weight:600;">{ku["sku"]}</span>'
                                                       f'<span style="color:#546E7A; font-size:11px;"> — {ku.get("urun_adi","")[:40]}</span>',
                                                       unsafe_allow_html=True)
                                        with kf2:
                                            adet_girisleri[ku["id"]] = st.number_input(
                                                "Satılan Adet",
                                                value=int(ku.get("satilan_adet",0) or 0),
                                                min_value=0, step=1,
                                                key=f"kapat_adet_{kid}_{ku['id']}",
                                                label_visibility="collapsed"
                                            )
                                        with kf3:
                                            st.markdown(f'<span style="color:#78909C; font-size:11px;">mevcut: {ku.get("satilan_adet",0)}</span>', unsafe_allow_html=True)
    
                                    kk1, kk2 = st.columns(2)
                                    with kk1:
                                        if st.form_submit_button("✅ Kaydet ve Kapat", type="primary", use_container_width=True):
                                            for ku_id_k, adet_k in adet_girisleri.items():
                                                ku_bilgi = next((x for x in k_urunler if x["id"] == ku_id_k), {})
                                                guncelle_kampanya_urun(ku_id_k,
                                                    ku_bilgi.get("satis_fiyati",0),
                                                    ku_bilgi.get("birim_firma_destek",0),
                                                    ku_bilgi.get("birim_ek_destek",0),
                                                    adet_k,
                                                    ku_bilgi.get("notlar",""))
                                            kapat_kampanya(kid)
                                            st.session_state.pop(f"kapat_onay_{kid}", None)
                                            st.cache_data.clear()
                                            st.toast("🔒 Kampanya kapatıldı ve satış adetleri kaydedildi!")
                                            st.session_state.pop("_kamp_detay_sec", None)
                                            st.session_state["_kamp_detay_ac"] = False
                                            st.rerun()
                                    with kk2:
                                        if st.form_submit_button("İptal", use_container_width=True):
                                            st.session_state.pop(f"kapat_onay_{kid}", None)
                                            st.session_state['_kamp_detay_ac'] = True
                                            st.rerun()
    
                            # ── Ürün Bazında Düzenleme ──
                            if k_urunler:
                                st.markdown("---")
                                st.markdown("**🛍️ Ürün Bilgilerini Düzenle**")
                                st.caption("Her ürün için SKU, satış fiyatı, birim destek ve satılan adeti ayrı ayrı düzenleyebilirsiniz.")
    
                                for ku in k_urunler:
                                    ku_id = ku["id"]
                                    pacal_ku = ku.get("pacal_maliyet") or 0
                                    # Paçal 0 ise güncel değer
                                    if pacal_ku == 0:
                                        pacal_ku = urun_dict_k.get(ku["sku"], {}).get("final_cost_price", 0)
    
                                    with st.form(f"urun_gun_{kid}_{ku_id}", clear_on_submit=False):
                                        st.markdown(f'<div style="background:rgba(255,255,255,0.04); border-radius:8px; padding:8px 16px; margin-bottom:8px;">'
                                                    f'<span style="color:#90CAF9; font-weight:700; font-size:13px;">📦 {ku.get("urun_adi","")}</span>'
                                                    f'<span style="color:#546E7A; font-size:11px; margin-left:8px;">SKU: {ku["sku"]}</span>'
                                                    f'</div>', unsafe_allow_html=True)
    
                                        ug1, ug2, ug3, ug4 = st.columns(4)
                                        with ug1:
                                            ug_satis = st.number_input(
                                                "Müşteriye Fiyat ($)",
                                                value=float(ku.get("satis_fiyati", 0) or 0),
                                                step=0.01, format="%.2f",
                                                key=f"ug_s_{kid}_{ku_id}"
                                            )
                                        with ug2:
                                            ug_fd = st.number_input(
                                                "Birim Firma Desteği ($)",
                                                value=float(ku.get("birim_firma_destek", 0) or 0),
                                                step=0.01, format="%.2f",
                                                key=f"ug_fd_{kid}_{ku_id}"
                                            )
                                        with ug3:
                                            ug_ed = st.number_input(
                                                "Birim Ek Destek ($)",
                                                value=float(ku.get("birim_ek_destek", 0) or 0),
                                                step=0.01, format="%.2f",
                                                key=f"ug_ed_{kid}_{ku_id}"
                                            )
                                        with ug4:
                                            ug_satilan = st.number_input(
                                                "Satılan Adet",
                                                value=int(ku.get("satilan_adet", 0) or 0),
                                                min_value=0, step=1,
                                                key=f"ug_sa_{kid}_{ku_id}"
                                            )
    
                                        # Canlı hesap göster
                                        if pacal_ku > 0 and ug_satis > 0:
                                            toplam_destek = ug_fd + ug_ed
                                            net_kar = (ug_satis - pacal_ku) - toplam_destek
                                            net_marj = (net_kar / ug_satis * 100) if ug_satis > 0 else 0
                                            toplam_net = net_kar * ug_satilan
                                            renk = "#A5D6A7" if net_kar >= 0 else "#FFCDD2"
                                            st.markdown(
                                                f'<div class="info-box" style="font-size:13px; margin:4px 0;">'
                                                f'⭐ Paçal: <b>${pacal_ku:.2f}</b> &nbsp;|&nbsp; '
                                                f'Net Kar/Adet: <span style="color:{renk}; font-weight:700;">${net_kar:.2f} (%{net_marj:.1f})</span> &nbsp;|&nbsp; '
                                                f'Toplam Net: <span style="color:{renk}; font-weight:700;">${toplam_net:.0f}</span>'
                                                f'</div>',
                                                unsafe_allow_html=True
                                            )
    
                                        ub1, ub2 = st.columns([3, 1])
                                        with ub2:
                                            if st.form_submit_button("💾 Güncelle", use_container_width=True, type="primary"):
                                                guncelle_kampanya_urun(ku_id, ug_satis, ug_fd, ug_ed, ug_satilan, ku.get("notlar",""))
                                                st.cache_data.clear()
                                                st.toast(f"✅ Güncellendi!")
                                                st.session_state['_kamp_detay_ac'] = True
                                                st.rerun()
    
                        # ── Kampanyayı kopyala (ürünler + destekler ile YENİ kampanya) ──
                        with st.expander(f"📋 Kampanyayı Kopyala — {kamp['kampanya_adi']}"):
                            st.caption("Bu kampanyanın TÜM ürünlerini birim firma desteği ve ek desteğiyle birlikte "
                                       "yeni bir kampanyaya kopyalar (satılan adetler sıfırlanır). Kopyaladıktan sonra "
                                       "yeni kampanyayı seçip ad/tarihi düzenler, gerekirse ek destekleri güncellersin.")
                            _kop_c1, _kop_c2 = st.columns([2, 1])
                            _kop_ad = _kop_c1.text_input("Yeni kampanya adı",
                                                         value=f"{kamp['kampanya_adi']} (KOPYA)", key=f"kop_ad_{kid}")
                            _kop_c2.markdown('<div style="height:28px"></div>', unsafe_allow_html=True)
                            if _kop_c2.button("📋 Kopyala ve Oluştur", type="primary",
                                              use_container_width=True, key=f"kop_btn_{kid}"):
                                _kop_hata, _yeni_id = None, None
                                try:
                                    _yeni_id = ekle_kampanya(
                                        (_kop_ad.strip() or f"{kamp['kampanya_adi']} (KOPYA)"),
                                        kamp.get("firma", ""), kamp.get("baslangic_tarihi"), kamp.get("bitis_tarihi"),
                                        kamp.get("notlar", "") or "", kamp.get("kategori", "") or "",
                                        kampanya_turu=(kamp.get("kampanya_turu", "") or ""),
                                        spiff_tl=(kamp.get("spiff_tl") or 0), spiff_kur=(kamp.get("spiff_kur") or 0),
                                        spiff_fatura=bool(kamp.get("spiff_fatura")))
                                except Exception as _e:
                                    _kop_hata = str(_e)
                                if _yeni_id:
                                    _kn = 0
                                    for _ku in get_kampanya_urunler(kid):
                                        try:
                                            ekle_kampanya_urun(
                                                _yeni_id, _ku.get("sku", ""), _ku.get("urun_adi", ""),
                                                _ku.get("pacal_maliyet", 0), _ku.get("satis_fiyati", 0),
                                                _ku.get("birim_firma_destek", 0), _ku.get("birim_ek_destek", 0),
                                                _ku.get("notlar", "") or "")
                                            _kn += 1
                                        except Exception:
                                            pass
                                    st.cache_data.clear()
                                    st.success(f"✅ Kampanya kopyalandı — {_kn} ürün taşındı. "
                                               f"Yeni kampanya: '{_kop_ad.strip()}'. Listeden seçip düzenleyebilirsin.")
                                    st.session_state['_kamp_detay_ac'] = True
                                    st.rerun()
                                elif _kop_hata:
                                    st.error(f"Kopyalama başarısız — {_kop_hata}")
                                else:
                                    st.error("Kopyalama başarısız — yeni kampanya oluşturulamadı (tablo izni/kolon olabilir).")

                        # Ürün ekleme formu
                        with st.expander(f"➕ Ürün Ekle — {kamp['kampanya_adi']}"):
                            if not sku_listesi_k:
                                st.warning("Önce 'Veri Yükleme' sekmesinden ürün yükleyin.")
                            else:
                                with st.form(f"urun_ekle_{kid}", clear_on_submit=True):
                                    uf1, uf2 = st.columns(2)
                                    with uf1:
                                        u_secim = st.selectbox("Ürün *", list(sku_listesi_k.keys()), key=f"u_sec_{kid}")
                                        u_sku = sku_listesi_k.get(u_secim, "")
                                        u_bilgi = urun_dict_k.get(u_sku, {})
                                        pacal = u_bilgi.get("final_cost_price", 0)
                                        pacal_son = u_bilgi.get("son_final", 0) or 0
    
                                        # Top-down margin hesaplama yardımcısı
                                        if pacal > 0:
                                            hedef_marj_giris = st.number_input(
                                                "Hedef Net Marj % (opsiyonel)",
                                                min_value=0.0, max_value=99.0, value=0.0, step=0.5, format="%.1f",
                                                help="Top-down hesap: Satış = Paçal / (1 - marj%). Örn: %20 → $100 paçal → $125 satış",
                                                key=f"u_marj_{kid}"
                                            )
                                            if hedef_marj_giris > 0:
                                                onerilen_satis = pacal / (1 - hedef_marj_giris / 100)
                                                st.markdown(f"""<div class="info-box" style="font-size:13px">
                                                💡 %{hedef_marj_giris:.1f} marj için önerilen satış: <b>${onerilen_satis:.2f}</b>
                                                </div>""", unsafe_allow_html=True)
    
                                        u_satis = st.number_input("Satış Fiyatı ($) *", min_value=0.0, value=0.0, step=0.01, format="%.2f", key=f"u_satis_{kid}")
                                        u_not = st.text_area("Notlar", key=f"u_not_{kid}")
                                    with uf2:
                                        u_firma_destek = st.number_input("Birim Firma Desteği ($)", min_value=0.0, value=0.0, step=0.01, format="%.2f",
                                            help="Firma tarafından ürün başına verilen destek tutarı", key=f"u_fd_{kid}")
                                        u_ek_destek = st.number_input("Birim Ek Destek ($)", min_value=0.0, value=0.0, step=0.01, format="%.2f",
                                            help="Ek destek tutarı (ürün başına)", key=f"u_ed_{kid}")
    
                                        # Seçilen ürünün paçal maliyetini göster
                                        if pacal > 0:
                                            toplam_destek = u_firma_destek + u_ek_destek
                                            net_kar_birim = (u_satis - pacal) - toplam_destek if u_satis > 0 else 0
                                            net_marj = (net_kar_birim / u_satis * 100) if u_satis > 0 else 0  # Top-down: kar/satış
                                            st.markdown(f"""
                                            <div class="info-box" style="font-size:13px">
                                            ⭐ Paçal: <b>${pacal:.2f}</b>{f' &nbsp;·&nbsp; 🆕 Son: <b>${pacal_son:.2f}</b>' if pacal_son else ''}<br>
                                            💸 Toplam Destek: <b>${toplam_destek:.2f}</b><br>
                                            📈 Net Kar/Adet: <b>${net_kar_birim:.2f} (%{net_marj:.1f})</b>
                                            </div>
                                            """, unsafe_allow_html=True)
                                        else:
                                            st.info(f"⭐ Paçal: Henüz satın alma kaydı yok")
    
                                    if st.form_submit_button("➕ Ürünü Kampanyaya Ekle", type="primary", use_container_width=True):
                                        if not u_secim or u_satis <= 0:
                                            st.error("Ürün ve satış fiyatı zorunludur.")
                                        else:
                                            ekle_kampanya_urun(
                                                kid, u_sku, u_bilgi.get("urun_adi", u_sku),
                                                pacal,
                                                u_satis, u_firma_destek, u_ek_destek, u_not
                                            )
                                            st.toast("✅ Ürün eklendi!")
                                            st.session_state['_kamp_detay_ac'] = True
                                            st.rerun()
    
                        # ── Excel ile toplu ürün ekle + şablon indir ──

                        # Kampanya ürünleri tablosu
                        if k_urunler:
                            st.markdown(f"**Kampanya Ürünleri ({len(k_urunler)} ürün)**")
    
                            # Özet hesaplar
                            toplam_net_kar = 0
                            toplam_destek_verilen = 0
                            toplam_ek_destek_verilen = 0  # yalnızca ek destek (firma desteği hariç) — "Toplam Destek Verilen" göstergesi için
                            toplam_satilan = 0
    
                            rows_ku = []
                            for ku in k_urunler:
                                pacal = ku.get("pacal_maliyet") or 0
    
                                # Paçal 0 ise güncel değeri çek ve güncelle
                                if pacal == 0:
                                    u_bilgi_c = urun_dict_k.get(ku["sku"], {})
                                    pacal_guncel = u_bilgi_c.get("final_cost_price", 0)
                                    if pacal_guncel > 0:
                                        try:
                                            get_client().table("kampanya_urunler").update(
                                                {"pacal_maliyet": pacal_guncel}).eq("id", ku["id"]).execute()
                                        except Exception:
                                            pass
                                        pacal = pacal_guncel
    
                                satis = ku.get("satis_fiyati") or 0
                                fd = ku.get("birim_firma_destek") or 0
                                ed = ku.get("birim_ek_destek") or 0
                                satilan = ku.get("satilan_adet") or 0
                                # YÖNTEM: destekler satış fiyatından düşülür → NET SATIŞ.
                                # Net Kâr/Adet = (satış − destekler) − paçal
                                # Net Marj     = Net Kâr / (satış − destekler)   [örn. 1 − 5,10/(12−3) = %43,3]
                                toplam_destek_birim = fd + ed
                                net_satis_birim = satis - toplam_destek_birim
                                net_kar_birim = net_satis_birim - pacal if satis > 0 and pacal > 0 else 0
                                net_marj = (net_kar_birim / net_satis_birim * 100) if net_satis_birim > 0 else 0
                                toplam_destek_urun = toplam_destek_birim * satilan
                                toplam_net_urun = net_kar_birim * satilan
    
                                toplam_net_kar += toplam_net_urun
                                toplam_destek_verilen += toplam_destek_urun
                                toplam_ek_destek_verilen += (ed * satilan)
                                toplam_satilan += satilan
    
                                rows_ku.append({
                                    "ID": ku["id"],
                                    "SKU": ku["sku"],
                                    "Ürün": ku.get("urun_adi",""),
                                    "⭐ Paçal ($)": f"${pacal:.2f}" if pacal else "—",
                                    "Satış ($)": f"${satis:.2f}",
                                    "Firma Destek ($)": f"${fd:.2f}",
                                    "Ek Destek ($)": f"${ed:.2f}",
                                    "Net Kar/Adet ($)": f"${net_kar_birim:.2f}",
                                    "Net Marj (%)": f"%{net_marj:.1f}",
                                    "Satılan Adet": satilan,
                                    "Toplam Destek ($)": f"${toplam_destek_urun:.0f}",
                                    "Toplam Net Kar ($)": f"${toplam_net_urun:.0f}",
                                    "Notlar": ku.get("notlar","") or "",
                                })
    
                            def _pf(x):
                                try:
                                    return float(str(x).replace("$", "").replace("%", "").replace(",", "").strip())
                                except Exception:
                                    return 0.0
                            k_rows = ""
                            for rk in rows_ku:
                                urun = str(rk.get("Ürün", "") or "")
                                urun_k = urun if len(urun) <= 40 else urun[:39] + "…"
                                urun_t = urun.replace(chr(34), "&quot;")
                                nkb = _pf(rk.get("Net Kar/Adet ($)"))
                                nkb_cls = "kc-pos" if nkb > 0 else ("kc-neg" if nkb < 0 else "kc-num")
                                tnk = _pf(rk.get("Toplam Net Kar ($)"))
                                tnk_cls = "kc-pos" if tnk > 0 else ("kc-neg" if tnk < 0 else "kc-num")
                                notlar = str(rk.get("Notlar", "") or "")
                                notlar_k = notlar if len(notlar) <= 24 else notlar[:23] + "…"
                                notlar_t = notlar.replace(chr(34), "&quot;")
                                k_rows += (
                                    "<tr>"
                                    f'<td class="kc-sku">{rk.get("SKU","")}</td>'
                                    f'<td class="kc-name" title="{urun_t}">{urun_k}</td>'
                                    f'<td class="kc-gold">{rk.get("⭐ Paçal ($)","")}</td>'
                                    f'<td class="kc-money">{rk.get("Satış ($)","")}</td>'
                                    f'<td class="kc-dim">{rk.get("Firma Destek ($)","")}</td>'
                                    f'<td class="kc-dim">{rk.get("Ek Destek ($)","")}</td>'
                                    f'<td class="{nkb_cls}">{rk.get("Net Kar/Adet ($)","")}</td>'
                                    f'<td class="kc-dim">{rk.get("Net Marj (%)","")}</td>'
                                    f'<td class="kc-num">{rk.get("Satılan Adet","")}</td>'
                                    f'<td class="kc-dim">{rk.get("Toplam Destek ($)","")}</td>'
                                    f'<td class="{tnk_cls}">{rk.get("Toplam Net Kar ($)","")}</td>'
                                    f'<td class="kc-note" title="{notlar_t}">{notlar_k}</td>'
                                    "</tr>"
                                )
                            k_css = (
                                "<style>"
                                ".kw{overflow-x:auto;border-radius:12px;box-shadow:0 2px 14px rgba(0,0,0,0.25);margin:4px 0}"
                                ".kt{width:100%;border-collapse:collapse;font-family:Inter,sans-serif}"
                                ".kt thead tr{background:linear-gradient(135deg,#1E293B,#0F172A)}"
                                ".kt thead th{padding:8px 12px;color:#CBD5E1;font-size:11px;font-weight:700;letter-spacing:.4px;text-transform:uppercase;white-space:nowrap;text-align:right}"
                                ".kt thead th:nth-child(1),.kt thead th:nth-child(2),.kt thead th:last-child{text-align:left}"
                                ".kt tbody{background:#131C35}"
                                ".kt td{padding:8px 12px;font-size:11px}"
                                ".kt tbody tr{border-bottom:1px solid rgba(255,255,255,0.05)}"
                                ".kt tbody tr:hover{background:rgba(99,102,241,0.06)}"
                                ".kc-sku{color:#E2E8F0;font-family:'JetBrains Mono',monospace;font-weight:600;white-space:nowrap}"
                                ".kc-name{color:#CBD5E1;max-width:220px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}"
                                ".kc-note{color:#78909C;max-width:160px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;font-size:11px}"
                                ".kc-num{text-align:right;color:#CBD5E1;font-family:'JetBrains Mono',monospace}"
                                ".kc-dim{text-align:right;color:#94A3B8;font-family:'JetBrains Mono',monospace}"
                                ".kc-money{text-align:right;color:#E2E8F0;font-family:'JetBrains Mono',monospace;font-weight:600}"
                                ".kc-gold{text-align:right;color:#FFD54F;font-weight:600;font-family:'JetBrains Mono',monospace}"
                                ".kc-pos{text-align:right;color:#4ADE80;font-weight:700;font-family:'JetBrains Mono',monospace}"
                                ".kc-neg{text-align:right;color:#F87171;font-weight:700;font-family:'JetBrains Mono',monospace}"
                                "</style>"
                            )
                            k_head = (
                                '<div class="kw"><table class="kt"><thead><tr>'
                                "<th>SKU</th><th>Ürün</th><th>⭐ Paçal</th><th>Satış</th><th>Firma Destek</th>"
                                "<th>Ek Destek</th><th>Net Kar/Adet</th><th>Net Marj</th><th>Satılan</th>"
                                "<th>Toplam Destek</th><th>Toplam Net Kar</th><th>Notlar</th>"
                                "</tr></thead><tbody>"
                            )
                            st.html(k_css + k_head + k_rows + "</tbody></table></div>")
    
                            # Kampanya özet metrikleri (+ Spiff kampanya maliyeti)
                            _spiff_tl_p = float(kamp.get("spiff_tl") or 0)
                            _spiff_kur_p = float(kamp.get("spiff_kur") or 0)
                            _spiff_usd_p = (_spiff_tl_p / _spiff_kur_p) if _spiff_kur_p > 0 else 0.0
                            _spiff_fat_p = bool(kamp.get("spiff_fatura"))
                            _kamp_turu_p = (kamp.get("kampanya_turu") or "").strip()
                            _net_final = toplam_net_kar - _spiff_usd_p
                            _metrikler = [
                                {"label": "📦 Toplam Satılan", "value": f"{toplam_satilan:,} adet", "renk": "#818CF8"},
                                {"label": "💸 Toplam Destek Verilen", "value": f"${toplam_ek_destek_verilen:,.0f}",
                                 "renk": "#FB923C", "alt": "yalnızca ek destek"},
                            ]
                            if _spiff_usd_p > 0:
                                _metrikler.append({
                                    "label": "🎟️ Spiff Maliyeti", "value": f"${_spiff_usd_p:,.0f}",
                                    "renk": "#FBBF24",
                                    "alt": f"₺{_spiff_tl_p:,.0f} @ {_spiff_kur_p:.2f} · {'fatura/kesin' if _spiff_fat_p else 'tahmini'}"})
                            _metrikler.append({
                                "label": "📈 Net Kar" + (" (Spiff sonrası)" if _spiff_usd_p > 0 else ""),
                                "value": f"${_net_final:,.0f}",
                                "renk": "#34D399" if _net_final > 0 else "#F87171",
                                "alt": "Kârlı" if _net_final > 0 else "Zararlı"})
                            _metrikler.append({
                                "label": "🏪 Firma", "value": kamp["firma"], "renk": "#A78BFA",
                                "alt": (f"Tür: {_kamp_turu_p}" if _kamp_turu_p else None)})
                            metrik_satiri(_metrikler)
    
                            # Ürün düzenleme
                            with st.expander("✏️ Ürün Bilgilerini Güncelle"):
                                st.caption("Satılan adet, fiyat veya destek tutarlarını güncelleyebilirsiniz.")
                                guncelle_id = st.selectbox(
                                    "Güncellenecek Ürün",
                                    [f"ID:{r['ID']} — {r['Ürün']}" for r in rows_ku],
                                    key=f"gun_sec_{kid}"
                                )
                                g_id = int(guncelle_id.split(":")[1].split(" ")[0])
                                g_urun = next(ku for ku in k_urunler if ku["id"] == g_id)
    
                                with st.form(f"gun_form_{kid}_{g_id}"):
                                    gf1, gf2 = st.columns(2)
                                    with gf1:
                                        g_satis = st.number_input("Satış Fiyatı ($)", value=float(g_urun.get("satis_fiyati",0) or 0), step=0.01, format="%.2f", key=f"g_s_{kid}_{g_id}")
                                        g_fd = st.number_input("Birim Firma Desteği ($)", value=float(g_urun.get("birim_firma_destek",0) or 0), step=0.01, format="%.2f", key=f"g_fd_{kid}_{g_id}")
                                    with gf2:
                                        g_ed = st.number_input("Birim Ek Destek ($)", value=float(g_urun.get("birim_ek_destek",0) or 0), step=0.01, format="%.2f", key=f"g_ed_{kid}_{g_id}")
                                        g_satilan = st.number_input("Satılan Adet", value=int(g_urun.get("satilan_adet",0) or 0), min_value=0, step=1, key=f"g_sa_{kid}_{g_id}")
                                    g_not = st.text_area("Notlar", value=g_urun.get("notlar","") or "", key=f"g_not_{kid}_{g_id}")
    
                                    gf_c1, gf_c2 = st.columns(2)
                                    with gf_c1:
                                        if st.form_submit_button("💾 Güncelle", type="primary", use_container_width=True):
                                            guncelle_kampanya_urun(g_id, g_satis, g_fd, g_ed, g_satilan, g_not)
                                            st.cache_data.clear()
                                            st.toast("Güncellendi!")
                                            st.session_state['_kamp_detay_ac'] = True
                                            st.rerun()
                                    with gf_c2:
                                        if st.form_submit_button("🗑️ Ürünü Sil", use_container_width=True):
                                            sil_kampanya_urun(g_id)
                                            st.cache_data.clear()
                                            st.warning("Ürün kaldırıldı.")
                                            st.session_state['_kamp_detay_ac'] = True
                                            st.rerun()
                        else:
                            st.info("Bu kampanyaya henüz ürün eklenmemiş.")
    
                        st.markdown("---")
    
                    if st.session_state.pop("_kamp_detay_ac", False):
                        _kampanya_detay_dialog()
        # ─────────────────────────────────────────────────────────────────
        # TAB 2: GEÇMİŞ KAMPANYALAR
        # ─────────────────────────────────────────────────────────────────
        if _kt_durum == "📁 Geçmiş Kampanyalar":
            gecmis = _kt_uygula(get_kampanyalar(durum="kapali"))
            if not gecmis:
                st.info("Henüz kapatılmış kampanya yok.")
            else:
                for kamp in gecmis:
                    kid = kamp["id"]
                    k_urunler = _ku_map.get(kid, [])
    
                    # Özet hesaplar
                    toplam_net = sum(
                        ((ku.get("satis_fiyati",0) or 0) - (ku.get("pacal_maliyet",0) or 0)
                         - (ku.get("birim_firma_destek",0) or 0) - (ku.get("birim_ek_destek",0) or 0))
                        * (ku.get("satilan_adet",0) or 0)
                        for ku in k_urunler
                    )
                    toplam_destek = sum(
                        ((ku.get("birim_firma_destek",0) or 0) + (ku.get("birim_ek_destek",0) or 0))
                        * (ku.get("satilan_adet",0) or 0)
                        for ku in k_urunler
                    )
                    toplam_satilan = sum(ku.get("satilan_adet",0) or 0 for ku in k_urunler)
    
                    net_renk = "#A5D6A7" if toplam_net >= 0 else "#FFCDD2"
                    net_bg = "#1B5E20" if toplam_net >= 0 else "#7F0000"
    
                    st.markdown(f"""
                    <div style="background:rgba(255,255,255,0.04); border-radius:12px; padding:16px 20px;
                                margin:8px 0; border-left:5px solid #546E7A;">
                      <div style="display:flex; justify-content:space-between; align-items:flex-start; flex-wrap:wrap; gap:8px;">
                        <div>
                          <span style="color:#CFD8DC; font-size:15px; font-weight:700;">📁 {kamp["kampanya_adi"]}</span>
                          <span style="background:#263238; color:#90A4AE; padding:0px 8px; border-radius:8px; font-size:11px; margin-left:8px;">{kamp["firma"]}</span>
                          <span style="background:#37474F; color:#CFD8DC; padding:0px 8px; border-radius:8px; font-size:11px; margin-left:4px;">● KAPALI</span>
                          <div style="color:#78909C; font-size:13px; margin-top:4px;">{kamp["baslangic_tarihi"]} → {kamp["bitis_tarihi"]}</div>
                        </div>
                        <div style="display:flex; gap:8px; flex-wrap:wrap;">
                          <div style="background:rgba(255,255,255,0.022); border:1px solid rgba(255,255,255,0.06); border-left:3px solid #818CF8; border-radius:10px; padding:8px 12px;">
                            <div style="color:#8B97A8; font-size:11px; font-weight:700; letter-spacing:.5px; text-transform:uppercase;">SATILAN</div>
                            <div style="color:#C7D2FE; font-size:15px; font-weight:800;">{toplam_satilan:,} adet</div>
                          </div>
                          <div style="background:rgba(255,255,255,0.022); border:1px solid rgba(255,255,255,0.06); border-left:3px solid #FB923C; border-radius:10px; padding:8px 12px;">
                            <div style="color:#8B97A8; font-size:11px; font-weight:700; letter-spacing:.5px; text-transform:uppercase;">TOPLAM DESTEK</div>
                            <div style="color:#FB923C; font-size:15px; font-weight:800;">${toplam_destek:,.0f}</div>
                          </div>
                          <div style="background:rgba(255,255,255,0.022); border:1px solid rgba(255,255,255,0.06); border-left:3px solid {'#34D399' if toplam_net >= 0 else '#F87171'}; border-radius:10px; padding:8px 12px;">
                            <div style="color:#8B97A8; font-size:11px; font-weight:700; letter-spacing:.5px; text-transform:uppercase;">NET KAR</div>
                            <div style="color:{'#34D399' if toplam_net >= 0 else '#F87171'}; font-size:15px; font-weight:800;">${toplam_net:,.0f}</div>
                          </div>
                        </div>
                      </div>
                    </div>
                    """, unsafe_allow_html=True)
    
                    if k_urunler:
                        with st.expander(f"📋 Detaylar — {kamp['kampanya_adi']}"):
                            rows_g = []
                            for ku in k_urunler:
                                pacal = ku.get("pacal_maliyet") or 0
                                # Paçal 0 ise güncel değeri çek
                                if pacal == 0:
                                    pacal = urun_dict_k.get(ku["sku"], {}).get("final_cost_price", 0)
                                satis = ku.get("satis_fiyati") or 0
                                fd = ku.get("birim_firma_destek") or 0
                                ed = ku.get("birim_ek_destek") or 0
                                satilan = ku.get("satilan_adet") or 0
                                toplam_d = (fd + ed) * satilan
                                # Top-down margin: kar/satış
                                net_b = (satis - pacal) - (fd + ed) if satis > 0 else 0
                                net_t = net_b * satilan
                                rows_g.append({
                                    "SKU": ku["sku"],
                                    "Ürün": ku.get("urun_adi",""),
                                    "⭐ Paçal ($)": f"${pacal:.2f}" if pacal else "—",
                                    "Satış ($)": f"${satis:.2f}",
                                    "Firma D. ($)": f"${fd:.2f}",
                                    "Ek D. ($)": f"${ed:.2f}",
                                    "Net Kar/Adet ($)": f"${net_b:.2f}",
                                    "Satılan": satilan,
                                    "Top. Destek ($)": f"${toplam_d:.0f}",
                                    "Top. Net Kar ($)": f"${net_t:.0f}",
                                })
                            st.dataframe(pd.DataFrame(rows_g), use_container_width=True, hide_index=True, height=tablo_h(len(rows_g)))
    
                            if st.button(f"🗑️ Kampanyayı Sil", key=f"sil_gecmis_{kid}"):
                                sil_kampanya(kid)
                                st.cache_data.clear()
                                st.warning("Silindi.")
                                st.rerun()
    
    
    elif sayfa == "📦  Sipariş Önerisi":
        from .database import get_uretim_suresi, set_uretim_suresi
        _esik = get_uretim_suresi()
        st.markdown('<div class="baslik">📦 Sipariş Önerisi</div>', unsafe_allow_html=True)
        st.markdown(f'<div class="alt-baslik">{_esik} günden az stok kalan ürünler · Otomatik öneri</div>', unsafe_allow_html=True)
        st.markdown('<div class="sayfa-baslik-cizgi"></div>', unsafe_allow_html=True)

        with st.popover(f"⚙️ Sipariş eşiği (üretim/tedarik süresi) — şu an {_esik} gün", use_container_width=False):
            st.caption("Stok bu kadar günde biteceği zaman 'sipariş ver' uyarısı çıkar. "
                       "Üretim/tedarik sürenize göre ayarlayın (varsayılan 135 gün).")
            _e1, _e2 = st.columns([2, 1])
            _yeni_esik = _e1.number_input("Eşik (gün)", min_value=1, max_value=730, value=int(_esik),
                                          step=5, key="uretim_suresi_input")
            _e2.markdown("<br>", unsafe_allow_html=True)
            if _e2.button("💾 Kaydet", use_container_width=True, key="uretim_suresi_kaydet"):
                if set_uretim_suresi(int(_yeni_esik)):
                    st.cache_data.clear()
                    st.toast(f"✅ Sipariş eşiği {int(_yeni_esik)} güne ayarlandı", icon="✅")
                    st.rerun()
                else:
                    st.error("Kaydedilemedi. 'pm_ayarlar' tablosu eksik olabilir — Supabase'de şu SQL'i çalıştırın:")
                    st.code("create table if not exists pm_ayarlar (anahtar text primary key, deger text);\n"
                            "alter table pm_ayarlar disable row level security;", language="sql")

        try:
            siparis_listesi = siparis_onerisi_listesi()
            urun_data = tum_urunler_listesi()
            urun_dict = {u["sku"]: u for u in urun_data}
        except Exception as e:
            _log.error("Hata: %s", e)
            st.error(f"Veri yüklenemedi: {e}")
            st.stop()
    
        if not siparis_listesi:
            st.success(f"✅ Tüm ürünlerde {_esik} günden fazla stok var, sipariş gerekmiyor!")
            st.stop()
    
        # Özet
        acil = [u for u in siparis_listesi if u.get("siparis_durum") == "acil"]
        yaklasan = [u for u in siparis_listesi if u.get("siparis_durum") == "yaklasıyor"]
        planlama = [u for u in siparis_listesi if u.get("siparis_durum") == "planlama"]
    
        metrik_satiri([
            {"label": "🔴 ACİL", "value": f"{len(acil):,}", "renk": "#F87171"},
            {"label": "🟠 Yaklaşıyor (30 gün)", "value": f"{len(yaklasan):,}", "renk": "#FB923C"},
            {"label": "🟡 Planlama (60 gün)", "value": f"{len(planlama):,}", "renk": "#FBBF24"},
        ])

        st.markdown("---")
    
        for urun in siparis_listesi:
            durum = urun.get("siparis_durum","")
            mesaj = urun.get("siparis_mesaj","")
            oneri = urun.get("oneri_miktar",0)
            oneri_mesaj = urun.get("oneri_mesaj","")
            sku = urun["sku"]
            ud = urun_dict.get(sku, {})
            fcp = ud.get("final_cost_price", 0)
            satis = ud.get("satis_fiyati", 0)
    
            # Stoku olan firmaları (kompakt)
            firma_detay = urun.get("firma_detay", [])
            firma_kisa = " · ".join(f'{fd["firma"]}:{fd["stok"]}' for fd in firma_detay if fd.get("stok", 0) > 0)

            if durum == "acil":
                brd, ik, dt = "rgba(239,68,68,0.6)", "🔴", "rgba(239,68,68,0.07)"
            elif durum == "yaklasıyor":
                brd, ik, dt = "rgba(245,158,11,0.6)", "🟠", "rgba(245,158,11,0.06)"
            else:
                brd, ik, dt = "rgba(234,179,8,0.55)", "🟡", "rgba(234,179,8,0.06)"

            alt = [f'📅 {mesaj}', f'G5F: {urun["bizim_stok"]}', f'Hft: {urun.get("ortalama_haftalik_satis",0):.1f}']
            if firma_kisa:
                alt.append(firma_kisa)
            if fcp > 0:
                alt.append(f'💵 ${fcp:,.0f}→${satis:,.0f}')
            alt_str = "  ·  ".join(alt)

            c1, c2, c3 = st.columns([6, 1.2, 1.7])
            with c1:
                st.markdown(
                    f'<div style="background:{dt};border-left:3px solid {brd};border-radius:8px;padding:8px 12px;">'
                    f'<div style="font-size:13px;font-weight:600;color:#E2E8F0;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;">'
                    f'{ik} {urun["urun_adi"]} <span style="color:#94A3B8;font-size:11px;font-family:monospace;">{sku}</span></div>'
                    f'<div style="font-size:11px;color:#94A3B8;margin-top:4px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;">'
                    f'{alt_str}  ·  <span style="color:#FBBF24;font-weight:600;">💡 {oneri} adet</span></div>'
                    f'</div>',
                    unsafe_allow_html=True
                )
            with c2:
                miktar = st.number_input("Miktar", min_value=1, value=max(oneri, 1),
                                         key=f"sp_miktar_{sku}", label_visibility="collapsed")
            with c3:
                if st.button("📦 Sipariş Ekle", key=f"sp_btn_{sku}", use_container_width=True):
                    from .database import ekle_siparis_onerisi
                    ekle_siparis_onerisi("G5F", sku, urun["urun_adi"], miktar)
                    st.cache_data.clear()
                    st.toast(f"✅ {urun['urun_adi']} için {miktar} adet sipariş önerisi oluşturuldu!")
                    st.rerun()
    
        # Onaylanan/Bekleyen geçmiş
        st.markdown("---")
        st.markdown("#### 📋 Sipariş Önerisi Geçmişi")
        from .database import get_siparis_onerileri
        onceki = get_siparis_onerileri()
        if onceki:
            rows_sp = []
            for sp in onceki:
                rows_sp.append({
                    "ID": sp["id"],
                    "SKU": sp["sku"],
                    "Ürün": sp.get("urun_adi",""),
                    "Miktar": sp["oneri_miktari"],
                    "Durum": sp["durum"],
                    "Tarih": sp["olusturma_tarihi"],
                    "Onay": sp.get("onay_tarihi","") or "",
                })
            df_sp = pd.DataFrame(rows_sp)
    
            def sp_rengi(row):
                d = row.get("Durum","")
                if d == "onaylandi":   return ["background-color:#1B5E20; color:#A5D6A7"]*len(row)
                if d == "reddedildi":  return ["background-color:#7F0000; color:#FFCDD2"]*len(row)
                return ["background-color:#827717; color:#FFF176"]*len(row)
    
            render_renkli_tablo(
                df_sp,
                kisalt={"Ürün": 42},
                satir_durum=("Durum", {"onaylandi": "rk-grn", "reddedildi": "rk-red", "bekliyor": "rk-yel"}),
                gizle=["ID"],
            )
    
            col_o1, col_o2 = st.columns(2)
            with col_o1:
                onayla_id = st.number_input("Onaylanacak ID", min_value=1, step=1, key="onayla_id")
                if st.button("✅ Onayla", key="onayla_btn", use_container_width=True):
                    from .database import onayla_siparis
                    onayla_siparis(int(onayla_id))
                    st.toast("Onaylandı!")
                    st.rerun()
            with col_o2:
                reddet_id = st.number_input("Reddedilecek ID", min_value=1, step=1, key="reddet_id")
                if st.button("❌ Reddet", key="reddet_btn", use_container_width=True):
                    from .database import reddet_siparis
                    reddet_siparis(int(reddet_id))
                    st.warning("Reddedildi.")
                    st.rerun()

    
    
    elif sayfa == "🔖  Ref No Takibi":
        from .ref_no import render as _ref_render
        _ref_render()

    elif sayfa == "📂  Veri Yükleme":
        st.markdown('<div class="baslik">📂 Veri Yükleme</div>', unsafe_allow_html=True)
        st.markdown('<div class="alt-baslik">Excel yükle · Geçmiş yüklemeleri gör · Veriyi yönet</div>', unsafe_allow_html=True)
        st.markdown('<div class="sayfa-baslik-cizgi"></div>', unsafe_allow_html=True)

        # 💲 Toplu Satış Fiyatı & Marj
        @st.dialog("💲 Toplu Satış Fiyatı & Marj — paçal maliyetten fiyat öner", width="large")
        def _dlg_toplu_fiyat():
            from .database import get_client as _gc_s, toplu_satis_kaydet as _satis_kaydet
            try:
                from ithalat.database import get_sku_maliyet_ozet as _ith_maliyet
                _pacal_map = _ith_maliyet() or {}
            except Exception:
                _pacal_map = {}
            try:
                _ur_s = (_gc_s().table("urunler").select("sku, urun_adi, satis_fiyati, hedef_kar_marji")
                         .order("urun_adi").execute().data) or []
            except Exception as _e_s:
                _ur_s = []
                st.warning(f"Ürünler okunamadı: {_e_s}")
            if not _ur_s:
                st.info("Henüz ürün yok.")
            else:
                _fiyatsiz = sum(1 for u in _ur_s if not (u.get("satis_fiyati") or 0))
                st.caption(f"Toplam {len(_ur_s)} ürün · satış fiyatı girilmemiş {_fiyatsiz}. "
                           "Hedef marjı gir → 🪄 Öner ile paçaldan satış fiyatı hesapla → düzelt → 💾 Kaydet.")
                _sc1, _sc2, _sc3 = st.columns([1, 1, 1])
                _hedef_marj = _sc1.number_input("Hedef marj (%)", min_value=0.0, max_value=500.0,
                                                value=25.0, step=5.0, key="satis_marj")
                _sadece_fiyatsiz = _sc2.checkbox("Sadece fiyatsız ürünler", value=False, key="satis_sadece")
                if _sc3.button("🪄 Marj'dan Satış Öner", use_container_width=True, key="satis_oner"):
                    _on = {}
                    for u in _ur_s:
                        _p = (_pacal_map.get(u["sku"], {}) or {}).get("pacal_final", 0) or 0
                        if _p > 0:
                            _on[u["sku"]] = round(_p * (1 + _hedef_marj / 100.0), 2)
                    st.session_state["_satis_oneri"] = _on
                    st.session_state["_satis_oneri_v"] = st.session_state.get("_satis_oneri_v", 0) + 1
                    st.toast(f"🪄 {len(_on)} ürün için satış fiyatı önerildi (marj %{_hedef_marj:.0f})", icon="🪄")
                    st.rerun()
                _son = st.session_state.get("_satis_oneri", {})
                st.caption("💡 Satış ($) hücresini elle de değiştirebilirsin. Paçal = İthalat maliyeti · "
                           "Marj % satışı değiştirince kaydederken yeniden hesaplanır.")
                _liste_s = [u for u in _ur_s if not (u.get("satis_fiyati") or 0)] if _sadece_fiyatsiz else _ur_s
                _rows_s = []
                for u in _liste_s:
                    _p = (_pacal_map.get(u["sku"], {}) or {}).get("pacal_final", 0) or 0
                    _ps = (_pacal_map.get(u["sku"], {}) or {}).get("son_final", 0) or 0
                    _satis = _son.get(u["sku"]) if u["sku"] in _son else (u.get("satis_fiyati") or 0)
                    _marj = ((_satis / _p - 1) * 100) if (_p > 0 and _satis) else 0.0
                    _rows_s.append({
                        "SKU": u["sku"], "Ürün Adı": u.get("urun_adi", ""),
                        "Paçal ($)": round(_p, 2), "Son ($)": round(_ps, 2),
                        "Satış ($)": round(float(_satis or 0), 2),
                        "Marj %": round(_marj, 1),
                    })
                _df_s = pd.DataFrame(_rows_s)
                _ed_s_key = f"satis_editor_{int(_sadece_fiyatsiz)}_{st.session_state.get('_satis_oneri_v', 0)}"
                _edited_s = st.data_editor(
                    _df_s, use_container_width=True, height=420, hide_index=True, key=_ed_s_key,
                    column_config={
                        "SKU": st.column_config.TextColumn("SKU", disabled=True, width="small"),
                        "Ürün Adı": st.column_config.TextColumn("Ürün Adı", disabled=True, width="large"),
                        "Paçal ($)": st.column_config.NumberColumn("Paçal ($)", disabled=True, format="$%.2f"),
                        "Son ($)": st.column_config.NumberColumn("Son ($)", disabled=True, format="$%.2f", help="En yeni ithalat dosyasındaki maliyet (referans · öneri paçala göre)"),
                        "Satış ($)": st.column_config.NumberColumn("Satış ($)", min_value=0.0, step=1.0, format="$%.2f"),
                        "Marj %": st.column_config.NumberColumn("Marj %", disabled=True, format="%.1f%%"),
                    },
                )
                if st.button("💾 Satış Fiyatlarını Kaydet", type="primary", key="satis_kaydet_btn"):
                    _map_s = {}
                    for _, r in _edited_s.iterrows():
                        _satis_v = float(r.get("Satış ($)") or 0)
                        _p = float(r.get("Paçal ($)") or 0)
                        _marj_v = ((_satis_v / _p - 1) * 100) if (_p > 0 and _satis_v) else 0.0
                        _map_s[str(r["SKU"])] = {"satis_fiyati": _satis_v, "hedef_kar_marji": round(_marj_v, 1)}
                    with st.spinner("Kaydediliyor..."):
                        _oks, _hts = _satis_kaydet(_map_s)
                    st.session_state.pop("_satis_oneri", None)
                    st.toast(f"✅ {_oks} ürün fiyatı kaydedildi" + (f" · {_hts} hata" if _hts else ""), icon="✅")
                    st.rerun()
        if st.button("💲 Toplu Satış Fiyatı & Marj — paçal maliyetten fiyat öner", key="btn_top_fiy", use_container_width=True):
            _dlg_toplu_fiyat()

        with st.expander("📋 Excel Şablonunu İndir (ilk kez kullanıyorsanız buradan başlayın)", expanded=False):
            st.markdown('<div style="color:#94A3B8;font-size:13px;line-height:1.6;margin-bottom:8px">Aşağıdaki butona tıklayıp örnek şablonu indir, doldur ve yükle.</div>', unsafe_allow_html=True)
            sablon_bytes = create_sample_excel_bytes()
            st.download_button("📥 Şablonu İndir", sablon_bytes, "SABLON_STOK_TAKIP.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    
        st.markdown("---")
    
        # Disa Aktar (eski Raporlar)
        @st.dialog("📤 Dışa Aktar — Excel / PDF Rapor", width="large")
        def _dlg_disa_aktar():
            _de1, _de2 = st.columns(2)
            with _de1:
                st.markdown('<div style="color:#90CAF9;font-size:13px;font-weight:700;letter-spacing:.5px;margin-bottom:4px">📊 EXCEL RAPORU</div>', unsafe_allow_html=True)
                st.markdown('<div style="color:#94A3B8;font-size:11px;line-height:1.6;margin-bottom:8px">Dashboard, Stok Yayılımı ve Sipariş Önerileri — 3 sekme, renkli.</div>', unsafe_allow_html=True)
                if st.button("📊 Excel Raporu Oluştur", use_container_width=True, type="primary", key="vy_excel_rapor"):
                    from .rapor import excel_rapor_olustur
                    import tempfile
                    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as _tmp:
                        _tmp_path = _tmp.name
                    _ok, _msg = excel_rapor_olustur(_tmp_path)
                    if _ok:
                        with open(_tmp_path, "rb") as _f:
                            st.download_button("⬇️ Excel İndir", _f.read(), f"Stok_Raporu_{tr_now().strftime('%Y%m%d_%H%M')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True, key="vy_excel_dl")
                        os.unlink(_tmp_path)
                    else:
                        st.error(_msg)
            with _de2:
                st.markdown('<div style="color:#F9A8D4;font-size:13px;font-weight:700;letter-spacing:.5px;margin-bottom:4px">📑 PDF RAPORU</div>', unsafe_allow_html=True)
                st.markdown('<div style="color:#94A3B8;font-size:11px;line-height:1.6;margin-bottom:8px">A4 yatay, yazdırmaya hazır özet rapor.</div>', unsafe_allow_html=True)
                if st.button("📑 PDF Raporu Oluştur", use_container_width=True, key="vy_pdf_rapor"):
                    from .rapor import pdf_rapor_olustur
                    import tempfile
                    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as _tmp:
                        _tmp_path = _tmp.name
                    _ok, _msg = pdf_rapor_olustur(_tmp_path)
                    if _ok:
                        with open(_tmp_path, "rb") as _f:
                            st.download_button("⬇️ PDF İndir", _f.read(), f"Stok_Raporu_{tr_now().strftime('%Y%m%d_%H%M')}.pdf", mime="application/pdf", use_container_width=True, key="vy_pdf_dl")
                        os.unlink(_tmp_path)
                    else:
                        st.error(_msg)
        if st.button("📤 Dışa Aktar — Excel / PDF Rapor", key="btn_disa_akt", use_container_width=True):
            _dlg_disa_aktar()

        st.markdown("---")
        st.markdown("---")
        st.markdown('<div style="font-size:13px;font-weight:700;color:#A5B4FC;letter-spacing:1px;text-transform:uppercase;margin:8px 0 8px;display:flex;align-items:center;gap:8px"><span style="width:5px;height:16px;border-radius:3px;background:linear-gradient(180deg,#38BDF8,#818CF8);display:inline-block"></span>🏬 G5F Stok · Depo Kırılımlı (Bizim Depo)</div>', unsafe_allow_html=True)
        st.markdown('<div style="color:#94A3B8;font-size:13px;line-height:1.6;margin-bottom:12px">Bizim depo stoğu — <b style="color:#CBD5E1">tek sayfa</b>, her satır bir depo-ürün. Sütunlar: <b style="color:#CBD5E1">DEPO ADI · STOK KODU · STOK İSMİ · MİKTAR</b>. Bir SKU birden çok depoda olabilir; <b>genel toplam</b> ve <b>depo kırılımı</b> tüm depolardan; sipariş önerisindeki <b>"bizim stok"</b> = Merkez depo + Happy Life. (Ürünün fiyat/kategori/marka bilgisine dokunmaz.)</div>', unsafe_allow_html=True)

        dosya_g = st.file_uploader("G5F Stok Excel'ini Seç", type=["xlsx", "xls"], key="g5f_depo_dosya")
        st.warning("📌 **Hareket bazlı stok (Model B) aktif:** İthalat teslimi, satış ve iadeler depo stoğunu "
                   "otomatik günceller. Bu G5F yüklemesi artık **SAYIM / DÜZELTMEDİR** — canlı takip edilen stoğu "
                   "Excel'deki değerlere **eşitler** (üzerine yazar). Yalnızca fiziksel sayım sonrası ya da "
                   "düzeltme amacıyla yükle.")
        if dosya_g:
            if st.button("⬆️ G5F Stok Yükle (Depo Kırılımlı)", type="primary", use_container_width=True, key="g5f_depo_btn"):
                import tempfile
                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmpg:
                    tmpg.write(dosya_g.read())
                    tmpg_path = tmpg.name
                with st.spinner("🏬 G5F depo kırılımlı stok işleniyor — ürünler senkronlanıyor…"):
                    basari_g, mesaj_g = excel_yukle_g5f_depolar(tmpg_path)
                os.unlink(tmpg_path)
                st.cache_data.clear()
                if basari_g:
                    st.success(mesaj_g)
                else:
                    st.error(mesaj_g)

        st.markdown("---")
        st.markdown('<div style="font-size:13px;font-weight:700;color:#A5B4FC;letter-spacing:1px;text-transform:uppercase;margin:8px 0 8px;display:flex;align-items:center;gap:8px"><span style="width:5px;height:16px;border-radius:3px;background:linear-gradient(180deg,#6366F1,#A78BFA);display:inline-block"></span>📅 Geçmiş Yüklemeler</div>', unsafe_allow_html=True)
        st.markdown('<div style="color:#94A3B8;font-size:13px;line-height:1.6;margin-bottom:8px">Hangi tarihlerde veri yüklendiğini gör, gerekirse sil.</div>', unsafe_allow_html=True)
    
        try:
            sb_vy = get_client()
    
            # Firma stok yükleme tarihleri
            firma_tarihler = sb_vy.table("firma_stok").select("yukleme_tarihi, firma").execute().data or []
            tarih_firma = {}
            for r in firma_tarihler:
                t = r["yukleme_tarihi"]
                if t not in tarih_firma:
                    tarih_firma[t] = set()
                tarih_firma[t].add(r["firma"])
    
            # Ürün yükleme tarihleri
            urun_tarihler = sb_vy.table("urunler").select("guncelleme_tarihi").execute().data or []
            urun_tarih_set = set(r["guncelleme_tarihi"] for r in urun_tarihler if r.get("guncelleme_tarihi"))
    
            if not tarih_firma and not urun_tarih_set:
                st.info("Henüz veri yüklenmemiş.")
            else:
                tum_tarihler = sorted(set(list(tarih_firma.keys()) + list(urun_tarih_set)), reverse=True)
    
                rows_vy = []
                for t in tum_tarihler:
                    firmalar = ", ".join(sorted(tarih_firma.get(t, [])))
                    urun_sayisi = len([r for r in urun_tarihler if r.get("guncelleme_tarihi") == t])
                    firma_kayit = sum(1 for r in firma_tarihler if r["yukleme_tarihi"] == t)
                    rows_vy.append({
                        "Tarih": t,
                        "Yüklenen Ürün": urun_sayisi,
                        "Firma Kayıt Sayısı": firma_kayit,
                        "Firmalar": firmalar or "—",
                    })
    
                df_vy = pd.DataFrame(rows_vy)
                render_renkli_tablo(df_vy, sol=["Firmalar"], kisalt={"Firmalar": 60})
    
                # Tarih seçip sil
                @st.dialog("🗑️ Belirli Bir Tarihin Firma Stok Verisini Sil", width="large")
                def _dlg_tarih_sil():
                    st.caption("⚠️ Seçilen tarihe ait firma stok verileri silinir. Ürün listesi ve satın alma geçmişi etkilenmez.")
                    sil_tarih = st.selectbox("Silinecek Tarih", sorted(tarih_firma.keys(), reverse=True), key="vy_sil_tarih")
                    if st.button("🗑️ Bu Tarihin Verisini Sil", type="secondary", key="vy_sil_btn"):
                        sb_vy.table("firma_stok").delete().eq("yukleme_tarihi", sil_tarih).execute()
                        st.toast(f"✅ {sil_tarih} tarihli firma stok verisi silindi.")
                        st.rerun()
                if st.button("🗑️ Belirli Bir Tarihin Firma Stok Verisini Sil", key="btn_tarih_sil", use_container_width=True):
                    _dlg_tarih_sil()
    
        except Exception as e:
            _log.warning("Hata: %s", e)
            st.warning(f"Geçmiş yüklemeler yüklenemedi: {e}")
