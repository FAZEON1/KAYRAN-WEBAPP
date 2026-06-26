from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import cm
from datetime import datetime
from .analitik import dashboard_hesapla
from .database import get_siparis_onerileri
from shared.utils import tr_today, tr_now, tr_today_iso, pdf_turkce_font, pdf_stilleri_turkcele

RENKLER = {
    "kirmizi": "FFCCCC",
    "turuncu": "FFD580",
    "sari": "FFFF99",
    "yesil": "CCFFCC",
    "yok": "FFFFFF",
}

PERFORMANS_RENKLER = {
    "Çok İyi": "CCFFCC",
    "İyi": "FFFACD",
    "Düşük": "FFCCCC",
    "veri yok": "F0F0F0",
}

def thin_border():
    thin = Side(style='thin')
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def excel_rapor_olustur(kayit_yolu):
    try:
        data = dashboard_hesapla()
        wb = Workbook()
        
        # ---- SHEET 1: DASHBOARD ----
        ws = wb.active
        ws.title = "Dashboard"
        
        baslik_font = Font(bold=True, color="FFFFFF", size=10)
        baslik_fill = PatternFill("solid", start_color="1F4E79")
        merkez = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        basliklar = [
            "SKU", "Ürün Adı", "Kategori", "Marka",
            "Bizim Stok", "Trendyol Stok", "Stok Yaşı (Gün)",
            "Firma", "Firma Stok", "Haftalık Satış",
            "Kaç Günlük Satış", "Performans", "Sipariş Uyarısı"
        ]
        
        for i, b in enumerate(basliklar, 1):
            cell = ws.cell(row=1, column=i, value=b)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = merkez
            cell.border = thin_border()
        ws.row_dimensions[1].height = 35
        ws.freeze_panes = "A2"
        
        satir = 2
        for urun in data:
            for fd in urun["firma_detay"]:
                stok_fill = PatternFill("solid", start_color=RENKLER.get(urun["stok_renk"], "FFFFFF"))
                gun_fill = PatternFill("solid", start_color=RENKLER.get(fd["gun_renk"], "FFFFFF"))
                perf_fill = PatternFill("solid", start_color=PERFORMANS_RENKLER.get(fd["performans"], "FFFFFF"))
                
                degerler = [
                    urun["sku"],
                    urun["urun_adi"],
                    urun["kategori"],
                    urun["marka"],
                    urun["bizim_stok"],
                    urun["trendyol_stok"],
                    urun["stok_gun"],
                    fd["firma"],
                    fd["stok"],
                    fd["satis"],
                    fd["gun_sayisi"] if fd["gun_sayisi"] is not None else "-",
                    fd["performans"],
                    "⚠️ SİPARİŞ ÖNERİ!" if fd["siparis_uyarisi"] else "",
                ]
                
                for j, val in enumerate(degerler, 1):
                    cell = ws.cell(row=satir, column=j, value=val)
                    cell.border = thin_border()
                    cell.alignment = Alignment(vertical="center")
                    
                    if j == 7:  # Stok yaşı
                        cell.fill = stok_fill
                    elif j == 11:  # Kaç günlük satış
                        cell.fill = gun_fill
                    elif j == 12:  # Performans
                        cell.fill = perf_fill
                    elif j == 13 and fd["siparis_uyarisi"]:  # Uyarı
                        cell.fill = PatternFill("solid", start_color="FF0000")
                        cell.font = Font(bold=True, color="FFFFFF")
                
                satir += 1
        
        genislikler = [12, 30, 15, 15, 12, 14, 14, 12, 12, 14, 15, 12, 18]
        for i, g in enumerate(genislikler, 1):
            ws.column_dimensions[get_column_letter(i)].width = g
        
        # ---- SHEET 2: STOK YAYILIMI ----
        ws2 = wb.create_sheet("Stok Yayılımı")
        yayilim_baslik = ["SKU", "Ürün Adı", "Bizim Stok", "TRENDYOL", "ITOPYA", "HB", "VATAN", "MONDAY", "KANAL", "DİĞER", "Toplam Kanal Stok"]
        for i, b in enumerate(yayilim_baslik, 1):
            cell = ws2.cell(row=1, column=i, value=b)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = merkez
            cell.border = thin_border()
        
        for satir_no, urun in enumerate(data, 2):
            y = urun["yayilim"]
            toplam = sum(y.values())
            row_data = [
                urun["sku"], urun["urun_adi"], urun["bizim_stok"],
                y.get("TRENDYOL", 0), y.get("ITOPYA", 0), y.get("HB", 0),
                y.get("VATAN", 0), y.get("MONDAY", 0), y.get("KANAL", 0),
                y.get("DIGER", 0), toplam
            ]
            for j, val in enumerate(row_data, 1):
                cell = ws2.cell(row=satir_no, column=j, value=val)
                cell.border = thin_border()
        
        for i, g in enumerate([12, 30, 12, 12, 12, 10, 12, 12, 12, 10, 16], 1):
            ws2.column_dimensions[get_column_letter(i)].width = g
        
        # ---- SHEET 3: SİPARİŞ ÖNERİLERİ ----
        ws3 = wb.create_sheet("Sipariş Önerileri")
        sp_baslik = ["ID", "Firma", "SKU", "Ürün Adı", "Önerilen Miktar", "Durum", "Oluşturma Tarihi", "Onay Tarihi"]
        for i, b in enumerate(sp_baslik, 1):
            cell = ws3.cell(row=1, column=i, value=b)
            cell.font = baslik_font
            cell.fill = PatternFill("solid", start_color="1B5E20")
            cell.alignment = merkez
            cell.border = thin_border()
        
        onerileri = get_siparis_onerileri()
        for satir_no, sp in enumerate(onerileri, 2):
            row_data = [sp["id"], sp["firma"], sp["sku"], sp["urun_adi"],
                        sp["oneri_miktari"], sp["durum"], sp["olusturma_tarihi"], sp.get("onay_tarihi", "")]
            for j, val in enumerate(row_data, 1):
                cell = ws3.cell(row=satir_no, column=j, value=val)
                cell.border = thin_border()
                if sp["durum"] == "onaylandi":
                    cell.fill = PatternFill("solid", start_color="CCFFCC")
                elif sp["durum"] == "reddedildi":
                    cell.fill = PatternFill("solid", start_color="FFCCCC")
        
        wb.save(kayit_yolu)
        return True, f"Excel raporu oluşturuldu: {kayit_yolu}"
    except Exception as e:
        return False, f"Excel raporu oluşturulamadı: {str(e)}"


def pdf_rapor_olustur(kayit_yolu):
    try:
        data = dashboard_hesapla()
        doc = SimpleDocTemplate(
            kayit_yolu,
            pagesize=landscape(A4),
            rightMargin=1*cm, leftMargin=1*cm,
            topMargin=1.5*cm, bottomMargin=1*cm
        )
        
        styles = getSampleStyleSheet()
        PDF_NORMAL, PDF_BOLD = pdf_turkce_font()
        pdf_stilleri_turkcele(styles, PDF_NORMAL, PDF_BOLD)
        baslik_style = ParagraphStyle('baslik', parent=styles['Title'], fontName=PDF_BOLD, fontSize=16, spaceAfter=12, textColor=colors.HexColor('#1F4E79'))
        alt_baslik_style = ParagraphStyle('altbaslik', parent=styles['Heading2'], fontName=PDF_BOLD, fontSize=11, spaceAfter=6, textColor=colors.HexColor('#2E7D32'))
        
        story = []
        story.append(Paragraph("Ürün Yönetimi - Stok Takip Raporu", baslik_style))
        story.append(Paragraph(f"Oluşturulma Tarihi: {tr_now().strftime('%d.%m.%Y %H:%M')}", styles['Normal']))
        story.append(Spacer(1, 0.5*cm))
        
        # Ana tablo başlıkları
        tablo_baslik = ["SKU", "Ürün Adı", "Biz.\nStok", "Stok\nYaşı", "Firma", "Firma\nStok", "Haftalık\nSatış", "Kaç Gün", "Performans", "Uyarı"]
        tablo_veri = [tablo_baslik]
        
        PDF_RENKLER = {
            "kirmizi": colors.HexColor('#FFCCCC'),
            "turuncu": colors.HexColor('#FFD580'),
            "sari": colors.HexColor('#FFFF99'),
            "yesil": colors.HexColor('#CCFFCC'),
            "yok": colors.white,
        }
        
        satir_stilleri = []
        satir_no = 1
        
        for urun in data:
            for fd in urun["firma_detay"]:
                satir = [
                    urun["sku"],
                    urun["urun_adi"][:25] + ("..." if len(urun["urun_adi"]) > 25 else ""),
                    str(urun["bizim_stok"]),
                    f"{urun['stok_gun']}g",
                    fd["firma"],
                    str(fd["stok"]),
                    str(fd["satis"]),
                    f"{fd['gun_sayisi']}g" if fd["gun_sayisi"] is not None else "-",
                    fd["performans"],
                    "⚠️" if fd["siparis_uyarisi"] else "",
                ]
                tablo_veri.append(satir)
                
                # Stok yaşı rengi (sütun 3)
                stok_renk = PDF_RENKLER.get(urun["stok_renk"], colors.white)
                satir_stilleri.append(('BACKGROUND', (3, satir_no), (3, satir_no), stok_renk))
                
                # Kaç günlük satış rengi (sütun 7)
                gun_renk = PDF_RENKLER.get(fd["gun_renk"], colors.white)
                satir_stilleri.append(('BACKGROUND', (7, satir_no), (7, satir_no), gun_renk))
                
                # Sipariş uyarısı
                if fd["siparis_uyarisi"]:
                    satir_stilleri.append(('BACKGROUND', (9, satir_no), (9, satir_no), colors.HexColor('#FF4444')))
                    satir_stilleri.append(('TEXTCOLOR', (9, satir_no), (9, satir_no), colors.white))
                
                satir_no += 1
        
        col_widths = [2.5*cm, 6*cm, 1.5*cm, 1.5*cm, 2*cm, 1.5*cm, 2*cm, 1.8*cm, 2.5*cm, 1.5*cm]
        
        tablo = Table(tablo_veri, colWidths=col_widths, repeatRows=1)
        tablo_stil = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), PDF_BOLD),
            ('FONTNAME', (0, 1), (-1, -1), PDF_NORMAL),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
        ] + satir_stilleri)
        tablo.setStyle(tablo_stil)
        
        story.append(tablo)
        doc.build(story)
        return True, f"PDF raporu oluşturuldu: {kayit_yolu}"
    except Exception as e:
        return False, f"PDF raporu oluşturulamadı: {str(e)}"


# ══════════════════════════════════════════════════════════════════════
# TÜM ÜRÜNLER ÖZET — filtrelenmiş liste raporu (Excel / PDF)
# ══════════════════════════════════════════════════════════════════════
TUM_URUN_KOLONLAR = [
    ("SKU", "SKU"),
    ("Ürün Adı", "Ürün Adı"),
    ("Kategori", "Kategori"),
    ("Marka", "Marka"),
    ("Stok Yaşı (gün)", "_stok_yas"),
    ("G5F Depo", "G5F Depo"),
    ("Toplam Stok", "Toplam"),
    ("Paçal FOB ($)", "FOB ($)"),
    ("Son FOB ($)", "Son FOB ($)"),
    ("Maliyet %", "Maliyet %"),
    ("Paçal Maliyet ($)", "Final Cost ($)"),
    ("Son Maliyet ($)", "Son Maliyet ($)"),
    ("Satış ($)", "Satış ($)"),
    ("Net Marj (%)", "Net Marj (%)"),
    ("Net Kâr ($)", "Net Kar ($)"),
]


def tum_urunler_excel(rows, kayit_yolu, meta=""):
    """Filtrelenmiş Tüm Ürünler özetini renkli Excel olarak yazar."""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Tüm Ürünler"
        ws["A1"] = "KAYRAN — Tüm Ürünler Özet Raporu"
        ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
        ws["A2"] = f"{meta}  ·  {tr_now().strftime('%d.%m.%Y %H:%M')}  ·  {len(rows)} ürün"
        ws["A2"].font = Font(size=10, color="666666")

        basliklar = [b for b, _ in TUM_URUN_KOLONLAR]
        hdr_row = 4
        head_fill = PatternFill("solid", start_color="1F4E79")
        head_font = Font(bold=True, color="FFFFFF", size=10)
        for ci, b in enumerate(basliklar, start=1):
            c = ws.cell(row=hdr_row, column=ci, value=b)
            c.fill = head_fill
            c.font = head_font
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = thin_border()

        para_kol = {"Paçal FOB ($)", "Son FOB ($)", "Paçal Maliyet ($)", "Son Maliyet ($)", "Satış ($)", "Net Kâr ($)"}
        pct_kol = {"Maliyet %", "Net Marj (%)"}
        sayi_kol = {"G5F Depo", "Toplam Stok", "Stok Yaşı (gün)"}

        for ri, r in enumerate(rows, start=hdr_row + 1):
            stok_renk = r.get("_stok_renk", "yok")
            for ci, (b, key) in enumerate(TUM_URUN_KOLONLAR, start=1):
                v = r.get(key)
                if v is None or v == "":
                    val = ""
                elif b in para_kol or b in pct_kol:
                    val = float(v)
                elif b in sayi_kol:
                    val = int(v or 0)
                else:
                    val = v
                c = ws.cell(row=ri, column=ci, value=val)
                c.border = thin_border()
                c.font = Font(size=10)
                if b in para_kol:
                    c.number_format = '$#,##0.00'
                    c.alignment = Alignment(horizontal="right")
                elif b in pct_kol:
                    c.number_format = '0.0"%"'
                    c.alignment = Alignment(horizontal="right")
                elif b in ("G5F Depo", "Toplam Stok"):
                    c.alignment = Alignment(horizontal="right")
                if b == "Stok Yaşı (gün)":
                    c.alignment = Alignment(horizontal="center")
                    if stok_renk in RENKLER and stok_renk != "yok":
                        c.fill = PatternFill("solid", start_color=RENKLER[stok_renk])
                if b == "Net Kâr ($)" and isinstance(val, (int, float)) and val < 0:
                    c.font = Font(size=10, color="C00000", bold=True)

        for ci, w in enumerate([12, 38, 13, 11, 11, 9, 9, 11, 11, 9, 13, 13, 10, 10, 11], start=1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = "A5"
        wb.save(kayit_yolu)
        return True, "ok"
    except Exception as e:
        return False, f"Excel oluşturulamadı: {str(e)}"


def tum_urunler_pdf(rows, kayit_yolu, meta=""):
    """Filtrelenmiş Tüm Ürünler özetini A4 yatay kompakt PDF olarak yazar."""
    try:
        doc = SimpleDocTemplate(kayit_yolu, pagesize=landscape(A4),
                                leftMargin=1 * cm, rightMargin=1 * cm,
                                topMargin=1 * cm, bottomMargin=1 * cm)
        styles = getSampleStyleSheet()
        PDF_NORMAL, PDF_BOLD = pdf_turkce_font()
        pdf_stilleri_turkcele(styles, PDF_NORMAL, PDF_BOLD)
        h = ParagraphStyle("h", parent=styles["Title"], fontName=PDF_BOLD, fontSize=15,
                           textColor=colors.HexColor("#1F4E79"))
        sub = ParagraphStyle("sub", parent=styles["Normal"], fontName=PDF_NORMAL, fontSize=9,
                             textColor=colors.HexColor("#666666"))
        elements = [
            Paragraph("KAYRAN — Tüm Ürünler Özet Raporu", h),
            Paragraph(f"{meta} · {tr_now().strftime('%d.%m.%Y %H:%M')} · {len(rows)} ürün", sub),
            Spacer(1, 0.3 * cm),
        ]
        pdf_kol = [
            ("SKU", "SKU"), ("Ürün", "Ürün Adı"), ("Kat.", "Kategori"),
            ("Yaş", "_stok_yas"), ("G5F", "G5F Depo"), ("Top.", "Toplam"),
            ("PaçFOB$", "FOB ($)"), ("SonFOB$", "Son FOB ($)"),
            ("Paçal$", "Final Cost ($)"), ("SonMal$", "Son Maliyet ($)"),
            ("Satış$", "Satış ($)"),
            ("Marj%", "Net Marj (%)"), ("Net Kâr$", "Net Kar ($)"),
        ]
        para_keys = {"FOB ($)", "Son FOB ($)", "Final Cost ($)", "Son Maliyet ($)", "Satış ($)", "Net Kar ($)"}
        pct_keys = {"Net Marj (%)"}

        def _f2(v, para=False, pct=False):
            if v is None or v == "":
                return "—"
            try:
                if para:
                    return f"${float(v):,.0f}"
                if pct:
                    return f"%{float(v):.0f}"
                return str(v)
            except Exception:
                return str(v)

        data = [[k for k, _ in pdf_kol]]
        for r in rows:
            ad = str(r.get("Ürün Adı", "") or "")
            satir = []
            for k, key in pdf_kol:
                v = r.get(key)
                if key == "Ürün Adı":
                    satir.append(ad if len(ad) <= 28 else ad[:27] + "…")
                elif key in para_keys:
                    satir.append(_f2(v, para=True))
                elif key in pct_keys:
                    satir.append(_f2(v, pct=True))
                elif key == "_stok_yas":
                    satir.append(f"{int(v or 0)}g" if r.get("_stok_renk", "yok") != "yok" else "—")
                else:
                    satir.append(_f2(v))
            data.append(satir)

        tbl = Table(data, repeatRows=1)
        stil = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("FONTNAME", (0, 0), (-1, 0), PDF_BOLD),
            ("FONTNAME", (0, 1), (-1, -1), PDF_NORMAL),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F6FA")]),
            ("ALIGN", (3, 1), (-1, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]
        for i, r in enumerate(rows, start=1):
            nk = r.get("Net Kar ($)")
            if isinstance(nk, (int, float)) and nk < 0:
                stil.append(("TEXTCOLOR", (len(pdf_kol) - 1, i), (len(pdf_kol) - 1, i),
                             colors.HexColor("#C00000")))
        tbl.setStyle(TableStyle(stil))
        elements.append(tbl)
        doc.build(elements)
        return True, "ok"
    except Exception as e:
        return False, f"PDF oluşturulamadı: {str(e)}"
