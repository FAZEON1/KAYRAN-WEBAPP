"""
HESAP MAKİNESİ — Ürün Karlılık & Break-Even Modülü
Sadece ibrahim kullanıcısına görünür.
"""
import streamlit as st
import json
import io
from datetime import datetime

# ─────────────────────────────────────────────────
# SUPABASE
# ─────────────────────────────────────────────────
def _get_supabase():
    try:
        from supabase import create_client
        url = st.secrets["supabase"]["url"]
        key = st.secrets["supabase"]["key"]
        return create_client(url, key)
    except Exception:
        return None

def _hesap_kaydet(kullanici, urun_adi, alis_fiyati, masraflar, senaryolar, notlar=""):
    sb = _get_supabase()
    if not sb:
        return False, "Supabase baglanti yok"
    try:
        data = {
            "kullanici": kullanici,
            "urun_adi": urun_adi or "",
            "alis_fiyati": float(alis_fiyati),
            "masraflar": float(masraflar),
            "toplam_maliyet": float(alis_fiyati) + float(masraflar),
            "senaryolar": json.dumps(senaryolar, ensure_ascii=False),
            "notlar": notlar or "",
            "olusturma_tarihi": datetime.now().isoformat()
        }
        sb.table("hm_urun_karliligi").insert(data).execute()
        return True, "ok"
    except Exception as e:
        return False, str(e)

def _hesaplari_getir(kullanici):
    sb = _get_supabase()
    if not sb:
        return []
    try:
        r = sb.table("hm_urun_karliligi").select("*").eq("kullanici", kullanici).order("olusturma_tarihi", desc=True).execute()
        return r.data or []
    except Exception:
        return []

def _hesap_sil(kayit_id):
    sb = _get_supabase()
    if not sb:
        return False
    try:
        sb.table("hm_urun_karliligi").delete().eq("id", kayit_id).execute()
        return True
    except Exception:
        return False

def _hesap_guncelle(kayit_id, notlar):
    sb = _get_supabase()
    if not sb:
        return False
    try:
        sb.table("hm_urun_karliligi").update({"notlar": notlar}).eq("id", kayit_id).execute()
        return True
    except Exception:
        return False

# ─────────────────────────────────────────────────
# HESAPLAMA FONKSİYONLARI
# ─────────────────────────────────────────────────
def hesapla_satis_fiyati(maliyet, marj_yuzdesi):
    """Top-down marj: Fiyat = Maliyet / (1 - Marj%)"""
    if marj_yuzdesi >= 100:
        return None
    return maliyet / (1 - marj_yuzdesi / 100)

def hesapla_marj(maliyet, satis_fiyati):
    """Top-down marj yüzdesi: (Fiyat - Maliyet) / Fiyat * 100"""
    if satis_fiyati <= 0:
        return None
    return ((satis_fiyati - maliyet) / satis_fiyati) * 100

def hesapla_kar(maliyet, satis_fiyati):
    return satis_fiyati - maliyet

def uygula_indirim(satis_fiyati, indirim_yuzdesi=0, indirim_tutar=0):
    fiyat = satis_fiyati
    if indirim_yuzdesi > 0:
        fiyat = fiyat * (1 - indirim_yuzdesi / 100)
    elif indirim_tutar > 0:
        fiyat = fiyat - indirim_tutar
    return fiyat

# ─────────────────────────────────────────────────
# CSS
# ─────────────────────────────────────────────────
def _hm_css():
    return """
<style>
.hm-card{background:rgba(255,255,255,0.03);border:1px solid rgba(255,255,255,0.08);border-radius:16px;padding:20px 22px;margin-bottom:16px;}
.hm-card:hover{border-color:rgba(99,102,241,0.3);}
.hm-senaryo-kart{background:rgba(99,102,241,0.06);border:1px solid rgba(99,102,241,0.15);border-radius:12px;padding:14px 16px;margin-bottom:10px;}
.hm-kar-pozitif{color:#10B981!important;font-weight:700;}
.hm-kar-negatif{color:#EF4444!important;font-weight:700;}
.hm-label{font-size:10px;color:#64748B;letter-spacing:1.5px;text-transform:uppercase;font-weight:700;margin-bottom:4px;}
.hm-deger{font-size:22px;font-weight:800;color:#FFFFFF;font-family:JetBrains Mono,monospace;}
.hm-separator{height:1px;background:rgba(255,255,255,0.06);margin:20px 0;}
[data-testid="stNumberInput"] label,[data-testid="stTextInput"] label,[data-testid="stTextArea"] label,[data-testid="stSelectbox"] label{color:#CBD5E1!important;font-size:12px!important;font-weight:600!important;}
[data-testid="stNumberInput"] input,[data-testid="stTextInput"] input{background:rgba(255,255,255,0.04)!important;border:1px solid rgba(255,255,255,0.12)!important;color:#FFFFFF!important;border-radius:10px!important;}
[data-testid="stTextArea"] textarea{background:rgba(255,255,255,0.04)!important;border:1px solid rgba(255,255,255,0.12)!important;color:#FFFFFF!important;border-radius:10px!important;}
[data-testid="stNumberInput"] input:focus,[data-testid="stTextInput"] input:focus{border-color:#6366F1!important;box-shadow:0 0 0 3px rgba(99,102,241,0.15)!important;}
</style>
"""

# ─────────────────────────────────────────────────
# ÜRÜN KARLILIK BÖLÜMÜ
# ─────────────────────────────────────────────────
def _urun_karlilik_bolumu():
    aktif_kullanici = st.session_state.get("aktif_kullanici", "")

    st.markdown(
        '<div style="display:inline-block;padding:6px 14px;background:rgba(99,102,241,0.12);border:1px solid rgba(99,102,241,0.25);border-radius:20px;margin-bottom:18px">'
        '<span style="color:#A5B4FC;font-size:11px;font-weight:600;letter-spacing:1px;text-transform:uppercase">🧮 Ürün Karlılık Hesaplayıcı</span>'
        '</div>',
        unsafe_allow_html=True
    )
    st.markdown('<p style="color:#94A3B8;font-size:13px;margin-bottom:24px">Ürün maliyeti ve satış fiyatını girerek net marjı hesaplayın. Birden fazla fiyat senaryosu ekleyebilir, indirim senaryoları oluşturabilirsiniz.</p>', unsafe_allow_html=True)

    # SESSION STATE INIT
    if "hm_senaryolar" not in st.session_state:
        st.session_state.hm_senaryolar = [
            {"id": 1, "ad": "Senaryo 1", "fiyat": 0.0, "marj": 0.0,
             "indirim_tipi": "Yok", "indirim_yuzdesi": 0.0, "indirim_tutar": 0.0}
        ]
    if "hm_senaryo_sayac" not in st.session_state:
        st.session_state.hm_senaryo_sayac = 1

    # MALİYET GİRİŞİ
    st.markdown('<div class="hm-card">', unsafe_allow_html=True)
    st.markdown('<div style="color:#FFFFFF;font-size:15px;font-weight:700;margin-bottom:16px">💰 Maliyet Bilgileri (KDV Hariç)</div>', unsafe_allow_html=True)

    col1, col2, col3 = st.columns([1.5, 1.5, 1])
    with col1:
        urun_adi = st.text_input("Ürün Adı / Kodu (opsiyonel)", placeholder="Örn: Monitor XG27AQM", key="hm_urun_adi")
    with col2:
        alis_fiyati = st.number_input("Alış Fiyatı ($)", min_value=0.0, value=0.0, step=0.01, format="%.2f", key="hm_alis")
    with col3:
        masraflar = st.number_input("Ek Masraflar ($)", min_value=0.0, value=0.0, step=0.01, format="%.2f",
                                     help="Kargo, komisyon, ambalaj vb.", key="hm_masraf")

    toplam_maliyet = alis_fiyati + masraflar

    colm1, colm2, colm3 = st.columns(3)
    with colm1:
        st.markdown(f'<div class="hm-label">Alış Fiyatı</div><div class="hm-deger">{alis_fiyati:,.2f}</div>', unsafe_allow_html=True)
    with colm2:
        st.markdown(f'<div class="hm-label">Ek Masraflar</div><div class="hm-deger">{masraflar:,.2f}</div>', unsafe_allow_html=True)
    with colm3:
        st.markdown(f'<div class="hm-label">Toplam Maliyet</div><div class="hm-deger" style="color:#60A5FA">{toplam_maliyet:,.2f}</div>', unsafe_allow_html=True)

    st.markdown('</div>', unsafe_allow_html=True)

    # SENARYO YÖNETİMİ
    st.markdown('<div class="hm-card">', unsafe_allow_html=True)
    st.markdown('<div style="color:#FFFFFF;font-size:15px;font-weight:700;margin-bottom:4px">📊 Fiyat Senaryoları</div>', unsafe_allow_html=True)
    st.markdown('<p style="color:#94A3B8;font-size:12px;margin-bottom:16px">Her senaryo için satış fiyatı veya marj yüzdesi girin — biri girilince diğeri otomatik hesaplanır.</p>', unsafe_allow_html=True)

    senaryolar_sonuc = []

    for idx, sen in enumerate(st.session_state.hm_senaryolar):
        sen_id = sen["id"]
        st.markdown(f'<div class="hm-senaryo-kart"><div style="color:#A5B4FC;font-size:11px;font-weight:700;letter-spacing:1px;margin-bottom:10px">📌 {O}sen["ad"]}</div>', unsafe_allow_html=True)

        cs1, cs2, cs3, cs4, cs5 = st.columns([1.5, 1, 1, 1, 0.5])

        with cs1:
            sen_ad = st.text_input("Senaryo Adı", value=sen["ad"], key=f"sen_ad_{O}sen_id}")
        with cs2:
            sen_fiyat = st.number_input("Satış Fiyatı ($)", min_value=0.0, value=float(sen["fiyat"]), step=0.01, format="%.2f", key=f"sen_fiyat_{O}sen_id}")
        with cs3:
            sen_marj = st.number_input("Marj (%)", min_value=0.0, max_value=99.9, value=float(sen["marj"]), step=0.1, format="%.1f", key=f"sen_marj_{O}sen_id}")
        with cs4:
            indirim_tipi = st.selectbox("İndirim Tipi", ["Yok", "Yüzde (%)", "Tutar ($)"], index=["Yok","Yüzde (%)","Tutar ($)"].index(sen.get("indirim_tipi","Yok")), key=f"sen_ind_tip_{O}sen_id}")
        with cs5:
            sil_btn = st.button("🗑️", key=f"sen_sil_{O}sen_id}", help="Senaryoyu sil")

        # Fiyat-Marj senkronizasyonu (son değişen hangisiyse ona göre hesapla)
        prev_fiyat = sen.get("_prev_fiyat", sen["fiyat"])
        prev_marj = sen.get("_prev_marj", sen["marj"])

        if toplam_maliyet > 0:
            fiyat_degisti = abs(sen_fiyat - prev_fiyat) > 0.001
            marj_degisti = abs(sen_marj - prev_marj) > 0.001

            if fiyat_degisti and sen_fiyat > 0:
                hesaplanan_marj = hesapla_marj(toplam_maliyet, sen_fiyat)
                if hesaplanan_marj is not None:
                    sen_marj = hesaplanan_marj
            elif marj_degisti and sen_marj > 0:
                hesaplanan_fiyat = hesapla_satis_fiyati(toplam_maliyet, sen_marj)
                if hesaplanan_fiyat is not None:
                    sen_fiyat = hesaplanan_fiyat

        # İndirim girişi
        ind_yuzdesi = 0.0
        ind_tutar = 0.0
        if indirim_tipi == "Yüzde (%)":
            ind_yuzdesi = st.number_input("İndirim Yüzdesi (%)", min_value=0.0, max_value=99.9, value=float(sen.get("indirim_yuzdesi", 0)), step=0.1, format="%.1f", key=f"ind_yuz_{O}sen_id}")
        elif indirim_tipi == "Tutar ($)":
            ind_tutar = st.number_input("İndirim Tutarı ($)", min_value=0.0, value=float(sen.get("indirim_tutar", 0)), step=0.01, format="%.2f", key=f"ind_tut_{O}sen_id}")

        # İndirimli fiyat hesapla
        ind_fiyat = uygula_indirim(sen_fiyat, ind_yuzdesi, ind_tutar)
        ind_marj = hesapla_marj(toplam_maliyet, ind_fiyat) if ind_fiyat > 0 and toplam_maliyet > 0 else 0
        kar = hesapla_kar(toplam_maliyet, ind_fiyat)

        # Sonuç göster
        if toplam_maliyet > 0 and sen_fiyat > 0:
            kar_renk = "#10B981" if kar > 0 else "#EF4444" if kar < 0 else "#F59E0B"
            ind_goster = indirim_tipi != "Yok"
            cr1, cr2, cr3, cr4 = st.columns(4)
            with cr1:
                st.markdown(f'<div class="hm-label">Satış Fiyatı</div><div class="hm-deger">{sen_fiyat:,.2f}</div>', unsafe_allow_html=True)
            with cr2:
                if ind_goster:
                    st.markdown(f'<div class="hm-label">İnd. Fiyat</div><div class="hm-deger" style="color:#F59E0B">{ind_fiyat:,.2f}</div>', unsafe_allow_html=True)
                else:
                    st.markdown('<div class="hm-label">İndirim</div><div style="color:#64748B;font-size:14px">Yok</div>', unsafe_allow_html=True)
            with cr3:
                st.markdown(f'<div class="hm-label">Net Kâr</div><div class="hm-deger" style="color:{O}kar_renk}">{kar:,.2f}</div>', unsafe_allow_html=True)
            with cr4:
                st.markdown(f'<div class="hm-label">Marj</div><div class="hm-deger" style="color:{O}kar_renk}">%{O}ind_marj:.1f}</div>', unsafe_allow_html=True)

        # State güncelle
        if not sil_btn:
            senaryolar_sonuc.append({
                "id": sen_id,
                "ad": sen_ad,
                "fiyat": sen_fiyat,
                "marj": sen_marj,
                "indirim_tipi": indirim_tipi,
                "indirim_yuzdesi": ind_yuzdesi,
                "indirim_tutar": ind_tutar,
                "_prev_fiyat": sen_fiyat,
                "_prev_marj": sen_marj
            })

        st.markdown('</div>', unsafe_allow_html=True)

    st.session_state.hm_senaryolar = senaryolar_sonuc

    # Senaryo ekle butonu
    if st.button("➕ Yeni Senaryo Ekle", key="hm_senaryo_ekle"):
        st.session_state.hm_senaryo_sayac += 1
        yeni_id = st.session_state.hm_senaryo_sayac
        st.session_state.hm_senaryolar.append({
            "id": yeni_id, "ad": f"Senaryo {O}yeni_id}", "fiyat": 0.0, "marj": 0.0,
            "indirim_tipi": "Yok", "indirim_yuzdesi": 0.0, "indirim_tutar": 0.0
        })
        st.rerun()

    st.markdown('</div>', unsafe_allow_html=True)

    # KAYDET VE EXPORT
    st.markdown('<div class="hm-separator"></div>', unsafe_allow_html=True)
    cks1, cks2, cks3 = st.columns([2, 1, 1])
    with cks1:
        notlar_kayit = st.text_area("Not / Yorum (opsiyonel)", placeholder="Bu hesaplamaya dair notunuzu girin...", key="hm_notlar", height=80)
    with cks2:
        st.markdown('<div style="height:28px"></div>', unsafe_allow_html=True)
        kaydet_btn = st.button("💾 Hesaplamayı Kaydet", key="hm_kaydet", use_container_width=True, type="primary")
    with cks3:
        st.markdown('<div style="height:28px"></div>', unsafe_allow_html=True)
        sifirla_btn = st.button("🔄 Formu Sıfırla", key="hm_sifirla", use_container_width=True)

    if kaydet_btn:
        if toplam_maliyet <= 0:
            st.warning("⚠️ Lütfen önce maliyet bilgilerini girin.")
        elif not senaryolar_sonuc:
            st.warning("⚠️ En az bir senaryo eklemelisiniz.")
        else:
            ok, msg = _hesap_kaydet(
                aktif_kullanici, urun_adi, alis_fiyati, masraflar,
                senaryolar_sonuc, notlar_kayit
            )
            if ok:
                st.success("✅ Hesaplama başarıyla kaydedildi!")
                if "hm_gecmis_cache" in st.session_state:
                    del st.session_state["hm_gecmis_cache"]
            else:
                st.error(f"❌ Kaydetme hatası: {O}msg}")

    if sifirla_btn:
        for k in ["hm_urun_adi", "hm_alis", "hm_masraf", "hm_senaryolar", "hm_senaryo_sayac", "hm_notlar"]:
            if k in st.session_state:
                del st.session_state[k]
        st.rerun()

# ─────────────────────────────────────────────────
# GEÇMİŞ HESAPLAMALAR
# ─────────────────────────────────────────────────
def _gecmis_hesaplamalar():
    aktif_kullanici = st.session_state.get("aktif_kullanici", "")

    st.markdown('<div class="hm-separator"></div>', unsafe_allow_html=True)
    st.markdown(
        '<div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:16px">'
        '<div style="color:#FFFFFF;font-size:15px;font-weight:700">📋 Kaydedilmiş Hesaplamalar</div>'
        '</div>',
        unsafe_allow_html=True
    )

    # Cache'den veya DB'den yükle
    if "hm_gecmis_cache" not in st.session_state:
        with st.spinner("Yükleniyor..."):
            st.session_state.hm_gecmis_cache = _hesaplari_getir(aktif_kullanici)

    kayitlar = st.session_state.hm_gecmis_cache

    # Yenile ve Excel export butonları
    cg1, cg2, cg3 = st.columns([1, 1, 4])
    with cg1:
        if st.button("🔄 Yenile", key="hm_gecmis_yenile"):
            if "hm_gecmis_cache" in st.session_state:
                del st.session_state["hm_gecmis_cache"]
            st.rerun()
    with cg2:
        if kayitlar:
            excel_data = _kayitlari_excel_olustur(kayitlar)
            st.download_button(
                label="📥 Excel İndir",
                data=excel_data,
                file_name=f"hesap_makinesi_{O}datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="hm_excel_indir"
            )

    if not kayitlar:
        st.markdown(
            '<div style="text-align:center;padding:32px;color:#64748B;font-size:13px">'
            '📭 Henüz kaydedilmiş hesaplama yok.</div>',
            unsafe_allow_html=True
        )
        return

    # Tablo başlığı
    th1, th2, th3, th4, th5, th6 = st.columns([2, 1.2, 1.2, 1.2, 1.5, 1])
    headers = ["Ürün / Kod", "Alış ($)", "Masraflar ($)", "Top. Maliyet ($)", "Tarih", "İşlem"]
    header_cols = [th1, th2, th3, th4, th5, th6]
    for col, h in zip(header_cols, headers):
        with col:
            st.markdown(f'<div class="hm-label">{O}h}</div>', unsafe_allow_html=True)

    # Satırlar
    for kayit in kayitlar:
        rid = kayit.get("id")
        uad = kayit.get("urun_adi") or "—"
        ali = kayit.get("alis_fiyati", 0)
        mas = kayit.get("masraflar", 0)
        top = kayit.get("toplam_maliyet", 0)
        tar_raw = kayit.get("olusturma_tarihi", "")
        try:
            tar = datetime.fromisoformat(tar_raw[:19]).strftime("%d.%m.%Y %H:%M")
        except Exception:
            tar = tar_raw[:16] if tar_raw else "—"

        rc1, rc2, rc3, rc4, rc5, rc6 = st.columns([2, 1.2, 1.2, 1.2, 1.5, 1])
        with rc1:
            st.markdown(f'<div style="color:#E2E8F0;font-size:13px;padding:8px 0">{O}uad}</div>', unsafe_allow_html=True)
        with rc2:
            st.markdown(f'<div style="color:#CBD5E1;font-size:13px;padding:8px 0">{ali:,.2f}</div>', unsafe_allow_html=True)
        with rc3:
            st.markdown(f'<div style="color:#CBD5E1;font-size:13px;padding:8px 0">{mas:,.2f}</div>', unsafe_allow_html=True)
        with rc4:
            st.markdown(f'<div style="color:#60A5FA;font-size:13px;font-weight:600;padding:8px 0">{top:,.2f}</div>', unsafe_allow_html=True)
        with rc5:
            st.markdown(f'<div style="color:#94A3B8;font-size:12px;padding:8px 0">{O}tar}</div>', unsafe_allow_html=True)
        with rc6:
            if st.button("🔍", key=f"hm_detay_{O}rid}", help="Detayları görüntüle / düzenle"):
                st.session_state[f"hm_detay_ac_{O}rid}"] = not st.session_state.get(f"hm_detay_ac_{O}rid}", False)

        # Detay paneli
        if st.session_state.get(f"hm_detay_ac_{O}rid}", False):
            _kayit_detay_panel(kayit)

        st.markdown('<div class="hm-separator" style="margin:6px 0"></div>', unsafe_allow_html=True)

def _kayit_detay_panel(kayit):
    rid = kayit.get("id")
    senaryolar_raw = kayit.get("senaryolar", "[]")
    try:
        senaryolar = json.loads(senaryolar_raw) if isinstance(senaryolar_raw, str) else senaryolar_raw
    except Exception:
        senaryolar = []

    toplam_maliyet = kayit.get("toplam_maliyet", 0)

    st.markdown(
        '<div style="background:rgba(99,102,241,0.05);border:1px solid rgba(99,102,241,0.2);'
        'border-radius:12px;padding:16px 20px;margin:8px 0 12px">',
        unsafe_allow_html=True
    )

    # Senaryolar
    if senaryolar:
        st.markdown('<div style="color:#A5B4FC;font-size:12px;font-weight:700;margin-bottom:10px">📊 Senaryolar</div>', unsafe_allow_html=True)
        for sen in senaryolar:
            sen_fiyat = sen.get("fiyat", 0)
            ind_tipi = sen.get("indirim_tipi", "Yok")
            ind_yuz = sen.get("indirim_yuzdesi", 0)
            ind_tut = sen.get("indirim_tutar", 0)
            ind_fiyat = uygula_indirim(sen_fiyat, ind_yuz, ind_tut)
            ind_marj = hesapla_marj(toplam_maliyet, ind_fiyat) if ind_fiyat > 0 and toplam_maliyet > 0 else 0
            kar = hesapla_kar(toplam_maliyet, ind_fiyat)
            kar_renk = "#10B981" if kar > 0 else "#EF4444" if kar < 0 else "#F59E0B"

            sc1, sc2, sc3, sc4 = st.columns(4)
            with sc1:
                st.markdown(f'<div class="hm-label">{O}sen.get("ad","?")}</div><div style="color:#E2E8F0;font-size:14px;font-weight:600">{sen_fiyat:,.2f}</div>', unsafe_allow_html=True)
            with sc2:
                if ind_tipi != "Yok":
                    st.markdown(f'<div class="hm-label">İnd. Fiyat</div><div style="color:#F59E0B;font-size:14px;font-weight:600">{ind_fiyat:,.2f}</div>', unsafe_allow_html=True)
                else:
                    st.markdown('<div class="hm-label">İndirim</div><div style="color:#64748B;font-size:12px">Yok</div>', unsafe_allow_html=True)
            with sc3:
                st.markdown(f'<div class="hm-label">Net Kâr</div><div style="color:{O}kar_renk};font-size:14px;font-weight:600">{kar:,.2f}</div>', unsafe_allow_html=True)
            with sc4:
                st.markdown(f'<div class="hm-label">Marj</div><div style="color:{O}kar_renk};font-size:14px;font-weight:600">%{O}ind_marj:.1f}</div>', unsafe_allow_html=True)

    st.markdown('<div class="hm-separator" style="margin:12px 0"></div>', unsafe_allow_html=True)

    # Not düzenleme
    mevcut_not = kayit.get("notlar", "")
    yeni_not = st.text_area("📝 Not / Yorum", value=mevcut_not, key=f"hm_not_edit_{O}rid}", height=80)

    dp1, dp2, dp3 = st.columns([1, 1, 3])
    with dp1:
        if st.button("💾 Notu Kaydet", key=f"hm_not_kaydet_{O}rid}"):
            if _hesap_guncelle(rid, yeni_not):
                st.success("✅ Not kaydedildi.")
                if "hm_gecmis_cache" in st.session_state:
                    del st.session_state["hm_gecmis_cache"]
                st.rerun()
    with dp2:
        if st.button("🗑️ Kaydı Sil", key=f"hm_kayit_sil_{O}rid}", type="primary"):
            st.session_state[f"hm_sil_onay_{O}rid}"] = True

    if st.session_state.get(f"hm_sil_onay_{O}rid}", False):
        st.warning("⚠️ Bu kaydı kalıcı olarak silmek istiyor musunuz?")
        eonay1, eonay2 = st.columns(2)
        with eonay1:
            if st.button("✅ Evet, Sil", key=f"hm_sil_evet_{O}rid}"):
                if _hesap_sil(rid):
                    if "hm_gecmis_cache" in st.session_state:
                        del st.session_state["hm_gecmis_cache"]
                    st.session_state.pop(f"hm_detay_ac_{O}rid}", None)
                    st.session_state.pop(f"hm_sil_onay_{O}rid}", None)
                    st.rerun()
        with eonay2:
            if st.button("❌ İptal", key=f"hm_sil_iptal_{O}rid}"):
                st.session_state.pop(f"hm_sil_onay_{O}rid}", None)
                st.rerun()

    st.markdown('</div>', unsafe_allow_html=True)


def _kayitlari_excel_olustur(kayitlar):
    """Kayıtları Excel dosyasına dönüştürür."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        # openpyxl yoksa basit CSV döndür
        lines = ["Urun Adi,Alis Fiyati,Masraflar,Toplam Maliyet,Tarih,Notlar"]
        for k in kayitlar:
            lines.append(f'{O}k.get("urun_adi","")},{O}k.get("alis_fiyati",0)},{O}k.get("masraflar",0)},{O}k.get("toplam_maliyet",0)},{O}k.get("olusturma_tarihi","")},{O}k.get("notlar","")}')
        return "\n".join(lines).encode("utf-8")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Karlılık Hesaplamaları"

    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border_side = Side(style="thin", color="374151")
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    headers = ["Ürün Adı/Kodu", "Alış Fiyatı ($)", "Ek Masraflar ($)", "Toplam Maliyet ($)",
               "Senaryo Adı", "Satış Fiyatı ($)", "İndirimli Fiyat ($)", "Net Kâr ($)", "Marj (%)",
               "Tarih", "Notlar"]

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    row_num = 2
    for kayit in kayitlar:
        senaryolar_raw = kayit.get("senaryolar", "[]")
        try:
            sens = json.loads(senaryolar_raw) if isinstance(senaryolar_raw, str) else senaryolar_raw
        except Exception:
            sens = []

        toplam_maliyet = kayit.get("toplam_maliyet", 0)
        tar_raw = kayit.get("olusturma_tarihi", "")
        try:
            tar = datetime.fromisoformat(tar_raw[:19]).strftime("%d.%m.%Y %H:%M")
        except Exception:
            tar = tar_raw[:16] if tar_raw else ""

        if not sens:
            sens = [{"ad": "", "fiyat": 0, "indirim_tipi": "Yok", "indirim_yuzdesi": 0, "indirim_tutar": 0}]

        for sen in sens:
            sen_fiyat = sen.get("fiyat", 0)
            ind_fiyat = uygula_indirim(sen_fiyat, sen.get("indirim_yuzdesi", 0), sen.get("indirim_tutar", 0))
            marj = hesapla_marj(toplam_maliyet, ind_fiyat) if ind_fiyat > 0 and toplam_maliyet > 0 else 0
            kar = hesapla_kar(toplam_maliyet, ind_fiyat)

            row_data = [
                kayit.get("urun_adi", ""),
                kayit.get("alis_fiyati", 0),
                kayit.get("masraflar", 0),
                toplam_maliyet,
                sen.get("ad", ""),
                round(sen_fiyat, 2),
                round(ind_fiyat, 2),
                round(kar, 2),
                round(marj or 0, 2),
                tar,
                kayit.get("notlar", "")
            ]
            for ci, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=ci, value=val)
                cell.border = thin_border
                if ci == 9 and isinstance(val, (int, float)):
                    cell.number_format = "0.00%"
                    cell.value = val / 100
            row_num += 1

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

# ─────────────────────────────────────────────────
# BREAK-EVEN SIDEBAR
# ─────────────────────────────────────────────────
def _breakeven_sidebar():
    with st.sidebar:
        st.markdown('<div class="hm-separator"></div>', unsafe_allow_html=True)
        st.markdown(
            '<div style="font-size:10px;color:#F59E0B;letter-spacing:2px;font-weight:700;text-transform:uppercase;margin:4px 0 10px;padding-left:6px">📈 KIRILMA NOKTASI</div>',
            unsafe_allow_html=True
        )

        # SESSION INIT
        if "be_senaryolar" not in st.session_state:
            st.session_state.be_senaryolar = [{
                "id": 1, "ad": "Senaryo 1",
                "gider_modu": "Toplam",
                "toplam_gider": 0.0,
                "kalemler": [],
                "ortalama_marj": 30.0,
                "mevcut_ciro": 0.0,
                "ort_urun_fiyati": 0.0,
                "periyot": "Aylık"
            }]
        if "be_sayac" not in st.session_state:
            st.session_state.be_sayac = 1

        # Aktif senaryo seçimi
        if len(st.session_state.be_senaryolar) > 1:
            sen_isimleri = [s["ad"] for s in st.session_state.be_senaryolar]
            if "be_aktif_idx" not in st.session_state:
                st.session_state.be_aktif_idx = 0
            aktif_idx = st.session_state.be_aktif_idx
            if aktif_idx >= len(sen_isimleri):
                aktif_idx = 0
        else:
            aktif_idx = 0

        # Senaryo sekmeleri
        for si, sen in enumerate(st.session_state.be_senaryolar):
            sen_id = sen["id"]

            # Senaryo başlığı (tıklanabilir)
            is_aktif = (si == aktif_idx)
            btn_style = "primary" if is_aktif else "secondary"
            if st.button(f'📋 {O}sen["ad"]}', key=f"be_sen_sec_{O}sen_id}", type=btn_style, use_container_width=True):
                st.session_state.be_aktif_idx = si
                st.rerun()

        if is_aktif or len(st.session_state.be_senaryolar) == 1:
            sen = st.session_state.be_senaryolar[aktif_idx]
            sen_id = sen["id"]

            # Senaryo adı
            sen_ad_be = st.text_input("Senaryo Adı", value=sen["ad"], key=f"be_sen_ad_{O}sen_id}")

            # Periyot seçimi
            periyot = st.selectbox("Periyot", ["Günlük", "Haftalık", "Aylık", "Yıllık"], index=["Günlük","Haftalık","Aylık","Yıllık"].index(sen.get("periyot","Aylık")), key=f"be_periyot_{O}sen_id}")

            # Periyot çarpanı (aylık bazda)
            periyot_carpan = {"Günlük": 1/30, "Haftalık": 1/4.33, "Aylık": 1, "Yıllık": 12}
            carpan = periyot_carpan.get(periyot, 1)

            # Gider modu
            gider_modu = st.selectbox("Gider Girişi", ["Toplam", "Kalem Kalem"], index=["Toplam","Kalem Kalem"].index(sen.get("gider_modu","Toplam")), key=f"be_gider_modu_{O}sen_id}")

            if gider_modu == "Toplam":
                toplam_gider_be = st.number_input(f"Aylık Sabit Gider ($)", min_value=0.0, value=float(sen.get("toplam_gider",0)), step=10.0, format="%.2f", key=f"be_toplam_gider_{O}sen_id}")
                kalemler_be = sen.get("kalemler", [])
            else:
                kalemler_be = sen.get("kalemler", [{"ad": "Kira", "tutar": 0.0}, {"ad": "Maaşlar", "tutar": 0.0}, {"ad": "Fatura", "tutar": 0.0}])
                yeni_kalemler = []
                for ki, kalem in enumerate(kalemler_be):
                    ck1, ck2 = st.columns([2, 1])
                    with ck1:
                        kalem_ad = st.text_input("Kalem", value=kalem.get("ad",""), key=f"be_kal_ad_{O}sen_id}_{O}ki}")
                    with ck2:
                        kalem_tut = st.number_input("$", min_value=0.0, value=float(kalem.get("tutar",0)), step=10.0, format="%.0f", key=f"be_kal_tut_{O}sen_id}_{O}ki}")
                    yeni_kalemler.append({"ad": kalem_ad, "tutar": kalem_tut})
                if st.button("+ Kalem Ekle", key=f"be_kalem_ekle_{O}sen_id}"):
                    yeni_kalemler.append({"ad": "Yeni Kalem", "tutar": 0.0})
                kalemler_be = yeni_kalemler
                toplam_gider_be = sum(k.get("tutar", 0) for k in kalemler_be)

            # Gider periyot dönüşümü
            gider_donem = toplam_gider_be * carpan
            st.markdown(f'<div class="hm-label">Dönem Gideri ({O}periyot})</div><div style="color:#F59E0B;font-size:16px;font-weight:700">{gider_donem:,.2f}</div>', unsafe_allow_html=True)

            st.markdown('<div style="height:8px"></div>', unsafe_allow_html=True)

            # Ortalama marj
            ort_marj = st.number_input("Ortalama Marj (%)", min_value=0.1, max_value=99.9, value=float(sen.get("ortalama_marj",30)), step=0.1, format="%.1f", key=f"be_marj_{O}sen_id}")

            # Mevcut ciro
            mevcut_ciro = st.number_input(f"Mevcut {O}periyot} Cirom ($)", min_value=0.0, value=float(sen.get("mevcut_ciro",0)), step=100.0, format="%.2f", key=f"be_ciro_{O}sen_id}")

            # Ortalama ürün fiyatı (adet hesabı için)
            ort_urun = st.number_input("Ort. Ürün Fiyatı ($)", min_value=0.0, value=float(sen.get("ort_urun_fiyati",0)), step=10.0, format="%.2f", key=f"be_urun_{O}sen_id}", help="Kaç adet satmalısın hesabı için")

            # HESAPLAMA
            if ort_marj > 0 and gider_donem > 0:
                hedef_ciro = gider_donem / (ort_marj / 100)
                kalan = max(0, hedef_ciro - mevcut_ciro)
                ilerleme = min(1.0, mevcut_ciro / hedef_ciro) if hedef_ciro > 0 else 0

                st.markdown('<div class="hm-separator"></div>', unsafe_allow_html=True)
                st.markdown('<div style="color:#FFFFFF;font-size:12px;font-weight:700;margin-bottom:8px">🎯 Kırılma Noktası</div>', unsafe_allow_html=True)
                st.markdown(f'<div class="hm-label">Hedef Ciro ({O}periyot})</div><div style="color:#A5B4FC;font-size:18px;font-weight:800">{hedef_ciro:,.0f}</div>', unsafe_allow_html=True)

                if mevcut_ciro > 0:
                    durum_renk = "#10B981" if mevcut_ciro >= hedef_ciro else "#EF4444"
                    durum_ikon = "✅" if mevcut_ciro >= hedef_ciro else "⚠️"
                    st.markdown(f'<div class="hm-label" style="margin-top:8px">Kalan</div><div style="color:{O}durum_renk};font-size:16px;font-weight:700">{O}durum_ikon} {kalan:,.0f}</div>', unsafe_allow_html=True)

                    # Progress bar
                    st.markdown(f'<div style="margin:10px 0 4px;height:8px;background:rgba(255,255,255,0.08);border-radius:4px;overflow:hidden"><div style="height:100%;width:{O}ilerleme*100:.1f}%;background:linear-gradient(90deg,#6366F1,#10B981);border-radius:4px;transition:width 0.5s"></div></div>', unsafe_allow_html=True)
                    st.markdown(f'<div style="color:#94A3B8;font-size:10px;text-align:right">%{O}ilerleme*100:.0f} tamamlandı</div>', unsafe_allow_html=True)

                if ort_urun > 0:
                    hedef_adet = hedef_ciro / ort_urun
                    kalan_adet = max(0, kalan / ort_urun)
                    st.markdown(f'<div class="hm-label" style="margin-top:8px">Hedef Adet</div><div style="color:#C4B5FD;font-size:16px;font-weight:700">{O}hedef_adet:,.0f} adet</div>', unsafe_allow_html=True)
                    if mevcut_ciro > 0:
                        st.markdown(f'<div style="color:#94A3B8;font-size:11px">Kalan: {O}kalan_adet:,.0f} adet</div>', unsafe_allow_html=True)
            else:
                st.markdown('<div style="color:#475569;font-size:12px;margin-top:8px">Gider ve marj girerek kırılma noktasını hesaplayın.</div>', unsafe_allow_html=True)

            # Senaryo state kaydet
            st.session_state.be_senaryolar[aktif_idx] = {
                "id": sen_id,
                "ad": sen_ad_be,
                "gider_modu": gider_modu,
                "toplam_gider": toplam_gider_be,
                "kalemler": kalemler_be,
                "ortalama_marj": ort_marj,
                "mevcut_ciro": mevcut_ciro,
                "ort_urun_fiyati": ort_urun,
                "periyot": periyot
            }

        # Senaryo ekle/sil
        bcol1, bcol2 = st.columns(2)
        with bcol1:
            if st.button("+ Senaryo", key="be_senaryo_ekle", use_container_width=True):
                st.session_state.be_sayac += 1
                yid = st.session_state.be_sayac
                st.session_state.be_senaryolar.append({
                    "id": yid, "ad": f"Senaryo {O}yid}",
                    "gider_modu": "Toplam", "toplam_gider": 0.0, "kalemler": [],
                    "ortalama_marj": 30.0, "mevcut_ciro": 0.0, "ort_urun_fiyati": 0.0, "periyot": "Aylık"
                })
                st.session_state.be_aktif_idx = len(st.session_state.be_senaryolar) - 1
                st.rerun()
        with bcol2:
            if len(st.session_state.be_senaryolar) > 1:
                if st.button("🗑️ Sil", key="be_senaryo_sil", use_container_width=True):
                    st.session_state.be_senaryolar.pop(aktif_idx)
                    st.session_state.be_aktif_idx = max(0, aktif_idx - 1)
                    st.rerun()

# ─────────────────────────────────────────────────
# ANA RUN FONKSİYONU
# ─────────────────────────────────────────────────
def run():
    aktif_kullanici = st.session_state.get("aktif_kullanici", "").lower().strip()

    st.markdown(_hm_css(), unsafe_allow_html=True)

    # Başlık
    st.markdown(
        '<div style="margin-bottom:28px">'
        '<div style="display:inline-block;padding:6px 14px;background:rgba(245,158,11,0.12);border:1px solid rgba(245,158,11,0.3);border-radius:20px;margin-bottom:16px">'
        '<span style="color:#FCD34D;font-size:11px;font-weight:600;letter-spacing:1px;text-transform:uppercase">🧮 Hesap Makinesi</span>'
        '</div>'
        '<h1 style="font-family:Inter,sans-serif;font-size:clamp(22px,4vw,36px);font-weight:800;color:#FFFFFF;margin:0">Hesap Makinesi</h1>'
        '<p style="color:#94A3B8;font-size:13px;margin-top:6px">Ürün karlılık analizi ve kırılma noktası hesaplama aracı</p>'
        '</div>',
        unsafe_allow_html=True
    )

    # Break-even sidebar'ı çiz (her zaman görünür)
    _breakeven_sidebar()

    # Ana içerik: Ürün karlılık hesaplayıcı
    _urun_karlilik_bolumu()

    # Geçmiş hesaplamalar
    _gecmis_hesaplamalar()

